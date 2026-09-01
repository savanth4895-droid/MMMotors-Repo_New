"""
MM Motors — Backend (FastAPI + MongoDB Atlas)

Structure:
  database.py — Shared config, DB helpers, auth, GST utils, serialization
  server.py   — App setup, Pydantic models, all route handlers

Render env vars required:
  MONGO_URL       — MongoDB Atlas connection string
  DB_NAME         — Database name (default: mmmotors)
  JWT_SECRET_KEY  — Random 64-char hex (generate: python3 -c "import secrets; print(secrets.token_hex(32))")
  ALLOW_ORIGINS   — Your Vercel URL (e.g. https://yourapp.vercel.app)
"""

import asyncio
import traceback
import re
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from typing import Optional, List, Any, Dict

from fastapi import FastAPI, Depends, HTTPException, status, Query, Request, UploadFile, File, Form, Response, Cookie
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field
import os
import certifi
import io

# PDF generation
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ── Register fonts once at module level ──────────────────────────────────────
_FONTS_REGISTERED = False
def _register_fonts():
    global _FONTS_REGISTERED
    if _FONTS_REGISTERED:
        return
    try:
        FONT_DIR = "/usr/share/fonts/truetype/liberation/"
        DJVU_DIR = "/usr/share/fonts/truetype/dejavu/"
        pdfmetrics.registerFont(TTFont('Sans',       FONT_DIR + 'LiberationSans-Regular.ttf'))
        pdfmetrics.registerFont(TTFont('Sans-Bold',  FONT_DIR + 'LiberationSans-Bold.ttf'))
        pdfmetrics.registerFont(TTFont('Sans-Italic',FONT_DIR + 'LiberationSans-Italic.ttf'))
        pdfmetrics.registerFont(TTFont('Mono',       DJVU_DIR + 'DejaVuSansMono.ttf'))
        pdfmetrics.registerFont(TTFont('Mono-Bold',  DJVU_DIR + 'DejaVuSansMono-Bold.ttf'))
        _FONTS_REGISTERED = True
    except Exception as e:
        print(f"[MM Motors] Font registration warning: {e}")

# ── Shared database module ────────────────────────────────────────────────────
import database as _db
from database import (
    MONGO_URL, DB_NAME, JWT_SECRET_KEY, JWT_ALGORITHM, JWT_EXPIRE_MIN,
    MAX_LOGIN_ATTEMPTS, LOGIN_LOCKOUT_MIN, BRANDS, GST_RATES,
    pwd_ctx, create_token, verify_token, require_admin, require_roles,
    next_sequence, _sync_counter,
    oid, oids, obj_id, paginate_params, now, utcnow,
    calc_gst_line, calc_bill_totals, amount_in_words,
)
from motor.motor_asyncio import AsyncIOMotorClient, AsyncIOMotorGridFSBucket

# Module-level db/fs aliases — set in lifespan
db = None
fs = None

@asynccontextmanager
async def lifespan(app):
    global db, fs
    try:
        _db.client = AsyncIOMotorClient(
            MONGO_URL,
            serverSelectionTimeoutMS=10000,
            tls=True,
            tlsCAFile=certifi.where(),
        )
        _db.db = _db.client[DB_NAME]
        _db.fs = AsyncIOMotorGridFSBucket(_db.db)
        db = _db.db
        fs = _db.fs
        await _ensure_indexes()
        await _seed_owner()
        await _seed_badge_types()
        print(f"[MM Motors] Connected to MongoDB · DB: {DB_NAME}")
    except Exception as e:
        print(f"[MM Motors] WARNING: DB connection failed: {e}")

    yield

    if _db.client:
        _db.client.close()

# ─── App ──────────────────────────────────────────────────────────────────────
_raw_origins = os.getenv("ALLOW_ORIGINS", "").strip()
ALLOW_ORIGINS = [o.strip() for o in _raw_origins.split(",") if o.strip()] or ["*"]

app = FastAPI(title="MM Motors API", version="2.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOW_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Total-Count"],
)

# Cache invalidation — clears /service/due cache on mutations to sales/service_jobs
@app.middleware("http")
async def _service_due_cache_invalidator(request: Request, call_next):
    response = await call_next(request)
    if request.method in ("POST", "PUT", "PATCH", "DELETE") and 200 <= response.status_code < 300:
        path = request.url.path
        if "/sales" in path or "/service" in path:
            _invalidate_service_due_cache()
    return response

from fastapi import APIRouter
api_router = APIRouter(prefix="/api")

async def _ensure_indexes():
    # users
    await db.users.create_index("username", unique=True)
    await db.users.create_index("mobile")
    # customers
    await db.customers.create_index("mobile")
    await db.customers.create_index([("name","text"),("mobile","text")])
    # vehicles
    await db.vehicles.create_index("chassis_number", unique=True, sparse=True)
    await db.vehicles.create_index("vehicle_number")
    await db.vehicles.create_index("status")
    await db.vehicles.create_index("brand")
    # sales
    await db.sales.create_index("invoice_number", unique=True)
    await db.sales.create_index("customer_id")
    await db.sales.create_index("sale_date")
    # service_jobs
    await db.service_jobs.create_index("job_number", unique=True)
    await db.service_jobs.create_index("status")
    await db.service_jobs.create_index("customer_id")
    await db.service_jobs.create_index("vehicle_number")
    # service_bills
    await db.service_bills.create_index("bill_number", unique=True)
    await db.service_bills.create_index("job_id")
    # spare_parts
    await db.spare_parts.create_index("part_number", unique=True)
    await db.spare_parts.create_index("category")
    await db.spare_parts.create_index([("name","text"),("part_number","text")])
    # parts_sales
    await db.parts_sales.create_index("bill_number", unique=True)
    await db.parts_sales.create_index("sale_date")
    # parts_bills  ← PATCH 7: new collection index
    await db.parts_bills.create_index("bill_number")
    await db.parts_bills.create_index("customer_mobile")
    await db.parts_bills.create_index("bill_date")
    # vendors — spare-parts suppliers
    await db.vendors.create_index("name")
    await db.vendors.create_index("gstin", sparse=True)
    # purchase_bills — one doc per vendor invoice; multi-line
    await db.purchase_bills.create_index([("vendor_id", 1), ("bill_number", 1)], unique=True, sparse=True)
    await db.purchase_bills.create_index("bill_date")
    await db.purchase_bills.create_index("vendor_id")
    # spare_parts additions: index the alias list for vendor-scoped search
    await db.spare_parts.create_index("aliases.vendor_id")
    await db.spare_parts.create_index("aliases.part_number")
    # login_attempts — TTL 30 min
    await db.login_attempts.create_index(
        "created_at", expireAfterSeconds=LOGIN_LOCKOUT_MIN * 60
    )
    await db.login_attempts.create_index("username")
    await db.login_attempts.create_index([("username", 1), ("ip", 1)])
    await db.badge_types.create_index("name", unique=True)
    print("[MM Motors] Indexes ensured")

async def _seed_owner():
    """Always ensure the default owner account exists (upsert — never overwrites password if already set)."""
    existing = await db.users.find_one({"username": "owner"})
    if not existing:
        await db.users.insert_one({
            "username":   "owner",
            "name":       "Owner",
            "mobile":     "",
            "email":      "",
            "role":       "owner",
            "password":   pwd_ctx.hash("mm@123456"),
            "status":     "active",
            "salary":     0,
            "join_date":  utcnow().strftime("%d %b %Y"),
            "created_at": utcnow().isoformat(),
        })
        print("[MM Motors] Default owner created  username=owner  password=mm@123456")
    else:
        # Make sure the existing owner account is active
        await db.users.update_one(
            {"username": "owner"},
            {"$set": {"status": "active", "role": "owner"}}
        )
        print(f"[MM Motors] Owner account verified  status=active")

async def _seed_badge_types():
    """Seed default badge types once. Owner can add/edit more from Settings later."""
    if await db.badge_types.count_documents({}) > 0:
        return
    defaults = [
        {"name": "VIP",           "color": "#B8860B", "sort_order":  1},
        {"name": "Family",        "color": "#9333EA", "sort_order":  2},
        {"name": "Repeat",        "color": "#16A34A", "sort_order":  3},
        {"name": "Referral",      "color": "#0EA5E9", "sort_order":  4},
        {"name": "Special Offer", "color": "#EF4444", "sort_order":  5},
        {"name": "Corporate",     "color": "#3B82F6", "sort_order":  6},
    ]
    now = utcnow().isoformat()
    for b in defaults:
        b["created_at"] = now
    await db.badge_types.insert_many(defaults)
    print(f"[MM Motors] Seeded {len(defaults)} default badge types")




# ── Auth ──────────────────────────────────────────────────────────────────────
class LoginIn(BaseModel):
    username: str
    password: str

class TokenOut(BaseModel):
    access_token: str
    token_type:   str = "bearer"
    user:         dict

# ── Users / Staff ──────────────────────────────────────────────────────────────
class UserCreate(BaseModel):
    username:      str
    name:          str
    mobile:        str
    email:         Optional[str]       = ""
    role:          str                 = "sales"
    password:      str
    salary:        Optional[float]     = Field(0, ge=0)
    join_date:     Optional[str]       = ""
    status:        Optional[str]       = "active"
    allowed_pages: Optional[list[str]] = None  # None = use role defaults

class UserUpdate(BaseModel):
    name:          Optional[str]       = None
    mobile:        Optional[str]       = None
    email:         Optional[str]       = None
    role:          Optional[str]       = None
    salary:        Optional[float]     = Field(None, ge=0)
    status:        Optional[str]       = None
    join_date:     Optional[str]       = None
    allowed_pages: Optional[list[str]] = None

class PasswordChange(BaseModel):
    new_password: str

# ── Customers ─────────────────────────────────────────────────────────────────
class CustomerCreate(BaseModel):
    name:    str
    mobile:  str
    email:   Optional[str] = ""
    address: Optional[str] = ""
    gstin:   Optional[str] = ""
    tags:    Optional[List[str]] = []
    id_proof_file_id: Optional[str] = ""

class CustomerUpdate(BaseModel):
    name:    Optional[str] = None
    mobile:  Optional[str] = None
    email:   Optional[str] = None
    address: Optional[str] = None
    gstin:   Optional[str] = None
    # tags are managed via the dedicated admin-only PUT /customers/{id}/tags endpoint
    id_proof_file_id: Optional[str] = None

class CustomerTagsUpdate(BaseModel):
    tags: List[str]

# ── Badge Types (owner-managed customer badges) ────────────────────────────────
class BadgeTypeCreate(BaseModel):
    name:       str    = Field(..., min_length=1, max_length=40)
    color:      str    = Field("#888888", min_length=4, max_length=9)   # hex like #RRGGBB
    sort_order: Optional[int] = Field(100, ge=0)

class BadgeTypeUpdate(BaseModel):
    name:       Optional[str] = Field(None, min_length=1, max_length=40)
    color:      Optional[str] = Field(None, min_length=4, max_length=9)
    sort_order: Optional[int] = Field(None, ge=0)

# ── Vehicles ──────────────────────────────────────────────────────────────────
class VehicleCreate(BaseModel):
    brand:          str
    model:          str
    variant:        Optional[str] = ""
    color:          Optional[str] = ""
    chassis_number: Optional[str] = ""
    engine_number:  Optional[str] = ""
    purchase_price: Optional[float] = 0
    inbound_date:      Optional[str] = ""
    location:          Optional[str] = ""
    outbound_date:     Optional[str] = ""
    outbound_location: Optional[str] = ""
    status:            Optional[str] = "Instock"
    type:           Optional[str] = "new"
    notes:          Optional[str] = ""

class VehicleUpdate(BaseModel):
    brand:          Optional[str] = None
    model:          Optional[str] = None
    variant:        Optional[str] = None
    color:          Optional[str] = None
    chassis_number: Optional[str] = None
    engine_number:  Optional[str] = None
    purchase_price: Optional[float] = None
    inbound_date:      Optional[str] = None
    location:          Optional[str] = None
    outbound_date:     Optional[str] = None
    outbound_location: Optional[str] = None
    key_number:     Optional[str] = None
    type:           Optional[str] = None
    status:            Optional[str] = None
    notes:          Optional[str] = None

# ── Sales ─────────────────────────────────────────────────────────────────────
class NomineeInfo(BaseModel):
    name:     Optional[str] = ""
    relation: Optional[str] = ""
    age:      Optional[str] = ""
    number:   Optional[str] = ""

class SaleCreate(BaseModel):
    customer_id:       str
    vehicle_id:        str
    vehicle_number:    Optional[str]   = ""
    vehicle_brand:     Optional[str]   = None   # override DB value if provided
    vehicle_model:     Optional[str]   = None
    vehicle_variant:   Optional[str]   = None
    vehicle_color:     Optional[str]   = None
    chassis_number:    Optional[str]   = None
    engine_number:     Optional[str]   = None
    sale_price:        Optional[float] = 0
    total_amount:      Optional[float] = None
    finance_type:      Optional[str]   = "cash"
    financier:         Optional[str]   = ""
    loan_amount:       Optional[float] = 0
    nominee:           Optional[NomineeInfo] = None
    payment_mode:      Optional[str]   = "Cash"
    sold_by:           Optional[str]   = ""
    sale_date:         Optional[str]   = ""
    notes:             Optional[str]   = ""
    care_of:           Optional[str]   = ""
    hsrp_front:        Optional[str]   = ""
    hsrp_back:         Optional[str]   = ""
    hsrp_front_id:     Optional[str]   = ""
    hsrp_back_id:      Optional[str]   = ""
    hsrp_date:         Optional[str]   = ""
    hsrp_notes:        Optional[str]   = ""
    milestones:        Optional[Dict[str, bool]] = None
    milestone_dates:   Optional[Dict[str, str]]  = None

class SaleUpdate(BaseModel):
    status:            Optional[str]   = None
    delivery_date:     Optional[str]   = None
    vehicle_number:    Optional[str]   = None
    payment_mode:      Optional[str]   = None
    notes:             Optional[str]   = None
    customer_name:     Optional[str]   = None
    customer_mobile:   Optional[str]   = None
    customer_address:  Optional[str]   = None
    vehicle_id:        Optional[str]   = None
    total_amount:      Optional[float] = None
    sale_price:        Optional[float] = None
    finance_type:      Optional[str]   = None
    financier:         Optional[str]   = None
    loan_amount:       Optional[float] = None
    nominee:           Optional[NomineeInfo] = None
    sale_date:         Optional[str]   = None
    sold_by:           Optional[str]   = None
    care_of:           Optional[str]   = None
    hsrp_front:        Optional[str]   = None
    hsrp_back:         Optional[str]   = None
    hsrp_front_id:     Optional[str]   = None
    hsrp_back_id:      Optional[str]   = None
    hsrp_date:         Optional[str]   = None
    hsrp_notes:        Optional[str]   = None
    milestones:        Optional[Dict[str, bool]] = None
    milestone_dates:   Optional[Dict[str, str]]  = None


# ─── Sale Milestones ───────────────────────────────────────────────────────
SALE_MILESTONE_KEYS = ["documents", "invoice", "insurance", "tax_paid", "number_plate"]

def default_milestones() -> Dict[str, bool]:
    return {k: False for k in SALE_MILESTONE_KEYS}


class MilestoneUpdate(BaseModel):
    key:   str
    value: bool
    date:  Optional[str] = None   # YYYY-MM-DD; required-ish when value=True (defaults to today)

# ── Service Jobs ──────────────────────────────────────────────────────────────
class ServiceJobCreate(BaseModel):
    customer_id:    str
    vehicle_number: str
    brand:          str
    model:          str
    variant:        Optional[str] = ""
    chassis_number: Optional[str] = ""
    odometer_km:    Optional[int] = 0
    complaint:      str
    advisor_id:     Optional[str] = ""
    technician:     Optional[str] = ""
    check_in_date:  Optional[str] = ""
    estimated_delivery: Optional[str] = ""
    notes:          Optional[str] = ""
    vehicle_photo_id: Optional[str] = ""
    save_vehicle:   Optional[bool] = False

class ServiceJobUpdate(BaseModel):
    status:             Optional[str] = None
    technician:         Optional[str] = None
    estimated_delivery: Optional[str] = None
    delivery_date:      Optional[str] = None
    notes:              Optional[str] = None
    complaint:          Optional[str] = None
    vehicle_number:     Optional[str] = None
    chassis_number:     Optional[str] = None
    brand:              Optional[str] = None
    model:              Optional[str] = None
    odometer_km:        Optional[int] = None
    odometer_out:       Optional[int] = None
    vehicle_photo_id:   Optional[str] = None
    check_in_date:      Optional[str] = None
    created_at:         Optional[str] = None   # correcting imported service dates
    customer_name:      Optional[str] = None
    customer_mobile:    Optional[str] = None

# ── Service Bills ─────────────────────────────────────────────────────────────
class BillLineItem(BaseModel):
    description:   str
    part_number:   Optional[str] = ""
    hsn_code:      Optional[str] = ""
    qty:           int           = Field(1, ge=1)
    unit_price:    float         = Field(..., ge=0)
    gst_rate:      float         = Field(18, ge=0, le=100)
    discount:      Optional[float] = Field(0, ge=0)
    complimentary: Optional[bool] = False

class ServiceBillCreate(BaseModel):
    job_id:          str
    labour_charges:  Optional[float] = Field(0, ge=0)
    labour_gst_rate: Optional[float] = Field(18, ge=0, le=100)
    labour_discount: Optional[float] = Field(0, ge=0)
    items:           Optional[List[BillLineItem]] = []
    payment_mode:    Optional[str] = "Cash"
    notes:           Optional[str] = ""
    discount:        Optional[float] = Field(0, ge=0)

class ServiceBillUpdate(BaseModel):
    items:           Optional[List[BillLineItem]] = None
    payment_mode:    Optional[str]   = None
    notes:           Optional[str]   = None
    discount:        Optional[float] = Field(None, ge=0)

# ── Spare Parts ───────────────────────────────────────────────────────────────
class SparePartCreate(BaseModel):
    part_number:     Optional[str] = ""
    name:            str
    category:        Optional[str] = ""
    brand:           Optional[str] = ""
    compatible_with: Optional[List[str]] = []
    stock:           int    = Field(0, ge=0)
    reorder_level:   int    = Field(5, ge=0)
    purchase_price:  float  = Field(..., ge=0)
    selling_price:   float  = Field(..., ge=0)
    gst_rate:        float  = Field(18, ge=0, le=100)
    hsn_code:        Optional[str] = ""
    location:        Optional[str] = ""

class SparePartUpdate(BaseModel):
    name:            Optional[str] = None
    category:        Optional[str] = None
    brand:           Optional[str] = None
    compatible_with: Optional[List[str]] = None
    stock:           Optional[int]   = Field(None, ge=0)
    reorder_level:   Optional[int]   = Field(None, ge=0)
    purchase_price:  Optional[float] = Field(None, ge=0)
    selling_price:   Optional[float] = Field(None, ge=0)
    gst_rate:        Optional[float] = Field(None, ge=0, le=100)
    hsn_code:        Optional[str] = None
    location:        Optional[str] = None

# PATCH 4: Added action field ─────────────────────────────────────────────────
class StockAdjust(BaseModel):
    qty:    int    # positive = stock in, negative = adjustment
    action: Optional[str] = "add"   # "add" | "subtract"
    reason: Optional[str] = ""


# ── Vendors ──────────────────────────────────────────────────────────────────
class VendorCreate(BaseModel):
    name:         str
    gstin:        Optional[str] = ""
    address:      Optional[str] = ""
    phone:        Optional[str] = ""
    email:        Optional[str] = ""
    bank_name:    Optional[str] = ""
    bank_account: Optional[str] = ""
    bank_ifsc:    Optional[str] = ""
    notes:        Optional[str] = ""

class VendorUpdate(BaseModel):
    name:         Optional[str] = None
    gstin:        Optional[str] = None
    address:      Optional[str] = None
    phone:        Optional[str] = None
    email:        Optional[str] = None
    bank_name:    Optional[str] = None
    bank_account: Optional[str] = None
    bank_ifsc:    Optional[str] = None
    notes:        Optional[str] = None


# ── Purchase Bills (from vendors) ────────────────────────────────────────────
class PurchaseBillItemIn(BaseModel):
    part_id:        Optional[str] = None   # null = create new spare_part
    part_name:      str
    part_number:    Optional[str] = ""     # vendor's SKU
    hsn:            Optional[str] = ""
    qty:            float           = Field(..., gt=0)
    unit:           Optional[str] = "PCS"
    rate:           float           = Field(..., ge=0)   # per-unit rate before discount
    discount_pct:   Optional[float] = Field(0, ge=0, le=100)
    gst_pct:        Optional[float] = Field(18, ge=0, le=100)
    # Optional per-line sale price update, and new-part hints
    new_sale_price: Optional[float] = Field(None, ge=0)
    new_part_selling_price: Optional[float] = Field(None, ge=0)  # only used if creating new part

class PurchaseBillCreate(BaseModel):
    vendor_id:        str
    bill_number:      str
    bill_date:        str                  # DD Mon YYYY preferred
    place_of_supply:  Optional[str] = ""
    items:            List[PurchaseBillItemIn]
    round_off:        Optional[float] = 0
    notes:            Optional[str] = ""

class PurchaseBillUpdate(BaseModel):
    bill_number:      Optional[str] = None
    bill_date:        Optional[str] = None
    place_of_supply:  Optional[str] = None
    items:            Optional[List[PurchaseBillItemIn]] = None
    round_off:        Optional[float] = None
    notes:            Optional[str] = None

# ── Parts Sales ───────────────────────────────────────────────────────────────
class PartsSaleItem(BaseModel):
    part_id:     str
    part_number: str
    name:        str
    hsn_code:    Optional[str] = ""
    qty:         int    = Field(..., ge=1)
    unit_price:  float  = Field(..., ge=0)
    gst_rate:    float  = Field(18, ge=0, le=100)

class PartsSaleCreate(BaseModel):
    customer_name:  Optional[str] = ""
    customer_mobile:Optional[str] = ""
    items:          List[PartsSaleItem]
    payment_mode:   Optional[str] = "Cash"
    sold_by:        Optional[str] = ""
    notes:          Optional[str] = ""

# PATCH 3: Parts Bills models ─────────────────────────────────────────────────
class PartsBillItem(BaseModel):
    part_id:         Optional[str]   = None
    part_number:     Optional[str]   = ""
    name:            str
    hsn_code:        Optional[str]   = "8714"
    qty:             int             = Field(1, ge=1)
    unit_price:      float           = Field(..., ge=0)
    gst_rate:        float           = Field(18.0, ge=0, le=100)
    discount:        Optional[float] = Field(0, ge=0)
    complimentary:   Optional[bool]  = False

class PartsBillCreate(BaseModel):
    customer_name:    Optional[str] = ""
    customer_mobile:  Optional[str] = ""
    customer_vehicle: Optional[str] = ""
    payment_mode:     Optional[str] = "Cash"
    items:            List[PartsBillItem] = []
    discount:         Optional[float] = Field(0, ge=0)

class PartsBillUpdate(BaseModel):
    customer_name:    Optional[str] = None
    customer_mobile:  Optional[str] = None
    customer_vehicle: Optional[str] = None
    payment_mode:     Optional[str] = None
    items:            Optional[List[PartsBillItem]] = None
    discount:         Optional[float] = Field(None, ge=0)


# ─── Accident Estimates ───────────────────────────────────────────────────────
class AccidentEstimatePartItem(BaseModel):
    part_name:  str   = ""
    part_number: str  = ""
    condition:  str   = "New OEM"
    qty:        float = Field(1, ge=0)
    unit_price: float = Field(0, ge=0)
    gst:        float = Field(18, ge=0, le=100)

class AccidentEstimateLabourItem(BaseModel):
    description: str   = ""
    hours:       float = Field(1, ge=0)
    rate:        float = Field(0, ge=0)

class AccidentEstimateCreate(BaseModel):
    vehicle:     dict = {}
    customer:    dict = {}
    incident:    dict = {}
    parts:       List[AccidentEstimatePartItem]   = []
    labour:      List[AccidentEstimateLabourItem] = []
    additional:  dict = {}
    notes:       str  = ""
    grand_total: float = 0
    status:      str  = "draft"

class AccidentEstimateUpdate(BaseModel):
    vehicle:     Optional[dict]  = None
    customer:    Optional[dict]  = None
    incident:    Optional[dict]  = None
    parts:       Optional[List[AccidentEstimatePartItem]]   = None
    labour:      Optional[List[AccidentEstimateLabourItem]] = None
    additional:  Optional[dict]  = None
    notes:       Optional[str]   = None
    grand_total: Optional[float] = None
    status:      Optional[str]   = None


# ═══════════════════════════════════════════════════════════════════════════════
#  HEALTH
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/health")
async def health():
    return {"status": "ok", "service": "MM Motors API", "version": "2.0.0"}

@app.get("/ready")
async def ready():
    try:
        await db.command("ping")
        return {"status": "ready", "db": DB_NAME}
    except Exception as e:
        return JSONResponse(status_code=503, content={"status": "not_ready", "error": str(e)})


# ═══════════════════════════════════════════════════════════════════════════════
#  FILES (GridFS Uploads)
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), current_user=Depends(verify_token)):
    file_id = await fs.upload_from_stream(
        file.filename,
        file.file,
        metadata={"content_type": file.content_type, "uploaded_by": current_user["username"]}
    )
    return {"file_id": str(file_id), "filename": file.filename}

@api_router.get("/files/{file_id}")
async def get_file(file_id: str, current_user=Depends(verify_token)):
    try:
        grid_out = await fs.open_download_stream(obj_id(file_id))
        async def file_stream():
            while chunk := await grid_out.readchunk():
                yield chunk
        return StreamingResponse(
            file_stream(),
            media_type=grid_out.metadata.get("content_type", "application/octet-stream")
        )
    except Exception:
        raise HTTPException(status_code=404, detail="File not found")


# ═══════════════════════════════════════════════════════════════════════════════
#  AUTH
# ═══════════════════════════════════════════════════════════════════════════════

def _client_ip(request: Request) -> str:
    """Extract client IP behind Render/Vercel proxy chain."""
    xff = request.headers.get("x-forwarded-for", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


@api_router.post("/auth/login", response_model=TokenOut)
async def login(body: LoginIn, request: Request):
    username = body.username.strip().lower()
    ip       = _client_ip(request)
    # Per-(username, IP) lockout — attacker cannot lock real user from elsewhere
    attempt_count = await db.login_attempts.count_documents({"username": username, "ip": ip})
    if attempt_count >= MAX_LOGIN_ATTEMPTS:
        raise HTTPException(
            status_code=429,
            detail=f"Too many failed attempts from this location. Try again in {LOGIN_LOCKOUT_MIN} minutes."
        )
    user = await db.users.find_one({"username": username})
    if not user or not pwd_ctx.verify(body.password, user.get("password", "")):
        await db.login_attempts.insert_one({
            "username":   username,
            "ip":         ip,
            "created_at": utcnow(),
        })
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if user.get("status") != "active":
        raise HTTPException(status_code=403, detail="Account is inactive or deactivated")
    # Clear only this IP's failures for this user — preserve others as evidence
    await db.login_attempts.delete_many({"username": username, "ip": ip})
    token = create_token({"sub": str(user["_id"]), "role": user["role"]})
    user_data = {
        "id":       str(user["_id"]),
        "username": user["username"],
        "name":     user["name"],
        "role":     user["role"],
        "mobile":   user.get("mobile", ""),
        "allowed_pages": user.get("allowed_pages", None),
    }
    response = JSONResponse(content={"access_token": token, "token_type": "bearer", "user": user_data})
    response.set_cookie(
        key="mm_token", value=token, httponly=True, secure=True,
        samesite="none", max_age=60 * JWT_EXPIRE_MIN, path="/",
    )
    return response

@api_router.get("/auth/me")
async def me(current_user: dict = Depends(verify_token)):
    user = dict(current_user)
    user.pop("password", None)
    return user

@api_router.post("/auth/logout")
async def logout(current_user: dict = Depends(verify_token)):
    response = Response(content='{"message":"Logged out"}', media_type="application/json")
    response.delete_cookie(key="mm_token", path="/")
    return response


# ═══════════════════════════════════════════════════════════════════════════════
#  USERS / STAFF
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/users")
async def list_users(p=Depends(paginate_params), current_user=Depends(require_admin)):
    docs  = await db.users.find({}, {"password": 0}).skip(p["skip"]).limit(p["limit"]).to_list(p["limit"])
    total = await db.users.count_documents({})
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})

@api_router.post("/users", status_code=201)
async def create_user(body: UserCreate, current_user=Depends(require_admin)):
    existing = await db.users.find_one({"username": body.username.strip().lower()})
    if existing:
        raise HTTPException(status_code=409, detail="Username already exists")
    doc = body.dict()
    doc["username"] = doc["username"].strip().lower()
    doc["password"] = pwd_ctx.hash(doc["password"])
    doc["created_at"] = utcnow()
    result = await db.users.insert_one(doc)
    doc["id"] = str(result.inserted_id)
    doc.pop("_id", None); doc.pop("password", None)
    return doc

@api_router.get("/users/{user_id}")
async def get_user(user_id: str, current_user=Depends(require_admin)):
    doc = await db.users.find_one({"_id": obj_id(user_id)}, {"password": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="User not found")
    return oid(doc)

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, body: UserUpdate, current_user=Depends(require_admin)):
    update = {k: v for k, v in body.dict().items() if v is not None}
    # allowed_pages can be an empty list — treat None as "not provided", [] as "no pages"
    if body.allowed_pages is not None:
        update["allowed_pages"] = body.allowed_pages
    if not update:
        raise HTTPException(status_code=400, detail="Nothing to update")
    await db.users.update_one({"_id": obj_id(user_id)}, {"$set": update})
    return oid(await db.users.find_one({"_id": obj_id(user_id)}, {"password": 0}))

@api_router.post("/users/{user_id}/password")
async def change_password(user_id: str, body: PasswordChange, current_user=Depends(require_admin)):
    await db.users.update_one({"_id": obj_id(user_id)}, {"$set": {"password": pwd_ctx.hash(body.new_password)}})
    return {"message": "Password updated"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user=Depends(require_admin)):
    if str(current_user["_id"]) == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    result = await db.users.delete_one({"_id": obj_id(user_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "Deleted"}


# ═══════════════════════════════════════════════════════════════════════════════
#  CUSTOMERS
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/customers")
async def list_customers(
    search: Optional[str] = Query(None),
    tag:    Optional[str] = Query(None),
    limit:  int           = Query(200, ge=1, le=2000),
    current_user=Depends(verify_token),
):
    query: dict = {}
    if search:
        s = re.escape(search)
        query["$or"] = [
            {"name":   {"$regex": s, "$options": "i"}},
            {"mobile": {"$regex": s, "$options": "i"}},
            {"email":  {"$regex": s, "$options": "i"}},
        ]
    if tag:
        query["tags"] = tag
    docs  = await db.customers.find(query).sort("name", 1).limit(limit).to_list(limit)
    total = await db.customers.count_documents(query)
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})

@api_router.post("/customers", status_code=201)
async def create_customer(body: CustomerCreate, current_user=Depends(verify_token)):
    doc = body.dict()
    doc["created_at"] = utcnow().isoformat()
    result = await db.customers.insert_one(doc)
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)
    return doc

@api_router.get("/customers/{cust_id}")
async def get_customer(cust_id: str, current_user=Depends(verify_token)):
    doc = await db.customers.find_one({"_id": obj_id(cust_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Customer not found")
    return oid(doc)

@api_router.put("/customers/{cust_id}")
async def update_customer(cust_id: str, body: CustomerUpdate, current_user=Depends(verify_token)):
    update = {k: v for k, v in body.dict().items() if v is not None}
    await db.customers.update_one({"_id": obj_id(cust_id)}, {"$set": update})
    return oid(await db.customers.find_one({"_id": obj_id(cust_id)}))

@api_router.delete("/customers/{cust_id}")
async def delete_customer(cust_id: str, current_user=Depends(require_admin)):
    result = await db.customers.delete_one({"_id": obj_id(cust_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Deleted"}


@api_router.put("/customers/{cust_id}/tags")
async def set_customer_tags(cust_id: str, body: CustomerTagsUpdate, current_user=Depends(require_admin)):
    """Owner-only. Replace a customer's badge list. Silently drops any name not in badge_types."""
    valid = set()
    async for b in db.badge_types.find({}, {"name": 1}):
        valid.add(b["name"])
    clean = [t for t in (body.tags or []) if t in valid]
    cust = await db.customers.find_one({"_id": obj_id(cust_id)}, {"_id": 1})
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")
    await db.customers.update_one({"_id": obj_id(cust_id)}, {"$set": {"tags": clean}})
    return {"id": cust_id, "tags": clean}


@api_router.get("/customer-badges-map")
async def customer_badges_map(current_user=Depends(verify_token)):
    """Lightweight lookup: mobile → [badge names]. Used by list pages that show a customer name
    but don't already fetch the full customer record. Empty tag lists are omitted."""
    out = {}
    cursor = db.customers.find({"tags": {"$exists": True, "$ne": []}}, {"mobile": 1, "tags": 1})
    async for c in cursor:
        mob = (c.get("mobile") or "").strip()
        tags = c.get("tags") or []
        if mob and tags:
            out[mob] = tags
    return out


# ═══════════════════════════════════════════════════════════════════════════════
#  BADGE TYPES  (owner-managed set of labels applied to customers)
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/badge-types")
async def list_badge_types(current_user=Depends(verify_token)):
    """All roles can read — needed to render chips throughout the app."""
    docs = await db.badge_types.find({}).sort([("sort_order", 1), ("name", 1)]).to_list(None)
    return oids(docs)

@api_router.post("/badge-types")
async def create_badge_type(body: BadgeTypeCreate, current_user=Depends(require_admin)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    if await db.badge_types.find_one({"name": name}):
        raise HTTPException(status_code=409, detail=f"Badge '{name}' already exists")
    doc = {
        "name":       name,
        "color":      body.color,
        "sort_order": body.sort_order if body.sort_order is not None else 100,
        "created_at": utcnow().isoformat(),
    }
    res = await db.badge_types.insert_one(doc)
    doc["id"] = str(res.inserted_id); doc.pop("_id", None)
    return doc

@api_router.patch("/badge-types/{badge_id}")
async def update_badge_type(badge_id: str, body: BadgeTypeUpdate, current_user=Depends(require_admin)):
    existing = await db.badge_types.find_one({"_id": obj_id(badge_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Badge type not found")
    update = {k: v for k, v in body.dict().items() if v is not None}
    if "name" in update:
        update["name"] = update["name"].strip()
        if not update["name"]:
            raise HTTPException(status_code=400, detail="Name cannot be empty")
        clash = await db.badge_types.find_one({"name": update["name"], "_id": {"$ne": obj_id(badge_id)}})
        if clash:
            raise HTTPException(status_code=409, detail=f"Badge '{update['name']}' already exists")
        old_name = existing["name"]
        new_name = update["name"]
        if old_name != new_name:
            # Cascade rename across all customers
            await db.customers.update_many(
                {"tags": old_name},
                {"$set": {"tags.$[elem]": new_name}},
                array_filters=[{"elem": old_name}],
            )
    await db.badge_types.update_one({"_id": obj_id(badge_id)}, {"$set": update})
    return oid(await db.badge_types.find_one({"_id": obj_id(badge_id)}))

@api_router.delete("/badge-types/{badge_id}")
async def delete_badge_type(badge_id: str, current_user=Depends(require_admin)):
    doc = await db.badge_types.find_one({"_id": obj_id(badge_id)}, {"name": 1})
    if not doc:
        raise HTTPException(status_code=404, detail="Badge type not found")
    strip_res = await db.customers.update_many({"tags": doc["name"]}, {"$pull": {"tags": doc["name"]}})
    await db.badge_types.delete_one({"_id": obj_id(badge_id)})
    return {"message": "Deleted", "customers_updated": strip_res.modified_count}



async def customer_timeline(cust_id: str, current_user=Depends(verify_token)):
    sales, jobs = await asyncio.gather(
        db.sales.find({"customer_id": cust_id}).sort("sale_date", -1).to_list(None),
        db.service_jobs.find({"customer_id": cust_id}).sort("check_in_date", -1).to_list(None),
    )
    job_ids = [str(j["_id"]) for j in jobs]
    bills = []
    if job_ids:
        bills = await db.service_bills.find({"job_id": {"$in": job_ids}}).to_list(None)
    service_spend       = sum(b.get("grand_total", 0) for b in bills)
    total_sales_spend   = sum(s.get("total_amount", 0) for s in sales)
    return {
        "sales":             oids(sales),
        "service":           oids(jobs),
        "total_spent":       total_sales_spend + service_spend,
        "service_spend":     service_spend,
        "total_sales_spend": total_sales_spend,
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  VEHICLES
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/vehicles")
async def list_vehicles(
    brand:   Optional[str] = Query(None),
    status:  Optional[str] = Query(None),
    type:    Optional[str] = Query(None),
    search:  Optional[str] = Query(None),
    limit:   int           = Query(300, ge=1, le=5000),
    current_user=Depends(verify_token),
):
    query: dict = {}
    if brand:  query["brand"]  = brand.upper()
    if status: query["status"] = status
    if type:   query["type"]   = type
    if search:
        s = re.escape(search)
        query["$or"] = [
            {"model":          {"$regex": s, "$options": "i"}},
            {"chassis_number": {"$regex": s, "$options": "i"}},
            {"vehicle_number": {"$regex": s, "$options": "i"}},
            {"color":          {"$regex": s, "$options": "i"}},
        ]
    docs  = await db.vehicles.find(query).sort("created_at", -1).limit(limit).to_list(limit)
    total = await db.vehicles.count_documents(query)
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})

@api_router.post("/vehicles", status_code=201)
async def create_vehicle(body: VehicleCreate, current_user=Depends(verify_token)):
    chassis = body.chassis_number.strip().upper()
    if await db.vehicles.find_one({"chassis_number": chassis}):
        raise HTTPException(status_code=409, detail=f"Chassis number {chassis} already exists")
    doc = body.dict()
    doc["chassis_number"] = chassis
    doc["brand"]          = doc["brand"].upper()
    doc["created_at"]     = utcnow().isoformat()
    result = await db.vehicles.insert_one(doc)
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)
    return doc

@api_router.get("/vehicles/stats/summary")
async def vehicle_stats(current_user=Depends(verify_token)):
    in_stock, sold, in_service, new_count, used_count = await asyncio.gather(
        db.vehicles.count_documents({"status": "in_stock"}),
        db.vehicles.count_documents({"status": "sold"}),
        db.vehicles.count_documents({"status": "in_service"}),
        db.vehicles.count_documents({"type": "new"}),
        db.vehicles.count_documents({"type": "used"}),
    )
    pipeline = [{"$match": {"status": "in_stock"}}, {"$group": {"_id": None, "total": {"$sum": "$purchase_price"}}}]
    result   = await db.vehicles.aggregate(pipeline).to_list(1)
    stock_val = result[0]["total"] if result else 0
    return {"in_stock": in_stock, "sold": sold, "in_service": in_service, "new": new_count, "used": used_count, "stock_value": stock_val}

@api_router.get("/vehicles/{veh_id}")
async def get_vehicle(veh_id: str, current_user=Depends(verify_token)):
    doc = await db.vehicles.find_one({"_id": obj_id(veh_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    return oid(doc)

@api_router.put("/vehicles/{veh_id}")
async def update_vehicle(veh_id: str, body: VehicleUpdate, current_user=Depends(verify_token)):
    update = {k: v for k, v in body.dict().items() if v is not None}
    if "chassis_number" in update:
        update["chassis_number"] = update["chassis_number"].strip().upper()
    if "brand" in update:
        update["brand"] = update["brand"].upper()
    await db.vehicles.update_one({"_id": obj_id(veh_id)}, {"$set": update})
    return oid(await db.vehicles.find_one({"_id": obj_id(veh_id)}))

@api_router.delete("/vehicles/{veh_id}")
async def delete_vehicle(veh_id: str, current_user=Depends(require_admin)):
    result = await db.vehicles.delete_one({"_id": obj_id(veh_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    return {"message": "Deleted"}


# ═══════════════════════════════════════════════════════════════════════════════
#  PDF INVOICE GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def _generate_sale_pdf(sale: dict) -> bytes:
    _register_fonts()
    GOLD        = colors.HexColor('#B8860B')
    GOLD_LIGHT  = colors.HexColor('#F5E6C0')
    DARK        = colors.HexColor('#1A1A1A')
    MID         = colors.HexColor('#4A4A4A')
    DIM         = colors.HexColor('#8A8A8A')
    RULE        = colors.HexColor('#D0C090')
    STRIPE      = colors.HexColor('#F7F7F4')
    WHITE       = colors.white
    PAGE_BG     = colors.HexColor('#FDFCF8')
    RS          = 'Rs.'
    W, H = A4
    ML = 16*mm; MR = W - 16*mm; TW = MR - ML
    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    c.setFillColor(PAGE_BG); c.rect(0, 0, W, H, fill=1, stroke=0)
    c.setFillColor(DARK); c.rect(0, H - 5*mm, W, 5*mm, fill=1, stroke=0)
    c.setFillColor(GOLD); c.rect(0, H - 6.5*mm, W, 1.5*mm, fill=1, stroke=0)
    HEADER_Y = H - 30*mm
    logo_path = os.path.join(os.path.dirname(__file__), 'mm_logo.png')
    if os.path.exists(logo_path):
        LOGO_SIZE = 30*mm
        c.drawImage(logo_path, ML, HEADER_Y - 10*mm, width=LOGO_SIZE, height=LOGO_SIZE, preserveAspectRatio=True, mask='auto')
        name_x = ML + 30*mm
    else:
        name_x = ML
    c.setFont('Sans-Bold', 20); c.setFillColor(DARK); c.drawString(name_x, HEADER_Y + 4*mm, 'MM MOTORS')
    c.setFont('Sans', 7); c.setFillColor(DIM); c.drawString(name_x, HEADER_Y - 2*mm, 'MULTI-BRAND DEALERSHIP  ·  MALUR')
    c.setFont('Mono-Bold', 8); c.setFillColor(GOLD); c.drawRightString(MR, HEADER_Y + 5*mm, 'SALE  INVOICE')
    c.setFont('Mono-Bold', 13); c.setFillColor(DARK); c.drawRightString(MR, HEADER_Y - 2*mm, sale.get('invoice_number', '—'))
    c.setFont('Sans', 7.5); c.setFillColor(DIM); c.drawRightString(MR, HEADER_Y - 7*mm, f"Date: {sale.get('sale_date', '—')}")
    DIV_Y = HEADER_Y - 12*mm
    c.setStrokeColor(GOLD); c.setLineWidth(0.8); c.line(ML, DIV_Y, MR, DIV_Y)
    def sec_label(x, y, text):
        c.setFont('Sans-Bold', 6.5); c.setFillColor(GOLD); c.drawString(x, y, text.upper())
        c.setStrokeColor(GOLD); c.setLineWidth(0.35); c.line(x, y - 1*mm, x + 38*mm, y - 1*mm)
    def irow(x, y, label, val, mono=False, max_w=None):
        c.setFont('Sans', 7); c.setFillColor(DIM); c.drawString(x, y, label)
        c.setFont('Mono' if mono else 'Sans-Bold', 7); c.setFillColor(DARK)
        text = str(val) if val else '—'
        # Strip rupee unicode — use "Rs." instead so Liberation Sans renders it
        text = text.replace('₹', 'Rs.')
        if max_w:
            # Truncate with ellipsis if text exceeds column width
            while c.stringWidth(text, 'Mono' if mono else 'Sans-Bold', 7) > max_w and len(text) > 4:
                text = text[:-2] + '…'
        c.drawString(x + 24*mm, y, text)
    def irow_wrap(x, y, label, val, max_w, line_h=4.5*mm):
        """Like irow but wraps long values across multiple lines."""
        c.setFont('Sans', 7); c.setFillColor(DIM); c.drawString(x, y, label)
        c.setFont('Sans-Bold', 7); c.setFillColor(DARK)
        text = str(val) if val else '—'
        words = text.split()
        line = ''; lines = []
        for w in words:
            test = (line + ' ' + w).strip()
            if c.stringWidth(test, 'Sans-Bold', 7) <= max_w:
                line = test
            else:
                if line: lines.append(line)
                line = w
        if line: lines.append(line)
        for li, ln in enumerate(lines):
            c.drawString(x + 24*mm, y - li * line_h, ln)
        return len(lines)
    COL1 = ML; COL2 = W/2 + 2*mm; RH = 5*mm
    COL1_MAX = (W/2 - 2*mm) - (ML + 24*mm)   # max text width for COL1 values
    COL2_MAX = MR - (COL2 + 24*mm)            # max text width for COL2 values
    INFO_Y = DIV_Y - 6*mm
    nominee = sale.get('nominee', {}) or {}
    sec_label(COL1, INFO_Y, 'Customer Details')
    # Address uses wrap; others use single-line with truncation guard
    irow(COL1, INFO_Y - 1.6*RH, 'Name',    sale.get('customer_name',''),   max_w=COL1_MAX)
    irow(COL1, INFO_Y - 2.6*RH, 'C/O',     sale.get('care_of',''),          max_w=COL1_MAX)
    irow(COL1, INFO_Y - 3.6*RH, 'Mobile',  sale.get('customer_mobile',''),  max_w=COL1_MAX)
    addr_lines = irow_wrap(COL1, INFO_Y - 4.6*RH, 'Address', sale.get('customer_address',''), max_w=COL1_MAX)
    addr_extra = max(0, addr_lines - 1)
    irow(COL1, INFO_Y - (5.6 + addr_extra)*RH, 'Payment', sale.get('payment_mode',''), max_w=COL1_MAX)
    sec_label(COL2, INFO_Y, 'Vehicle Details')
    for i, (l, v) in enumerate([('Brand', sale.get('vehicle_brand','')), ('Model', sale.get('vehicle_model','')), ('Variant', sale.get('vehicle_variant','')), ('Colour', sale.get('vehicle_color','')), ('Financier', sale.get('financier',''))]):
        irow(COL2, INFO_Y - (i+1.6)*RH, l, v, max_w=COL2_MAX)
    SEC2_Y = INFO_Y - (7.8 + addr_extra)*RH   # extra row for wrapped address
    sec_label(COL1, SEC2_Y, 'Registration / Chassis')
    for i, (l, v, m) in enumerate([
        ('Vehicle No', sale.get('vehicle_number',''), False),
        ('RTO',        sale.get('rto',''),            False),
        ('Chassis No', sale.get('chassis_number',''), True),
        ('Engine No',  sale.get('engine_number',''),  True),
    ]):
        irow(COL1, SEC2_Y - (i+1.6)*RH, l, v, mono=m, max_w=COL1_MAX)
    sec_label(COL2, SEC2_Y, 'Nominee (Insurance)')
    for i, (l, v) in enumerate([('Name', nominee.get('name','')), ('Relation', nominee.get('relation','')), ('Age', nominee.get('age','')), ('Mobile', nominee.get('number',''))]):
        irow(COL2, SEC2_Y - (i+1.6)*RH, l, v)
    TBL_Y = SEC2_Y - 5.8*RH; TBL_TY = TBL_Y + 1*mm
    c.setFillColor(DARK); c.rect(ML, TBL_TY - 6.5*mm, TW, 6.5*mm, fill=1, stroke=0)
    c.setFont('Sans-Bold', 7); c.setFillColor(WHITE)
    TC = [ML+3*mm, ML+65*mm, ML+100*mm, ML+128*mm, MR-3*mm]
    c.drawString(TC[0], TBL_TY - 4.5*mm, 'DESCRIPTION'); c.drawString(TC[1], TBL_TY - 4.5*mm, 'CHASSIS / DETAILS')
    c.drawString(TC[2], TBL_TY - 4.5*mm, 'PAYMENT'); c.drawString(TC[3], TBL_TY - 4.5*mm, 'MODE')
    c.drawRightString(TC[4], TBL_TY - 4.5*mm, 'AMOUNT')
    DR_Y = TBL_TY - 14*mm
    c.setFillColor(STRIPE); c.rect(ML, DR_Y - 1*mm, TW, 9*mm, fill=1, stroke=0)
    c.setFont('Sans-Bold', 8); c.setFillColor(DARK)
    c.drawString(TC[0], DR_Y + 4*mm, f"{sale.get('vehicle_brand','')} {sale.get('vehicle_model','')}")
    c.setFont('Sans', 7); c.setFillColor(DIM)
    c.drawString(TC[0], DR_Y + 0.8*mm, f"{sale.get('vehicle_variant','')}  ·  {sale.get('vehicle_color','')}")
    c.setFont('Mono', 7); c.setFillColor(DARK); c.drawString(TC[1], DR_Y + 2.5*mm, sale.get('chassis_number',''))
    c.setFont('Sans', 7.5); c.drawString(TC[2], DR_Y + 2.5*mm, sale.get('payment_mode','')); c.drawString(TC[3], DR_Y + 2.5*mm, 'Full Payment')
    total = sale.get('total_amount') or sale.get('sale_price') or 0
    total_str = f"{int(total):,}" if isinstance(total, (int, float)) else str(total)
    c.setFont('Sans-Bold', 10); c.setFillColor(GOLD); c.drawRightString(TC[4], DR_Y + 2.5*mm, f"{RS}{total_str}")
    c.setStrokeColor(GOLD); c.setLineWidth(0.7); c.line(ML, DR_Y - 2*mm, MR, DR_Y - 2*mm)
    TOT_Y = DR_Y - 12*mm
    c.setFont('Sans-Italic', 7); c.setFillColor(DIM); c.drawString(ML, TOT_Y + 4*mm, sale.get('amount_in_words',''))
    c.setFillColor(GOLD_LIGHT); c.setStrokeColor(GOLD); c.setLineWidth(0.6)
    c.roundRect(MR - 58*mm, TOT_Y - 1*mm, 58*mm, 11*mm, 1.5*mm, fill=1, stroke=1)
    c.setFont('Sans', 7); c.setFillColor(MID); c.drawString(MR - 55*mm, TOT_Y + 5.5*mm, 'TOTAL AMOUNT')
    c.setFont('Sans-Bold', 13); c.setFillColor(DARK); c.drawRightString(MR - 3*mm, TOT_Y + 4*mm, f"{RS}{total_str}")
    SIG_Y = TOT_Y - 14*mm
    c.setStrokeColor(RULE); c.setLineWidth(0.5)
    c.line(ML, SIG_Y, ML + 52*mm, SIG_Y)
    c.setFont('Sans', 6.5); c.setFillColor(DIM); c.drawString(ML, SIG_Y - 4*mm, "Customer's Signature")
    c.setFont('Sans-Bold', 7); c.setFillColor(DARK); c.drawString(ML, SIG_Y - 8.5*mm, sale.get('customer_name','').upper())
    c.setFont('Sans', 6.5); c.setFillColor(DIM); c.drawCentredString(W/2, SIG_Y - 4*mm, f"Sold by: {sale.get('sold_by','MM Motors')}")
    c.line(MR - 52*mm, SIG_Y, MR, SIG_Y)
    c.setFont('Sans', 6.5); c.setFillColor(DIM); c.drawRightString(MR, SIG_Y - 4*mm, 'Authorised Signatory')
    c.setFont('Sans-Bold', 7); c.setFillColor(DARK); c.drawRightString(MR, SIG_Y - 8.5*mm, 'MM MOTORS')
    SCHED_Y = SIG_Y - 20*mm
    c.setFillColor(DARK); c.rect(ML, SCHED_Y, TW, 6.5*mm, fill=1, stroke=0)
    c.setFillColor(GOLD); c.rect(ML, SCHED_Y + 6.5*mm, TW, 0.8*mm, fill=1, stroke=0)
    c.setFont('Sans-Bold', 8.5); c.setFillColor(WHITE); c.drawString(ML + 4*mm, SCHED_Y + 2*mm, 'SERVICE SCHEDULE')
    DEAR_Y = SCHED_Y - 13.5*mm
    c.setFillColor(GOLD_LIGHT); c.setStrokeColor(GOLD); c.setLineWidth(0.5)
    c.roundRect(ML, DEAR_Y, TW, 12*mm, 1*mm, fill=1, stroke=1)
    c.setFont('Sans-Bold', 7.5); c.setFillColor(DARK); c.drawString(ML + 3*mm, DEAR_Y + 8*mm, 'DEAR VALUED CUSTOMER,')
    c.setFont('Sans', 7); c.setFillColor(MID)
    c.drawString(ML + 3*mm, DEAR_Y + 4.5*mm, 'We thank you for choosing our world-class vehicle. To ensure optimal performance and longevity,')
    c.drawString(ML + 3*mm, DEAR_Y + 1.5*mm, 'please follow the service schedule below for a pleasant riding experience at all times.')
    TH_Y = DEAR_Y - 8*mm
    c.setFillColor(DARK); c.rect(ML, TH_Y, TW, 6.5*mm, fill=1, stroke=0)
    c.setFont('Sans-Bold', 7); c.setFillColor(WHITE)
    SC1 = ML + 3*mm; SC2 = ML + 52*mm; SC3 = ML + 115*mm
    c.drawString(SC1, TH_Y + 2*mm, 'SERVICE DATE'); c.drawString(SC2, TH_Y + 2*mm, 'SERVICE TYPE'); c.drawString(SC3, TH_Y + 2*mm, 'RECOMMENDED SCHEDULE')
    SERVICES = [('FIRST SERVICE','500-700 kms or 15-30 days'),('SECOND SERVICE','3000-3500 kms or 30-90 days'),('THIRD SERVICE','6000-6500 kms or 90-180 days'),('FOURTH SERVICE','9000-9500 kms or 180-270 days')]
    SRH = 6.5*mm
    for i, (stype, sched) in enumerate(SERVICES):
        ry = TH_Y - (i + 1) * SRH
        c.setFillColor(WHITE if i % 2 == 0 else STRIPE); c.rect(ML, ry, TW, SRH, fill=1, stroke=0)
        c.setStrokeColor(RULE); c.setLineWidth(0.3); c.line(ML, ry, MR, ry)
        c.setFont('Mono', 7); c.setFillColor(DIM); c.drawString(SC1, ry + 2*mm, '__/__/____')
        c.setFont('Sans-Bold', 7); c.setFillColor(GOLD); c.drawString(SC2, ry + 2*mm, stype)
        c.setFont('Sans', 7); c.setFillColor(DARK); c.drawString(SC3, ry + 2*mm, sched)
    c.setStrokeColor(RULE); c.setLineWidth(0.5)
    c.rect(ML, TH_Y - 4*SRH - 1, TW, 6.5*mm + 13.5*mm + 8*mm + 4*SRH + 1, fill=0, stroke=1)
    NOTE_Y = TH_Y - 4*SRH - 6.5*mm
    c.setFillColor(GOLD_LIGHT); c.setStrokeColor(GOLD); c.setLineWidth(0.5)
    c.roundRect(ML, NOTE_Y - 0.5*mm, TW, 6*mm, 1*mm, fill=1, stroke=1)
    c.setFont('Sans-Bold', 7.5); c.setFillColor(colors.HexColor('#7A5800'))
    c.drawCentredString(W/2, NOTE_Y + 1.8*mm, 'IMPORTANT: Follow whichever milestone comes first (km or days)')
    THANKS_Y = NOTE_Y - 17*mm
    c.setFillColor(STRIPE); c.setStrokeColor(RULE); c.setLineWidth(0.5)
    c.roundRect(ML, THANKS_Y - 1*mm, TW, 28*mm, 2*mm, fill=1, stroke=1)
    c.setFillColor(GOLD); c.rect(ML, THANKS_Y + 26*mm, TW, 1.5*mm, fill=1, stroke=0)
    for label, x in [('* Trusted Dealer', ML + 6*mm), ('* 24/7 Service Support', W/2 - 22*mm), ('* Quality Guaranteed', MR - 50*mm)]:
        c.setFont('Sans', 7); c.setFillColor(MID); c.drawString(x, THANKS_Y + 20.5*mm, label)
    c.setStrokeColor(RULE); c.setLineWidth(0.4); c.line(ML + 4*mm, THANKS_Y + 18.5*mm, MR - 4*mm, THANKS_Y + 18.5*mm)
    c.setFont('Sans-Bold', 12); c.setFillColor(DARK); c.drawCentredString(W/2, THANKS_Y + 13*mm, 'Thank You for Choosing M M Motors!')
    c.setFont('Sans-Italic', 7.5); c.setFillColor(DIM); c.drawCentredString(W/2, THANKS_Y + 8.5*mm, 'Your trust drives our excellence in two-wheeler sales and service.')
    c.setFont('Sans', 7.5); c.setFillColor(MID); c.drawCentredString(W/2, THANKS_Y + 4.5*mm, '* Premium Quality   *  Expert Service   *  Customer First')
    c.setFillColor(DARK); c.rect(0, 0, W, 7.5*mm, fill=1, stroke=0)
    c.setFillColor(GOLD); c.rect(0, 7.5*mm, W, 1*mm, fill=1, stroke=0)
    c.setFont('Sans', 6); c.setFillColor(colors.HexColor('#888888'))
    c.drawString(ML, 2.8*mm, 'This is a computer-generated document. No signature required if digitally authenticated.')
    c.drawRightString(MR, 2.8*mm, 'MM Motors  ·  Malur  ·  Multi-brand Dealership')
    c.save(); buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
#  SALES
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/sales")
async def list_sales(
    customer_id: Optional[str] = Query(None),
    status:      Optional[str] = Query(None),
    from_date:   Optional[str] = Query(None),
    to_date:     Optional[str] = Query(None),
    search:      Optional[str] = Query(None),
    limit:       int           = Query(300, ge=1, le=5000),
    current_user=Depends(verify_token),
):
    query: dict = {}
    if customer_id: query["customer_id"] = customer_id
    if status:      query["status"]      = status
    if search:
        s = re.escape(search)
        query["$or"] = [
            {"invoice_number": {"$regex": s, "$options": "i"}},
            {"customer_name":  {"$regex": s, "$options": "i"}},
            {"vehicle_model":  {"$regex": s, "$options": "i"}},
            {"vehicle_number": {"$regex": s, "$options": "i"}},
        ]
    docs  = await db.sales.find(query).sort("created_at", -1).limit(limit).to_list(limit)
    total = await db.sales.count_documents(query)
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})

@api_router.post("/sales", status_code=201)
async def create_sale(body: SaleCreate, current_user=Depends(verify_token)):
    customer = await db.customers.find_one({"_id": obj_id(body.customer_id)})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    vehicle = await db.vehicles.find_one({"_id": obj_id(body.vehicle_id)})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    if vehicle.get("status") == "sold":
        raise HTTPException(status_code=409, detail="Vehicle already sold")
    total_amount = body.total_amount if body.total_amount is not None else (body.sale_price or 0)
    inv_no    = await next_sequence("invoice")
    sale_date = body.sale_date or utcnow().strftime("%d %b %Y")
    doc = {
        "invoice_number":  inv_no,
        "customer_id":     body.customer_id,
        "customer_name":   customer["name"],
        "customer_mobile": customer.get("mobile",""),
        "customer_address":customer.get("address",""),
        "care_of":         body.care_of or "",
        "vehicle_id":      body.vehicle_id,
        "vehicle_brand":   body.vehicle_brand   if body.vehicle_brand   is not None else vehicle["brand"],
        "vehicle_model":   body.vehicle_model   if body.vehicle_model   is not None else vehicle["model"],
        "vehicle_variant": body.vehicle_variant if body.vehicle_variant is not None else vehicle.get("variant",""),
        "vehicle_color":   body.vehicle_color   if body.vehicle_color   is not None else vehicle.get("color",""),
        "chassis_number":  body.chassis_number  if body.chassis_number  is not None else vehicle.get("chassis_number",""),
        "engine_number":   body.engine_number   if body.engine_number   is not None else vehicle.get("engine_number",""),
        "vehicle_number":  body.vehicle_number or vehicle.get("vehicle_number",""),
        "sale_price":      body.sale_price,
        "total_amount":    round(total_amount, 2),
        "amount_in_words": amount_in_words(total_amount),
        "finance_type":    body.finance_type or "cash",
        "financier":       body.financier or "",
        "loan_amount":     body.loan_amount or 0,
        "nominee":         body.nominee.dict() if body.nominee else {},
        "payment_mode":    body.payment_mode or "Cash",
        "sold_by":         body.sold_by or current_user.get("name",""),
        "sale_date":       sale_date,
        "status":          "pending",
        "notes":           body.notes or "",
        "hsrp_front":      body.hsrp_front or "",
        "hsrp_back":       body.hsrp_back or "",
        "hsrp_front_id":   body.hsrp_front_id or "",
        "hsrp_back_id":    body.hsrp_back_id or "",
        "hsrp_date":       body.hsrp_date or "",
        "hsrp_notes":      body.hsrp_notes or "",
        "milestones":      {**default_milestones(), **(body.milestones or {})},
        "milestone_dates": dict(body.milestone_dates or {}),
        "created_at":      utcnow().isoformat(),
    }
    result = await db.sales.insert_one(doc)
    await db.vehicles.update_one(
        {"_id": obj_id(body.vehicle_id)},
        {"$set": {"status": "sold", "sold_date": sale_date, "invoice_number": inv_no}}
    )
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)
    return doc

@api_router.get("/sales/stats/summary")
async def sales_stats(current_user=Depends(verify_token)):
    today = utcnow().strftime("%d %b %Y")
    pipeline_total = [{"$group": {"_id": None, "total": {"$sum": "$total_amount"}, "count": {"$sum": 1}}}]
    result = await db.sales.aggregate(pipeline_total).to_list(1)
    all_time    = result[0] if result else {"total": 0, "count": 0}
    today_sales = await db.sales.count_documents({"sale_date": today})
    pending     = await db.sales.count_documents({"status": "pending"})
    return {"total_count": all_time.get("count",0), "total_revenue": all_time.get("total",0), "today_count": today_sales, "pending_delivery": pending}

@api_router.get("/sales/{sale_id}")
async def get_sale(sale_id: str, current_user=Depends(verify_token)):
    doc = await db.sales.find_one({"_id": obj_id(sale_id)})
    if not doc:
        doc = await db.sales.find_one({"invoice_number": sale_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Sale not found")
    return oid(doc)

@api_router.get("/sales/{sale_id}/pdf")
async def sale_pdf(sale_id: str, current_user=Depends(verify_token)):
    doc = await db.sales.find_one({"_id": obj_id(sale_id)})
    if not doc:
        doc = await db.sales.find_one({"invoice_number": sale_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Sale not found")
    sale = oid(doc)
    try:
        pdf_bytes = _generate_sale_pdf(sale)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {e}")
    invoice_no = sale.get("invoice_number","invoice").replace("/","-")
    return StreamingResponse(
        io.BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=MM_Motors_{invoice_no}.pdf"},
    )

@api_router.put("/sales/{sale_id}")
async def update_sale(sale_id: str, body: SaleUpdate, current_user=Depends(verify_token)):
    sale = await db.sales.find_one({"_id": obj_id(sale_id)})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    update: dict = {}
    for field in ("status","delivery_date","vehicle_number","payment_mode","notes","customer_name","customer_mobile","customer_address","care_of","total_amount","sale_price","finance_type","financier","loan_amount","sale_date","sold_by","hsrp_front","hsrp_back","hsrp_front_id","hsrp_back_id","hsrp_date","hsrp_notes"):
        val = getattr(body, field)
        if val is not None:
            update[field] = val
    if "total_amount" in update:
        update["amount_in_words"] = amount_in_words(update["total_amount"])
    if body.nominee is not None:
        update["nominee"] = body.nominee.dict()
    if body.vehicle_id is not None and body.vehicle_id != sale.get("vehicle_id",""):
        if current_user.get("role") != "owner":
            raise HTTPException(status_code=403, detail="Only owner can change vehicle on a sale")
        new_vehicle = await db.vehicles.find_one({"_id": obj_id(body.vehicle_id)})
        if not new_vehicle:
            raise HTTPException(status_code=404, detail="New vehicle not found")
        if new_vehicle.get("status") == "sold" and str(new_vehicle["_id"]) != sale.get("vehicle_id",""):
            raise HTTPException(status_code=409, detail="Vehicle already sold")
        if sale.get("vehicle_id"):
            await db.vehicles.update_one({"_id": obj_id(sale["vehicle_id"])}, {"$set": {"status": "in_stock"}, "$unset": {"sold_date":"","invoice_number":""}})
        await db.vehicles.update_one({"_id": obj_id(body.vehicle_id)}, {"$set": {"status":"sold","sold_date":sale.get("sale_date",""),"invoice_number":sale.get("invoice_number","")}})
        update["vehicle_id"]      = body.vehicle_id
        update["vehicle_brand"]   = new_vehicle.get("brand","")
        update["vehicle_model"]   = new_vehicle.get("model","")
        update["vehicle_variant"] = new_vehicle.get("variant","")
        update["vehicle_color"]   = new_vehicle.get("color","")
        update["chassis_number"]  = new_vehicle.get("chassis_number","")
        update["engine_number"]   = new_vehicle.get("engine_number","")
    if not update:
        raise HTTPException(status_code=400, detail="Nothing to update")
    await db.sales.update_one({"_id": obj_id(sale_id)}, {"$set": update})
    return oid(await db.sales.find_one({"_id": obj_id(sale_id)}))


@api_router.patch("/sales/{sale_id}/milestone")
async def update_sale_milestone(sale_id: str, body: MilestoneUpdate, current_user=Depends(verify_token)):
    """Toggle a single sale milestone (documents/invoice/insurance/tax_paid/number_plate).
    When value=True, records the completion date (defaults to today if not supplied).
    When value=False, removes any previously recorded date."""
    if body.key not in SALE_MILESTONE_KEYS:
        raise HTTPException(status_code=400, detail=f"Invalid milestone key. Allowed: {SALE_MILESTONE_KEYS}")
    sale = await db.sales.find_one(
        {"_id": obj_id(sale_id)},
        {"_id": 1, "milestones": 1, "milestone_dates": 1},
    )
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    # Backfill missing milestones dict if this is a legacy sale
    current       = {**default_milestones(), **(sale.get("milestones") or {})}
    current_dates = dict(sale.get("milestone_dates") or {})
    current[body.key] = bool(body.value)
    if body.value:
        d = (body.date or utcnow().strftime("%Y-%m-%d")).strip()
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", d):
            raise HTTPException(status_code=400, detail="date must be YYYY-MM-DD")
        current_dates[body.key] = d
    else:
        current_dates.pop(body.key, None)
    await db.sales.update_one(
        {"_id": obj_id(sale_id)},
        {"$set": {"milestones": current, "milestone_dates": current_dates}},
    )
    return {"id": sale_id, "milestones": current, "milestone_dates": current_dates}


@api_router.delete("/sales/{sale_id}")
async def delete_sale(sale_id: str, current_user=Depends(require_admin)):
    sale = await db.sales.find_one({"_id": obj_id(sale_id)})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    await db.vehicles.update_one({"_id": obj_id(sale["vehicle_id"])}, {"$set": {"status":"in_stock"}, "$unset": {"sold_date":"","invoice_number":""}})
    await db.sales.delete_one({"_id": obj_id(sale_id)})
    await _sync_counter("invoice", "sales", "invoice_number")
    return {"message": "Deleted"}


# ═══════════════════════════════════════════════════════════════════════════════
#  SERVICE JOBS
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/service")
async def list_service_jobs(
    status:      Optional[str] = Query(None),
    customer_id: Optional[str] = Query(None),
    technician:  Optional[str] = Query(None),
    search:      Optional[str] = Query(None),
    limit:       int           = Query(200, ge=1, le=5000),
    current_user=Depends(verify_token),
):
    query: dict = {}
    if status:      query["status"]      = status
    if customer_id: query["customer_id"] = customer_id
    if technician:  query["technician"]  = technician
    if search:
        s = re.escape(search)
        query["$or"] = [
            {"job_number":     {"$regex": s, "$options": "i"}},
            {"customer_name":  {"$regex": s, "$options": "i"}},
            {"vehicle_number": {"$regex": s, "$options": "i"}},
            {"model":          {"$regex": s, "$options": "i"}},
            {"complaint":      {"$regex": s, "$options": "i"}},
        ]
    docs  = await db.service_jobs.find(query).sort("created_at",-1).limit(limit).to_list(limit)
    total = await db.service_jobs.count_documents(query)
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})

@api_router.post("/service", status_code=201)
async def create_service_job(body: ServiceJobCreate, current_user=Depends(verify_token)):
    customer = await db.customers.find_one({"_id": obj_id(body.customer_id)})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    job_no   = await next_sequence("job")
    check_in = body.check_in_date or utcnow().strftime("%d %b %Y")
    veh_no   = body.vehicle_number.strip().upper() if body.vehicle_number else ""
    chassis  = body.chassis_number.strip().upper() if body.chassis_number else ""
    doc = {
        "job_number":         job_no,
        "customer_id":        body.customer_id,
        "customer_name":      customer["name"],
        "customer_mobile":    customer.get("mobile",""),
        "customer_address":   customer.get("address",""),
        "vehicle_number":     veh_no,
        "chassis_number":     chassis,
        "brand":              body.brand.upper(),
        "model":              body.model,
        "variant":            body.variant or "",
        "odometer_km":        body.odometer_km or 0,
        "odometer_out":       0,
        "complaint":          body.complaint,
        "advisor_id":         body.advisor_id or "",
        "technician":         body.technician or "",
        "check_in_date":      check_in,
        "estimated_delivery": body.estimated_delivery or "",
        "delivery_date":      "",
        "status":             "pending",
        "notes":              body.notes or "",
        "created_at":         utcnow().isoformat(),
    }
    result = await db.service_jobs.insert_one(doc)
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)

    # Option B: save vehicle to records if requested (service-only customers)
    if getattr(body, "save_vehicle", False):
        # Only create if not already in vehicles collection
        dedup_query = {"chassis_number": chassis} if chassis else {"vehicle_number": veh_no} if veh_no else None
        existing_veh = await db.vehicles.find_one(dedup_query) if dedup_query else None
        if not existing_veh:
            await db.vehicles.insert_one({
                "brand":            body.brand.upper(),
                "model":            body.model,
                "variant":          body.variant or "",
                "chassis_number":   chassis,
                "vehicle_number":   veh_no,
                "engine_number":    "",
                "color":            "",
                "key_number":       "",
                "type":             "used",
                "status":           "in_service",
                "customer_id":      body.customer_id,
                "customer_name":    customer["name"],
                "customer_mobile":  customer.get("mobile",""),
                "inbound_date":     check_in,
                "inbound_location": "Service",
                "return_date":      "",
                "returned_location":"",
                "created_at":       utcnow().isoformat(),
                "_source":          "service",
            })

    return doc


# ── Service Due ────────────────────────────────────────────────────────────────
# Module-level TTL cache — invalidated on sales/service_jobs writes
_SERVICE_DUE_CACHE: dict = {}   # key: days (int) → {"ts": datetime, "data": list}
_SERVICE_DUE_TTL_SEC = 60

def _invalidate_service_due_cache():
    _SERVICE_DUE_CACHE.clear()

@api_router.get("/service/due")
async def service_due(
    days: int               = Query(30, ge=1, le=365),  # lookahead window
    first_service_days: int = Query(30, ge=1, le=90),   # kept for API compat, ignored — hardcoded 30d
    current_user=Depends(verify_token),
):
    """
    Service schedule:
      Sold vehicles:
        1st service  = sale_date + 30 days
        2nd service  = 1st service date + 90 days
        3rd+ service = last service date + 90 days
      Service-only vehicles:
        next service = last service date + 90 days

    Returns vehicles whose next_due_date <= today + lookahead (days param).
    Cached in-memory for 60s; invalidated on sales/service_jobs write.
    """
    # Cache lookup
    cached = _SERVICE_DUE_CACHE.get(days)
    if cached and (utcnow() - cached["ts"]).total_seconds() < _SERVICE_DUE_TTL_SEC:
        return cached["data"]

    FIRST_SVC_INTERVAL  = 30   # days after sale for 1st service
    REPEAT_SVC_INTERVAL = 90   # days between 2nd+ services

    now         = utcnow()
    lookahead   = now + timedelta(days=days)

    def parse_date(s):
        for fmt in ("%d %b %Y", "%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"):
            try: return datetime.strptime(s[:len(fmt)+2].strip(), fmt)
            except (ValueError, TypeError): continue
        return None

    def ordinal(n):
        return {1:"1st",2:"2nd",3:"3rd"}.get(n, f"{n}th")

    # ── 1. Sales map: vehicle_number → most-recent sale ─────────────────────
    sales_docs = await db.sales.find(
        {}, {"vehicle_number":1,"sale_date":1,"customer_name":1,
             "customer_mobile":1,"vehicle_brand":1,"vehicle_model":1,"created_at":1}
    ).to_list(20000)

    # Cutoff from env var (YYYY-MM-DD), fallback 2026-05-01
    _cutoff_str = os.getenv("SERVICE_DUE_CUTOFF", "2026-05-01").strip()
    try:
        SOLD_SINCE = datetime.strptime(_cutoff_str, "%Y-%m-%d")
    except ValueError:
        SOLD_SINCE = datetime(2026, 5, 1)

    sales_map = {}
    for s in sales_docs:
        vn = (s.get("vehicle_number") or "").upper().strip()
        if not vn:
            continue
        dt = parse_date(s.get("sale_date","")) or parse_date(s.get("created_at",""))
        # Skip vehicles sold before May 1, 2026
        if not dt or dt < SOLD_SINCE:
            continue
        existing = sales_map.get(vn)
        if not existing or (dt and (not existing["dt"] or dt > existing["dt"])):
            sales_map[vn] = {
                "dt":     dt,
                "str":    s.get("sale_date",""),
                "name":   s.get("customer_name",""),
                "mobile": s.get("customer_mobile",""),
                "brand":  s.get("vehicle_brand",""),
                "model":  s.get("vehicle_model",""),
            }

    # ── 2. Service history per vehicle (oldest-first for counting) ───────────
    oldest_pipeline = [
        {"$match": {"status": "delivered"}},
        {"$sort": {"created_at": 1}},          # oldest first
        {"$group": {
            "_id":   "$vehicle_number",
            "jobs":  {"$push": {
                "job_id":   "$_id", "job_number": "$job_number",
                "created_at": "$created_at", "check_in_date": "$check_in_date",
                "customer_name": "$customer_name", "customer_mobile": "$customer_mobile",
                "brand": "$brand", "model": "$model", "complaint": "$complaint",
            }},
        }},
    ]
    grouped = await db.service_jobs.aggregate(oldest_pipeline).to_list(10000)
    service_map = {}
    for g in grouped:
        vn = (g["_id"] or "").upper().strip()
        if vn:
            service_map[vn] = g["jobs"]   # [oldest … newest]

    # ── 3. Compute next_due for every known vehicle ──────────────────────────
    result = []
    for vn in set(sales_map) | set(service_map):
        is_sold     = vn in sales_map
        jobs        = service_map.get(vn, [])   # oldest → newest
        job_count   = len(jobs)
        last_job    = jobs[-1] if jobs else None
        first_job   = jobs[0]  if jobs else None

        last_svc_dt  = parse_date(last_job["created_at"])  if last_job  else None
        first_svc_dt = parse_date(first_job["created_at"]) if first_job else None

        # ── Determine next due date ──────────────────────────────────────────
        if is_sold:
            sale   = sales_map[vn]
            sale_dt = sale["dt"]

            if job_count == 0:
                # Never serviced — 1st service 30d after sale
                if not sale_dt: continue
                next_due   = sale_dt + timedelta(days=FIRST_SVC_INTERVAL)
                next_label = "1st Service"
                ref_date   = sale["str"]
                ref_label  = "Delivery date"
                source     = "sale"
            elif job_count == 1:
                # 1st done — 2nd service 90d after 1st service
                if not first_svc_dt: continue
                next_due   = first_svc_dt + timedelta(days=REPEAT_SVC_INTERVAL)
                next_label = "2nd Service"
                ref_date   = last_job.get("check_in_date","")
                ref_label  = last_job.get("complaint","")
                source     = "sale"
            else:
                # 2nd+ done — 90d from last service
                if not last_svc_dt: continue
                next_due   = last_svc_dt + timedelta(days=REPEAT_SVC_INTERVAL)
                next_label = f"{ordinal(job_count + 1)} Service"
                ref_date   = last_job.get("check_in_date","")
                ref_label  = last_job.get("complaint","")
                source     = "sale"
        else:
            # Service-only — 90d from last service
            if not last_svc_dt: continue
            next_due   = last_svc_dt + timedelta(days=REPEAT_SVC_INTERVAL)
            next_label = f"{ordinal(job_count + 1)} Service"
            ref_date   = last_job.get("check_in_date","")
            ref_label  = last_job.get("complaint","")
            source     = "service"

        # Skip if not due within lookahead window
        if next_due > lookahead:
            continue

        due_in_days = (next_due - now).days   # negative = overdue
        days_since  = (now - last_svc_dt).days if last_svc_dt else None
        urgency     = "overdue" if due_in_days < 0 else "due_soon"

        # Customer / vehicle info — prefer latest service job over sale record
        info = sales_map.get(vn, {})
        if last_job:
            cname   = last_job.get("customer_name")   or info.get("name","")
            cmobile = last_job.get("customer_mobile") or info.get("mobile","")
            brand   = last_job.get("brand")           or info.get("brand","")
            model   = last_job.get("model")           or info.get("model","")
            jnum    = last_job.get("job_number","")
            jid     = str(last_job.get("job_id",""))
        else:
            cname   = info.get("name","")
            cmobile = info.get("mobile","")
            brand   = info.get("brand","")
            model   = info.get("model","")
            jnum    = ""
            jid     = ""

        result.append({
            "id":               jid,
            "job_number":       jnum,
            "vehicle_number":   vn,
            "customer_name":    cname,
            "customer_mobile":  cmobile,
            "brand":            brand,
            "model":            model,
            "check_in_date":    ref_date,
            "complaint":        ref_label,
            "created_at":       last_svc_dt.isoformat() if last_svc_dt else "",
            "days_since":       days_since,
            "due_in_days":      due_in_days,
            "next_due_date":    next_due.strftime("%d %b %Y"),
            "service_count":    job_count,
            "service_type":     next_label,
            "urgency":          urgency,
            "source":           source,
            "vehicle_type":     "sold" if is_sold else "service_only",
        })

    # Most overdue first, then soonest due
    result.sort(key=lambda x: x["due_in_days"])
    trimmed = result[:500]
    _SERVICE_DUE_CACHE[days] = {"ts": utcnow(), "data": trimmed}
    return trimmed

@api_router.post("/service/due/{vehicle_number}/notified")
async def mark_notified(vehicle_number: str, current_user=Depends(verify_token)):
    """Mark a vehicle as notified — stores timestamp so you know who was contacted."""
    key = vehicle_number.upper().strip()
    await db.service_notifications.update_one(
        {"vehicle_number": key},
        {"$set": {
            "vehicle_number":  key,
            "notified_at":     utcnow().isoformat(),
            "notified_by":     current_user.get("name", ""),
        }},
        upsert=True,
    )
    return {"message": "Marked as notified"}

@api_router.get("/service/due/notifications")
async def get_notifications(current_user=Depends(verify_token)):
    """Return vehicle_number → last notified_at map."""
    docs = await db.service_notifications.find({}).to_list(2000)
    return {d["vehicle_number"]: d["notified_at"] for d in docs if d.get("vehicle_number")}


@api_router.get("/service/stats")
async def service_stats(current_user=Depends(verify_token)):
    pending, in_progress, ready, delivered = await asyncio.gather(
        db.service_jobs.count_documents({"status":"pending"}),
        db.service_jobs.count_documents({"status":"in_progress"}),
        db.service_jobs.count_documents({"status":"ready"}),
        db.service_jobs.count_documents({"status":"delivered"}),
    )
    return {"pending":pending,"in_progress":in_progress,"ready":ready,"delivered":delivered,"total_active":pending+in_progress+ready}

@api_router.get("/service/{job_id}")
async def get_service_job(job_id: str, current_user=Depends(verify_token)):
    doc = await db.service_jobs.find_one({"_id": obj_id(job_id)})
    if not doc:
        doc = await db.service_jobs.find_one({"job_number": job_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Service job not found")
    return oid(doc)

@api_router.put("/service/{job_id}")
async def update_service_job(job_id: str, body: ServiceJobUpdate, current_user=Depends(verify_token)):
    update = {k: v for k, v in body.dict().items() if v is not None}
    if "status" in update and update["status"] == "delivered":
        update.setdefault("delivery_date", utcnow().strftime("%d %b %Y"))
    await db.service_jobs.update_one({"_id": obj_id(job_id)}, {"$set": update})
    return oid(await db.service_jobs.find_one({"_id": obj_id(job_id)}))

@api_router.delete("/service/{job_id}")
async def delete_service_job(job_id: str, current_user=Depends(require_admin)):
    result = await db.service_jobs.delete_one({"_id": obj_id(job_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Job not found")
    await db.service_bills.delete_many({"job_id": job_id})
    await _sync_counter("job", "service_jobs", "job_number")
    return {"message": "Deleted"}


# ═══════════════════════════════════════════════════════════════════════════════
#  SERVICE BILLS (GST)
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/service-bills")
async def list_service_bills(
    job_id: Optional[str] = Query(None),
    p=Depends(paginate_params),
    current_user=Depends(verify_token),
):
    query = {"job_id": job_id} if job_id else {}
    docs  = await db.service_bills.find(query).skip(p["skip"]).limit(p["limit"]).sort("created_at",-1).to_list(p["limit"])
    total = await db.service_bills.count_documents(query)
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})

# PATCH 1: Fixed indentation + returns existing bill instead of 409 ────────────
@api_router.post("/service-bills", status_code=201)
async def create_service_bill(body: ServiceBillCreate, current_user=Depends(verify_token)):
    job = await db.service_jobs.find_one({"_id": obj_id(body.job_id)})
    if not job:
        raise HTTPException(status_code=404, detail="Service job not found")
    existing = await db.service_bills.find_one({"job_id": body.job_id})
    if existing:
        # Return the existing bill instead of 409 — frontend uses GET to load anyway
        return JSONResponse(content=oid(existing), status_code=200)

    # Build line items with GST + per-line discount
    items_out = []
    for item in (body.items or []):
        if item.complimentary:
            line = {"gross":0.0,"discount":0.0,"taxable":0.0,"cgst":0.0,"sgst":0.0,"gst_total":0.0,"total":0.0}
            item_disc = 0.0
        else:
            item_disc = max(0.0, float(item.discount or 0))
            line = calc_gst_line(item.unit_price, item.qty, item.gst_rate, item_disc)
        items_out.append({
            "description":   item.description,
            "part_number":   item.part_number or "",
            "hsn_code":      item.hsn_code or "",
            "qty":           item.qty,
            "unit_price":    0 if item.complimentary else item.unit_price,
            "mrp":           item.unit_price if item.complimentary else item.unit_price,
            "gst_rate":      0 if item.complimentary else item.gst_rate,
            "discount":      0 if item.complimentary else item_disc,
            "complimentary": bool(item.complimentary),
            **line,
        })

    # Labour as separate line
    if body.labour_charges and body.labour_charges > 0:
        lab_disc = max(0.0, float(body.labour_discount or 0))
        lab_line = calc_gst_line(body.labour_charges, 1, body.labour_gst_rate or 18, lab_disc)
        items_out.insert(0, {
            "description": "Labour charges",
            "part_number": "", "hsn_code": "9987",
            "qty": 1, "unit_price": body.labour_charges,
            "gst_rate": body.labour_gst_rate or 18,
            "discount": lab_disc,
            **lab_line,
        })

    totals  = calc_bill_totals([{"unit_price": i["unit_price"], "qty": i["qty"], "gst_rate": i["gst_rate"], "discount": i.get("discount", 0)} for i in items_out])
    # Apply discount off grand_total (post-GST). Clamped 0..grand_total.
    discount = max(0.0, min(float(body.discount or 0), totals["grand_total"]))
    net_total = round(totals["grand_total"] - discount, 2)
    # Bill number derives from job number — no separate counter consumed
    bill_no = job.get("job_number", "").replace("SRV", "SRV-B")

    doc = {
        "bill_number":    bill_no,
        "job_id":         body.job_id,
        "job_number":     job.get("job_number",""),
        "customer_id":    job.get("customer_id",""),
        "customer_name":  job.get("customer_name",""),
        "customer_mobile":job.get("customer_mobile",""),
        "vehicle_number": job.get("vehicle_number",""),
        "brand":          job.get("brand",""),
        "model":          job.get("model",""),
        "items":          items_out,
        "labour_charges": body.labour_charges or 0,
        **totals,
        "discount":       discount,
        "net_total":      net_total,
        "amount_in_words": amount_in_words(net_total),
        "payment_mode":   body.payment_mode or "Cash",
        "notes":          body.notes or "",
        "bill_date":      utcnow().strftime("%d %b %Y"),
        "created_at":     utcnow().isoformat(),
    }
    result = await db.service_bills.insert_one(doc)
    await db.service_jobs.update_one(
        {"_id": obj_id(body.job_id)},
        {"$set": {"status": "ready", "bill_number": doc["bill_number"], "grand_total": net_total}}
    )
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)
    return doc

@api_router.get("/service-bills/{bill_id}")
async def get_service_bill(bill_id: str, current_user=Depends(verify_token)):
    doc = await db.service_bills.find_one({"_id": obj_id(bill_id)})
    if not doc:
        doc = await db.service_bills.find_one({"bill_number": bill_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Bill not found")
    return oid(doc)

# PATCH 2: PUT /service-bills/{bill_id} — was a string variable, now real code ─
@api_router.put("/service-bills/{bill_id}")
async def update_service_bill(bill_id: str, body: ServiceBillCreate, current_user=Depends(verify_token)):
    bill = await db.service_bills.find_one({"_id": obj_id(bill_id)})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")

    items_out = []
    for item in (body.items or []):
        if item.complimentary:
            line = {"gross":0.0,"discount":0.0,"taxable":0.0,"cgst":0.0,"sgst":0.0,"gst_total":0.0,"total":0.0}
            item_disc = 0.0
        else:
            item_disc = max(0.0, float(item.discount or 0))
            line = calc_gst_line(item.unit_price, item.qty, item.gst_rate, item_disc)
        items_out.append({
            "description":   item.description,
            "part_number":   item.part_number or "",
            "hsn_code":      item.hsn_code or "",
            "qty":           item.qty,
            "unit_price":    0 if item.complimentary else item.unit_price,
            "mrp":           item.unit_price,
            "gst_rate":      0 if item.complimentary else item.gst_rate,
            "discount":      0 if item.complimentary else item_disc,
            "complimentary": bool(item.complimentary),
            **line,
        })

    totals = calc_bill_totals([
        {"unit_price": i["unit_price"], "qty": i["qty"], "gst_rate": i["gst_rate"], "discount": i.get("discount", 0)}
        for i in items_out
    ])
    # Discount: use new value if provided, else keep existing
    discount_raw = body.discount if body.discount is not None else bill.get("discount", 0)
    discount = max(0.0, min(float(discount_raw or 0), totals["grand_total"]))
    net_total = round(totals["grand_total"] - discount, 2)

    update = {
        "items":           items_out,
        "payment_mode":    body.payment_mode or bill.get("payment_mode", "Cash"),
        "updated_at":      utcnow(),
        "discount":        discount,
        "net_total":       net_total,
        "amount_in_words": amount_in_words(net_total),
        **totals,
    }
    await db.service_bills.update_one({"_id": obj_id(bill_id)}, {"$set": update})
    await db.service_jobs.update_one(
        {"_id": obj_id(bill["job_id"])},
        {"$set": {"grand_total": net_total}}
    )
    updated = await db.service_bills.find_one({"_id": obj_id(bill_id)})
    return JSONResponse(content=oid(updated))

@api_router.delete("/service-bills/{bill_id}")
async def delete_service_bill(bill_id: str, current_user=Depends(require_admin)):
    bill = await db.service_bills.find_one({"_id": obj_id(bill_id)})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    # Restore stock for all parts used in this bill
    for item in bill.get("items", []):
        part_number = item.get("part_number", "").strip()
        qty = item.get("qty", 0)
        if part_number and qty > 0:
            await db.spare_parts.update_one(
                {"part_number": part_number},
                {"$inc": {"stock": qty}}
            )
    await db.service_bills.delete_one({"_id": obj_id(bill_id)})
    await db.service_jobs.update_one(
        {"_id": obj_id(bill["job_id"])},
        {"$set": {"status": "in_progress"}, "$unset": {"bill_number": ""}}
    )
    return {"message": "Deleted"}


# ═══════════════════════════════════════════════════════════════════════════════
#  VENDORS  (spare-parts suppliers)
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/vendors")
async def list_vendors(
    search: Optional[str] = Query(None),
    current_user=Depends(verify_token),
):
    query: dict = {}
    if search:
        esc = re.escape(search)
        query["$or"] = [
            {"name":  {"$regex": esc, "$options": "i"}},
            {"gstin": {"$regex": esc, "$options": "i"}},
            {"phone": {"$regex": esc, "$options": "i"}},
        ]
    docs = await db.vendors.find(query).sort("name", 1).to_list(500)
    return oids(docs)

@api_router.post("/vendors", status_code=201)
async def create_vendor(body: VendorCreate, current_user=Depends(verify_token)):
    doc = body.dict()
    doc["name"] = doc["name"].strip()
    if not doc["name"]:
        raise HTTPException(status_code=400, detail="Vendor name required")
    doc["gstin"] = (doc.get("gstin") or "").strip().upper()
    # Soft-uniqueness: warn if same name or GSTIN already exists, but don't block
    doc["created_at"] = utcnow().isoformat()
    result = await db.vendors.insert_one(doc)
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)
    return doc

@api_router.get("/vendors/{vendor_id}")
async def get_vendor(vendor_id: str, current_user=Depends(verify_token)):
    doc = await db.vendors.find_one({"_id": obj_id(vendor_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return oid(doc)

@api_router.put("/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, body: VendorUpdate, current_user=Depends(verify_token)):
    update = {k: v for k, v in body.dict().items() if v is not None}
    if "name" in update:
        update["name"] = update["name"].strip()
    if "gstin" in update:
        update["gstin"] = (update["gstin"] or "").strip().upper()
    if not update:
        raise HTTPException(status_code=400, detail="Nothing to update")
    r = await db.vendors.update_one({"_id": obj_id(vendor_id)}, {"$set": update})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return oid(await db.vendors.find_one({"_id": obj_id(vendor_id)}))

@api_router.delete("/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str, current_user=Depends(require_admin)):
    # Block delete if any purchase bills reference this vendor
    ref_count = await db.purchase_bills.count_documents({"vendor_id": vendor_id})
    if ref_count > 0:
        raise HTTPException(status_code=409,
            detail=f"Cannot delete: {ref_count} purchase bill(s) reference this vendor")
    r = await db.vendors.delete_one({"_id": obj_id(vendor_id)})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"deleted": True}

@api_router.get("/vendors/{vendor_id}/summary")
async def vendor_summary(vendor_id: str, current_user=Depends(verify_token)):
    """Total spend + bill count + last-bill date for this vendor."""
    pipeline = [
        {"$match": {"vendor_id": vendor_id}},
        {"$group": {
            "_id": None,
            "total_spend":  {"$sum": "$grand_total"},
            "bill_count":   {"$sum": 1},
            "last_bill":    {"$max": "$bill_date"},
        }},
    ]
    result = await db.purchase_bills.aggregate(pipeline).to_list(1)
    if not result:
        return {"total_spend": 0, "bill_count": 0, "last_bill": None}
    r = result[0]
    r.pop("_id", None)
    return r


# ═══════════════════════════════════════════════════════════════════════════════
#  PURCHASE BILLS  (incoming stock from vendors)
# ═══════════════════════════════════════════════════════════════════════════════

def _pb_compute_line(item: dict) -> dict:
    """Compute taxable_amt, cgst, sgst, line_total for one bill line.
    Rate is the per-unit price BEFORE discount. GST is added on top of the
    discounted taxable amount (standard trade invoice convention)."""
    qty          = float(item.get("qty") or 0)
    rate         = float(item.get("rate") or 0)
    discount_pct = float(item.get("discount_pct") or 0)
    gst_pct      = float(item.get("gst_pct") or 0)
    gross        = qty * rate
    discount     = gross * (discount_pct / 100.0)
    taxable      = round(gross - discount, 2)
    gst_total    = round(taxable * (gst_pct / 100.0), 2)
    cgst         = round(gst_total / 2, 2)
    sgst         = round(gst_total - cgst, 2)  # keep balance exact
    line_total   = round(taxable + cgst + sgst, 2)
    return {**item,
            "taxable_amt": taxable, "cgst": cgst, "sgst": sgst,
            "line_total":  line_total}

def _pb_compute_totals(items: List[dict], round_off: float = 0.0) -> dict:
    subtotal    = round(sum(i.get("taxable_amt", 0) for i in items), 2)
    total_cgst  = round(sum(i.get("cgst", 0)        for i in items), 2)
    total_sgst  = round(sum(i.get("sgst", 0)        for i in items), 2)
    grand       = round(subtotal + total_cgst + total_sgst + (round_off or 0), 2)
    return {"subtotal": subtotal, "total_cgst": total_cgst,
            "total_sgst": total_sgst, "grand_total": grand,
            "round_off": round(round_off or 0, 2)}

async def _pb_apply_stock_deltas(items: List[dict], vendor_id: str, vendor_name: str,
                                  bill_date: str, sign: int = +1):
    """Apply +qty (sign=+1) or -qty (sign=-1) to stock for each line.
    Also update aliases + last_purchase_rate on the linked part."""
    for it in items:
        part_id = it.get("part_id")
        if not part_id: continue
        qty     = float(it.get("qty") or 0) * sign
        # Net cost per unit AFTER discount, INCLUDING GST (what you actually paid)
        line_total = float(it.get("line_total") or 0)
        base_qty   = float(it.get("qty") or 0)
        net_unit   = round(line_total / base_qty, 2) if base_qty else 0
        try:
            oid_p = obj_id(part_id)
        except HTTPException:
            continue
        update: dict = {"$inc": {"stock": int(qty)}}
        # Only update purchase_price/aliases/last_purchase_rate on the ADD path
        if sign > 0:
            update["$set"] = {
                "last_purchase_rate":     net_unit,
                "last_purchase_date":     bill_date,
                "last_purchase_vendor":   vendor_name,
            }
            # Store this vendor's naming as an alias (upsert-style: add if new)
            alias_entry = {
                "vendor_id":    vendor_id,
                "vendor_name":  vendor_name,
                "part_number":  (it.get("part_number") or "").strip(),
                "part_name":    (it.get("part_name")   or "").strip(),
            }
            # Only add alias if part_number or part_name is non-empty AND not already present
            if alias_entry["part_number"] or alias_entry["part_name"]:
                await db.spare_parts.update_one(
                    {"_id": oid_p, "aliases": {
                        "$not": {"$elemMatch": {
                            "vendor_id":   vendor_id,
                            "part_number": alias_entry["part_number"],
                        }}
                    }},
                    {"$push": {"aliases": alias_entry}},
                )
            # Optional caller-driven sale-price change
            if it.get("new_sale_price") is not None:
                update["$set"]["selling_price"] = float(it["new_sale_price"])
        await db.spare_parts.update_one({"_id": oid_p}, update)

async def _pb_create_new_parts(items: List[dict], current_user: dict) -> List[dict]:
    """For lines with part_id=None: create a fresh spare_part and stamp part_id on the item.
    Returns the mutated items list."""
    out = []
    for it in items:
        item = dict(it)
        if item.get("part_id"):
            out.append(item); continue
        # Auto-generate part number if vendor didn't provide one
        pn = (item.get("part_number") or "").strip()
        if not pn:
            pn = await next_sequence("part")
        else:
            # Collision safety: append seq if number already taken
            if await db.spare_parts.find_one({"part_number": pn}):
                pn = f"{pn}-{await next_sequence('part')}"
        line_total = float(item.get("line_total") or 0)
        base_qty   = float(item.get("qty") or 0)
        net_unit   = round(line_total / base_qty, 2) if base_qty else 0
        # Default selling price = net_unit * 1.25 (25% margin) unless caller overrode
        selling  = item.get("new_part_selling_price")
        if selling is None:
            selling = round(net_unit * 1.25, 2)
        new_part = {
            "part_number":     pn,
            "name":            (item.get("part_name") or "").strip(),
            "category":        "",
            "brand":           "",
            "compatible_with": [],
            "stock":           0,   # will get incremented by _pb_apply_stock_deltas
            "reorder_level":   5,
            "purchase_price":  net_unit,
            "selling_price":   float(selling),
            "gst_rate":        float(item.get("gst_pct") or 18),
            "hsn_code":        (item.get("hsn") or "").strip(),
            "location":        "",
            "aliases":         [],
            "created_at":      utcnow().isoformat(),
            "created_by":      current_user.get("username", ""),
        }
        res = await db.spare_parts.insert_one(new_part)
        item["part_id"] = str(res.inserted_id)
        out.append(item)
    return out

@api_router.get("/purchase-bills")
async def list_purchase_bills(
    vendor_id: Optional[str] = Query(None),
    search:    Optional[str] = Query(None),
    p=Depends(paginate_params),
    current_user=Depends(verify_token),
):
    query: dict = {}
    if vendor_id: query["vendor_id"] = vendor_id
    if search:
        esc = re.escape(search)
        query["$or"] = [
            {"bill_number": {"$regex": esc, "$options": "i"}},
            {"vendor_name": {"$regex": esc, "$options": "i"}},
        ]
    docs  = await db.purchase_bills.find(query).skip(p["skip"]).limit(p["limit"]).sort("created_at", -1).to_list(p["limit"])
    total = await db.purchase_bills.count_documents(query)
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})

@api_router.get("/purchase-bills/{bill_id}")
async def get_purchase_bill(bill_id: str, current_user=Depends(verify_token)):
    doc = await db.purchase_bills.find_one({"_id": obj_id(bill_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Purchase bill not found")
    return oid(doc)

@api_router.post("/purchase-bills", status_code=201)
async def create_purchase_bill(body: PurchaseBillCreate, current_user=Depends(verify_token)):
    # Fetch vendor (denorm name onto bill)
    vendor = await db.vendors.find_one({"_id": obj_id(body.vendor_id)})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")

    if not body.items:
        raise HTTPException(status_code=400, detail="Bill must have at least one line")

    # Duplicate check: same vendor + same bill_number
    bill_num = body.bill_number.strip()
    if not bill_num:
        raise HTTPException(status_code=400, detail="Bill number required")
    dup = await db.purchase_bills.find_one({"vendor_id": body.vendor_id, "bill_number": bill_num})
    if dup:
        raise HTTPException(status_code=409, detail=f"Bill number {bill_num} already recorded for this vendor")

    # Compute per-line taxes + totals
    items = [_pb_compute_line(i.dict()) for i in body.items]
    # Auto-create any new parts (part_id=None); returns items with part_id filled
    items = await _pb_create_new_parts(items, current_user)
    totals = _pb_compute_totals(items, body.round_off or 0)

    vendor_name = vendor.get("name", "")
    doc = {
        "vendor_id":       body.vendor_id,
        "vendor_name":     vendor_name,
        "vendor_gstin":    vendor.get("gstin", ""),
        "bill_number":     bill_num,
        "bill_date":       body.bill_date.strip(),
        "place_of_supply": body.place_of_supply or "",
        "items":           items,
        "notes":           body.notes or "",
        **totals,
        "created_at":      utcnow().isoformat(),
        "created_by":      current_user.get("username", ""),
    }
    result = await db.purchase_bills.insert_one(doc)
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)

    # Apply stock deltas + aliases + last_purchase_rate
    await _pb_apply_stock_deltas(items, body.vendor_id, vendor_name, body.bill_date.strip(), sign=+1)
    return doc

@api_router.put("/purchase-bills/{bill_id}")
async def update_purchase_bill(bill_id: str, body: PurchaseBillUpdate,
                               current_user=Depends(require_admin)):
    """Owner-only. Reverses old stock deltas, then applies new."""
    existing = await db.purchase_bills.find_one({"_id": obj_id(bill_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Purchase bill not found")

    updates = {k: v for k, v in body.dict().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Nothing to update")

    # Reverse old stock
    await _pb_apply_stock_deltas(
        existing.get("items") or [],
        existing.get("vendor_id"),
        existing.get("vendor_name", ""),
        existing.get("bill_date", ""),
        sign=-1,
    )

    # Rebuild items if provided
    if "items" in updates:
        new_items = [_pb_compute_line(dict(i) if isinstance(i, dict) else i.dict())
                     for i in updates["items"]]
        new_items = await _pb_create_new_parts(new_items, current_user)
        updates["items"] = new_items
        round_off = updates.get("round_off", existing.get("round_off", 0))
        updates.update(_pb_compute_totals(new_items, round_off or 0))

    updates["updated_at"] = utcnow().isoformat()
    updates["updated_by"] = current_user.get("username", "")
    await db.purchase_bills.update_one({"_id": obj_id(bill_id)}, {"$set": updates})

    # Re-fetch fresh doc and re-apply stock in +direction
    fresh = await db.purchase_bills.find_one({"_id": obj_id(bill_id)})
    await _pb_apply_stock_deltas(
        fresh.get("items") or [],
        fresh.get("vendor_id"),
        fresh.get("vendor_name", ""),
        fresh.get("bill_date", ""),
        sign=+1,
    )
    return oid(fresh)

@api_router.delete("/purchase-bills/{bill_id}")
async def delete_purchase_bill(bill_id: str, current_user=Depends(require_admin)):
    """Owner-only. Reverses stock, then deletes."""
    existing = await db.purchase_bills.find_one({"_id": obj_id(bill_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Purchase bill not found")
    await _pb_apply_stock_deltas(
        existing.get("items") or [],
        existing.get("vendor_id"),
        existing.get("vendor_name", ""),
        existing.get("bill_date", ""),
        sign=-1,
    )
    await db.purchase_bills.delete_one({"_id": obj_id(bill_id)})
    return {"deleted": True}


# ─── Parts search: vendor-aware fuzzy match (for purchase-bill autocomplete) ──
@api_router.get("/parts/search-alias")
async def parts_search_alias(
    q:         str          = Query(..., min_length=1),
    vendor_id: Optional[str] = Query(None),
    limit:     int          = Query(20, ge=1, le=100),
    current_user=Depends(verify_token),
):
    """Search parts by name, part_number, OR alias.part_number (vendor-scoped).
    Returns up to `limit` matches; vendor-alias hits ranked first when vendor_id passed."""
    esc = re.escape(q)
    # Alias-match branch (vendor-scoped when vendor_id provided)
    alias_branch = {"aliases": {"$elemMatch": {
        "$or": [
            {"part_number": {"$regex": esc, "$options": "i"}},
            {"part_name":   {"$regex": esc, "$options": "i"}},
        ]
    }}}
    if vendor_id:
        alias_branch["aliases"]["$elemMatch"]["vendor_id"] = vendor_id
    query = {"$or": [
        {"name":        {"$regex": esc, "$options": "i"}},
        {"part_number": {"$regex": esc, "$options": "i"}},
        alias_branch,
    ]}
    docs = await db.spare_parts.find(query).limit(limit).to_list(limit)
    return oids(docs)


# ═══════════════════════════════════════════════════════════════════════════════
#  SPARE PARTS
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/parts")
async def list_parts(
    category:    Optional[str]  = Query(None),
    brand:       Optional[str]  = Query(None),
    low_stock:   Optional[bool] = Query(None),
    out_of_stock:Optional[bool] = Query(None),
    search:      Optional[str]  = Query(None),
    p=Depends(paginate_params),
    current_user=Depends(verify_token),
):
    query: dict = {}
    if category: query["category"] = category
    if brand:    query["brand"]    = brand
    if low_stock:
        query["$expr"] = {"$and":[{"$gt":["$stock",0]},{"$lte":["$stock","$reorder_level"]}]}
    if out_of_stock:
        query["stock"] = 0
    if search:
        s = re.escape(search)
        query["$or"] = [
            {"name":        {"$regex": s, "$options":"i"}},
            {"part_number": {"$regex": s, "$options":"i"}},
            {"brand":       {"$regex": s, "$options":"i"}},
        ]
    docs  = await db.spare_parts.find(query).skip(p["skip"]).limit(p["limit"]).sort("name",1).to_list(p["limit"])
    total = await db.spare_parts.count_documents(query)
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})

@api_router.post("/parts", status_code=201)
async def create_part(body: SparePartCreate, current_user=Depends(verify_token)):
    doc = body.dict()
    pn = (doc.get("part_number") or "").strip()
    if not pn:
        pn = await next_sequence("part")
    else:
        if await db.spare_parts.find_one({"part_number": pn}):
            raise HTTPException(status_code=409, detail="Part number already exists")
    doc["part_number"] = pn
    doc["created_at"]  = utcnow().isoformat()
    result = await db.spare_parts.insert_one(doc)
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)
    return doc

@api_router.get("/parts/stats/summary")
async def parts_stats(current_user=Depends(verify_token)):
    pipeline_val = [{"$group": {"_id":None,"stock_value":{"$sum":{"$multiply":["$purchase_price","$stock"]}},"selling_value":{"$sum":{"$multiply":["$selling_price","$stock"]}},"total_skus":{"$sum":1}}}]
    result = await db.spare_parts.aggregate(pipeline_val).to_list(1)
    stats  = result[0] if result else {}
    low, out = await asyncio.gather(
        db.spare_parts.count_documents({"$expr":{"$and":[{"$gt":["$stock",0]},{"$lte":["$stock","$reorder_level"]}]}}),
        db.spare_parts.count_documents({"stock": 0}),
    )
    return {"total_skus":stats.get("total_skus",0),"low_stock":low,"out_of_stock":out,"stock_value":round(stats.get("stock_value",0),2),"selling_value":round(stats.get("selling_value",0),2)}

@api_router.get("/parts/low-stock")
async def low_stock_parts(current_user=Depends(verify_token)):
    docs = await db.spare_parts.aggregate([{"$match":{"$expr":{"$and":[{"$gt":["$stock",0]},{"$lte":["$stock","$reorder_level"]}]}}}]).to_list(None)
    return oids(docs)

@api_router.get("/parts/out-of-stock")
async def out_of_stock_parts(current_user=Depends(verify_token)):
    docs = await db.spare_parts.find({"stock":0}).to_list(None)
    return oids(docs)

@api_router.get("/parts/{part_id}")
async def get_part(part_id: str, current_user=Depends(verify_token)):
    doc = await db.spare_parts.find_one({"_id": obj_id(part_id)})
    if not doc:
        doc = await db.spare_parts.find_one({"part_number": part_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Part not found")
    return oid(doc)

@api_router.put("/parts/{part_id}")
async def update_part(part_id: str, body: SparePartUpdate, current_user=Depends(verify_token)):
    update = {k: v for k, v in body.dict().items() if v is not None}
    await db.spare_parts.update_one({"_id": obj_id(part_id)}, {"$set": update})
    return oid(await db.spare_parts.find_one({"_id": obj_id(part_id)}))

# PATCH 5: adjust-stock by part_number (for service bill modal) ───────────────
@api_router.post("/parts/{part_number}/adjust-stock-by-number")
async def adjust_stock_by_number(
    part_number: str,
    body: StockAdjust,
    current_user=Depends(verify_token),
):
    """Adjust stock by part_number string — used by service bill modal for deduction/restore."""
    part = await db.spare_parts.find_one({"part_number": part_number})
    if not part:
        raise HTTPException(status_code=404, detail=f"Part '{part_number}' not found")
    current_stock = part.get("stock") or 0
    action = body.action or "add"
    if action == "subtract":
        new_stock = max(0, current_stock - abs(body.qty))
    else:
        new_stock = current_stock + abs(body.qty)
    await db.spare_parts.update_one(
        {"part_number": part_number},
        {"$set": {"stock": new_stock}, "$push": {"stock_log": {
            "qty": body.qty, "action": action, "reason": body.reason or "service_bill",
            "new_stock": new_stock, "date": utcnow().isoformat(),
        }}}
    )
    return JSONResponse(content={"part_number": part_number, "old_stock": current_stock, "new_stock": new_stock})

@api_router.post("/parts/{part_id}/adjust-stock")
async def adjust_stock(part_id: str, body: StockAdjust, current_user=Depends(verify_token)):
    part = await db.spare_parts.find_one({"_id": obj_id(part_id)})
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    # PATCH 4: respect action field (add/subtract), fall back to qty sign
    action = body.action or "add"
    if action == "subtract":
        new_stock = max(0, part["stock"] - abs(body.qty))
    else:
        new_stock = max(0, part["stock"] + body.qty)
    await db.spare_parts.update_one(
        {"_id": obj_id(part_id)},
        {"$set": {"stock": new_stock}, "$push": {"stock_log": {
            "qty": body.qty, "action": action, "reason": body.reason,
            "new_stock": new_stock, "date": utcnow().isoformat()
        }}}
    )
    return {"part_id": part_id, "new_stock": new_stock}

@api_router.delete("/parts/{part_id}")
async def delete_part(part_id: str, current_user=Depends(require_admin)):
    result = await db.spare_parts.delete_one({"_id": obj_id(part_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Part not found")
    return {"message": "Deleted"}


# ═══════════════════════════════════════════════════════════════════════════════
#  PARTS SALES (Counter bills with GST — requires part_id)
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/parts-sales")
async def list_parts_sales(
    search: Optional[str] = Query(None),
    p=Depends(paginate_params),
    current_user=Depends(verify_token),
):
    query: dict = {}
    if search:
        s = re.escape(search)
        query["$or"] = [
            {"bill_number":     {"$regex": s, "$options":"i"}},
            {"customer_name":   {"$regex": s, "$options":"i"}},
            {"customer_mobile": {"$regex": s, "$options":"i"}},
        ]
    docs  = await db.parts_sales.find(query).skip(p["skip"]).limit(p["limit"]).sort("created_at",-1).to_list(p["limit"])
    total = await db.parts_sales.count_documents(query)
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})

@api_router.post("/parts-sales", status_code=201)
async def create_parts_sale(body: PartsSaleCreate, current_user=Depends(verify_token)):
    if not body.items:
        raise HTTPException(status_code=400, detail="At least one item required")
    items_out = []
    for item in body.items:
        part = await db.spare_parts.find_one({"_id": obj_id(item.part_id)})
        if not part:
            raise HTTPException(status_code=404, detail=f"Part {item.part_id} not found")
        # Atomic guarded decrement — no TOCTOU race
        dec_res = await db.spare_parts.update_one(
            {"_id": obj_id(item.part_id), "stock": {"$gte": item.qty}},
            {"$inc": {"stock": -item.qty}},
        )
        if dec_res.matched_count == 0:
            fresh = await db.spare_parts.find_one({"_id": obj_id(item.part_id)}, {"stock": 1, "name": 1})
            have = (fresh or {}).get("stock", 0)
            raise HTTPException(status_code=409, detail=f"Insufficient stock for {part['name']} (have {have}, need {item.qty})")
        line = calc_gst_line(item.unit_price, item.qty, item.gst_rate)
        items_out.append({"part_id":item.part_id,"part_number":item.part_number,"name":item.name,"hsn_code":item.hsn_code or part.get("hsn_code",""),"qty":item.qty,"unit_price":item.unit_price,"gst_rate":item.gst_rate,**line})
    totals  = calc_bill_totals([{"unit_price":i["unit_price"],"qty":i["qty"],"gst_rate":i["gst_rate"]} for i in items_out])
    bill_no = await next_sequence("part_bill")
    doc = {"bill_number":bill_no,"customer_name":body.customer_name or "","customer_mobile":body.customer_mobile or "","items":items_out,**totals,"amount_in_words":amount_in_words(totals["grand_total"]),"payment_mode":body.payment_mode or "Cash","sold_by":body.sold_by or current_user.get("name",""),"sale_date":utcnow().strftime("%d %b %Y"),"notes":body.notes or "","created_at":utcnow().isoformat()}
    result = await db.parts_sales.insert_one(doc)
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)
    return doc

@api_router.get("/parts-sales/{bill_id}")
async def get_parts_sale(bill_id: str, current_user=Depends(verify_token)):
    doc = await db.parts_sales.find_one({"_id": obj_id(bill_id)})
    if not doc:
        doc = await db.parts_sales.find_one({"bill_number": bill_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Bill not found")
    return oid(doc)

@api_router.delete("/parts-sales/{bill_id}")
async def delete_parts_sale(bill_id: str, current_user=Depends(require_admin)):
    bill = await db.parts_sales.find_one({"_id": obj_id(bill_id)})
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    for item in bill.get("items", []):
        await db.spare_parts.update_one({"_id": obj_id(item["part_id"])}, {"$inc": {"stock": item["qty"]}})
    await db.parts_sales.delete_one({"_id": obj_id(bill_id)})
    await _sync_counter("part_bill", "parts_sales", "bill_number")
    return {"message": "Deleted — stock restored"}


# ═══════════════════════════════════════════════════════════════════════════════
#  PARTS BILLS — PATCH 6 (walk-in counter, vehicle number optional)
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/parts-bills")
async def list_parts_bills(
    search: Optional[str] = Query(None),
    p=Depends(paginate_params),
    current_user=Depends(verify_token),
):
    query: dict = {}
    if search:
        s = re.escape(search)
        query["$or"] = [
            {"bill_number":      {"$regex": s, "$options":"i"}},
            {"customer_name":    {"$regex": s, "$options":"i"}},
            {"customer_mobile":  {"$regex": s, "$options":"i"}},
            {"customer_vehicle": {"$regex": s, "$options":"i"}},
        ]
    docs  = await db.parts_bills.find(query).skip(p["skip"]).limit(p["limit"]).sort("created_at",-1).to_list(p["limit"])
    total = await db.parts_bills.count_documents(query)
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})


@api_router.post("/parts-bills", status_code=201)
async def create_parts_bill(body: PartsBillCreate, current_user=Depends(verify_token)):
    if not body.items:
        raise HTTPException(status_code=400, detail="At least one item required")

    items_out = []
    for item in body.items:
        # Resolve part by part_id first, fall back to part_number
        part = None
        if item.part_id:
            part = await db.spare_parts.find_one({"_id": obj_id(item.part_id)})
        if not part and item.part_number:
            part = await db.spare_parts.find_one({"part_number": item.part_number})

        if part:
            # Atomic guarded decrement — no TOCTOU race; log entry via $push in same op
            dec_res = await db.spare_parts.update_one(
                {"_id": part["_id"], "stock": {"$gte": item.qty}},
                {"$inc": {"stock": -item.qty}, "$push": {"stock_log": {
                    "qty": -item.qty, "action": "subtract",
                    "reason": "complimentary" if item.complimentary else "parts_bill",
                    "date": utcnow().isoformat(),
                }}}
            )
            if dec_res.matched_count == 0:
                fresh = await db.spare_parts.find_one({"_id": part["_id"]}, {"stock": 1})
                have = (fresh or {}).get("stock", 0)
                raise HTTPException(
                    status_code=409,
                    detail=f"Insufficient stock for {item.name} (have {have}, need {item.qty})"
                )

        if item.complimentary:
            line = {"gross":0.0,"discount":0.0,"taxable":0.0,"cgst":0.0,"sgst":0.0,"gst_total":0.0,"total":0.0}
            item_disc = 0.0
        else:
            item_disc = max(0.0, float(item.discount or 0))
            line = calc_gst_line(item.unit_price, item.qty, item.gst_rate, item_disc)
        items_out.append({
            "part_id":       str(part["_id"]) if part else (item.part_id or ""),
            "part_number":   item.part_number or (part.get("part_number","") if part else ""),
            "name":          item.name,
            "hsn_code":      item.hsn_code or (part.get("hsn_code","8714") if part else "8714"),
            "qty":           item.qty,
            "unit_price":    0 if item.complimentary else item.unit_price,
            "mrp":           item.unit_price,
            "gst_rate":      0 if item.complimentary else item.gst_rate,
            "discount":      0 if item.complimentary else item_disc,
            "complimentary": bool(item.complimentary),
            **line,
        })

    totals  = calc_bill_totals([{"unit_price":i["unit_price"],"qty":i["qty"],"gst_rate":i["gst_rate"],"discount":i.get("discount",0)} for i in items_out])
    # Bill-level discount off grand_total (post-GST). Clamped 0..grand_total.
    bill_discount = max(0.0, min(float(body.discount or 0), totals["grand_total"]))
    net_total = round(totals["grand_total"] - bill_discount, 2)
    bill_no = await next_sequence("part_bill")   # shared counter with parts_sales

    doc = {
        "bill_number":      bill_no,
        "customer_name":    body.customer_name or "",
        "customer_mobile":  body.customer_mobile or "",
        "customer_vehicle": body.customer_vehicle or "",
        "payment_mode":     body.payment_mode or "Cash",
        "items":            items_out,
        "discount":         bill_discount,
        "net_total":        net_total,
        "amount_in_words":  amount_in_words(net_total),
        "sold_by":          current_user.get("name",""),
        "bill_date":        utcnow().strftime("%d %b %Y"),
        "created_at":       utcnow().isoformat(),
        **totals,
    }
    res     = await db.parts_bills.insert_one(doc)
    created = await db.parts_bills.find_one({"_id": res.inserted_id})
    return JSONResponse(content=oid(created), status_code=201)


@api_router.get("/parts-bills/{bill_id}")
async def get_parts_bill(bill_id: str, current_user=Depends(verify_token)):
    doc = await db.parts_bills.find_one({"_id": obj_id(bill_id)})
    if not doc:
        doc = await db.parts_bills.find_one({"bill_number": bill_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Parts bill not found")
    return oid(doc)


@api_router.put("/parts-bills/{bill_id}")
async def update_parts_bill(bill_id: str, body: PartsBillUpdate, current_user=Depends(verify_token)):
    """Update parts bill. If items changed, restore old stock then re-deduct new stock.
    Discount recomputed per-line + bill-level. Complimentary rows ignore discount."""
    bill = await db.parts_bills.find_one({"_id": obj_id(bill_id)})
    if not bill:
        raise HTTPException(status_code=404, detail="Parts bill not found")

    update: dict = {"updated_at": utcnow().isoformat()}
    if body.customer_name    is not None: update["customer_name"]    = body.customer_name
    if body.customer_mobile  is not None: update["customer_mobile"]  = body.customer_mobile
    if body.customer_vehicle is not None: update["customer_vehicle"] = body.customer_vehicle
    if body.payment_mode     is not None: update["payment_mode"]     = body.payment_mode

    if body.items is not None:
        # Restore stock from old items first
        for old_it in bill.get("items", []):
            query = {}
            if old_it.get("part_id"):
                try:
                    from bson import ObjectId as _OID
                    query = {"_id": _OID(old_it["part_id"])}
                except Exception:
                    query = {"part_number": old_it.get("part_number","")}
            elif old_it.get("part_number"):
                query = {"part_number": old_it["part_number"]}
            if query and old_it.get("qty", 0) > 0:
                await db.spare_parts.update_one(query, {"$inc": {"stock": old_it["qty"]}})

        # Build new items + deduct stock
        items_out = []
        for item in body.items:
            part = None
            if item.part_id:
                part = await db.spare_parts.find_one({"_id": obj_id(item.part_id)})
            if not part and item.part_number:
                part = await db.spare_parts.find_one({"part_number": item.part_number})
            if part:
                # Atomic guarded decrement — no TOCTOU race
                dec_res = await db.spare_parts.update_one(
                    {"_id": part["_id"], "stock": {"$gte": item.qty}},
                    {"$inc": {"stock": -item.qty}, "$push": {"stock_log": {
                        "qty": -item.qty, "action": "subtract",
                        "reason": "complimentary" if item.complimentary else "parts_bill_edit",
                        "date": utcnow().isoformat(),
                    }}}
                )
                if dec_res.matched_count == 0:
                    fresh = await db.spare_parts.find_one({"_id": part["_id"]}, {"stock": 1})
                    have = (fresh or {}).get("stock", 0)
                    raise HTTPException(
                        status_code=409,
                        detail=f"Insufficient stock for {item.name} (have {have}, need {item.qty})"
                    )

            if item.complimentary:
                line = {"gross":0.0,"discount":0.0,"taxable":0.0,"cgst":0.0,"sgst":0.0,"gst_total":0.0,"total":0.0}
                item_disc = 0.0
            else:
                item_disc = max(0.0, float(item.discount or 0))
                line = calc_gst_line(item.unit_price, item.qty, item.gst_rate, item_disc)
            items_out.append({
                "part_id":       str(part["_id"]) if part else (item.part_id or ""),
                "part_number":   item.part_number or (part.get("part_number","") if part else ""),
                "name":          item.name,
                "hsn_code":      item.hsn_code or (part.get("hsn_code","8714") if part else "8714"),
                "qty":           item.qty,
                "unit_price":    0 if item.complimentary else item.unit_price,
                "mrp":           item.unit_price,
                "gst_rate":      0 if item.complimentary else item.gst_rate,
                "discount":      0 if item.complimentary else item_disc,
                "complimentary": bool(item.complimentary),
                **line,
            })
        totals = calc_bill_totals([{"unit_price":i["unit_price"],"qty":i["qty"],"gst_rate":i["gst_rate"],"discount":i.get("discount",0)} for i in items_out])
        update["items"] = items_out
        update.update(totals)
    else:
        totals = {k: bill.get(k, 0) for k in ("subtotal","gst_total","line_discount","grand_total","gst_breakup")}

    # Bill-level discount
    disc_raw = body.discount if body.discount is not None else bill.get("discount", 0)
    grand = update.get("grand_total", bill.get("grand_total", 0))
    bill_discount = max(0.0, min(float(disc_raw or 0), float(grand)))
    net_total = round(float(grand) - bill_discount, 2)
    update["discount"]        = bill_discount
    update["net_total"]       = net_total
    update["amount_in_words"] = amount_in_words(net_total)

    await db.parts_bills.update_one({"_id": obj_id(bill_id)}, {"$set": update})
    updated = await db.parts_bills.find_one({"_id": obj_id(bill_id)})
    return JSONResponse(content=oid(updated))


@api_router.delete("/parts-bills/{bill_id}")
async def delete_parts_bill(bill_id: str, current_user=Depends(require_admin)):
    bill = await db.parts_bills.find_one({"_id": obj_id(bill_id)})
    if not bill:
        raise HTTPException(status_code=404, detail="Parts bill not found")
    # Restore stock for each item
    for item in bill.get("items", []):
        query = {}
        if item.get("part_id"):
            try:
                from bson import ObjectId as _OID
                query = {"_id": _OID(item["part_id"])}
            except Exception:
                query = {"part_number": item.get("part_number","")}
        elif item.get("part_number"):
            query = {"part_number": item["part_number"]}
        if query:
            await db.spare_parts.update_one(query, {"$inc": {"stock": item["qty"]}})
    await db.parts_bills.delete_one({"_id": obj_id(bill_id)})
    return {"message": "Deleted — stock restored"}


# ═══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD STATS
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/dashboard/stats")
async def dashboard_stats(current_user=Depends(verify_token)):
    today = utcnow().strftime("%d %b %Y")
    (vehicles_in_stock, vehicles_sold_today, jobs_pending, jobs_in_progress, jobs_ready, customers_total, parts_low, parts_out, sales_today_count) = await asyncio.gather(
        db.vehicles.count_documents({"status":"in_stock"}),
        db.sales.count_documents({"sale_date":today}),
        db.service_jobs.count_documents({"status":"pending"}),
        db.service_jobs.count_documents({"status":"in_progress"}),
        db.service_jobs.count_documents({"status":"ready"}),
        db.customers.count_documents({}),
        db.spare_parts.count_documents({"$expr":{"$and":[{"$gt":["$stock",0]},{"$lte":["$stock","$reorder_level"]}]}}),
        db.spare_parts.count_documents({"stock":0}),
        db.sales.count_documents({"sale_date":today}),
    )
    pipeline_today_rev = [{"$match":{"sale_date":today}},{"$group":{"_id":None,"total":{"$sum":"$total_amount"}}}]
    pipeline_month_rev = [{"$match":{"sale_date":{"$regex":utcnow().strftime("%b %Y")}}},{"$group":{"_id":None,"total":{"$sum":"$total_amount"}}}]
    today_rev_r, month_rev_r = await asyncio.gather(
        db.sales.aggregate(pipeline_today_rev).to_list(1),
        db.sales.aggregate(pipeline_month_rev).to_list(1),
    )
    return {
        "vehicles":  {"in_stock":vehicles_in_stock,"sold_today":vehicles_sold_today},
        "service":   {"pending":jobs_pending,"in_progress":jobs_in_progress,"ready":jobs_ready,"active_total":jobs_pending+jobs_in_progress+jobs_ready},
        "customers": customers_total,
        "parts":     {"low_stock":parts_low,"out_of_stock":parts_out},
        "revenue":   {"today":today_rev_r[0]["total"] if today_rev_r else 0,"month":month_rev_r[0]["total"] if month_rev_r else 0},
        "sales_today_count": sales_today_count,
    }

@api_router.get("/dashboard/recent-activity")
async def recent_activity(limit: int = Query(10, le=50), current_user=Depends(verify_token)):
    sales_docs, job_docs, bill_docs = await asyncio.gather(
        db.sales.find({}).sort("created_at",-1).limit(limit).to_list(limit),
        db.service_jobs.find({}).sort("created_at",-1).limit(limit).to_list(limit),
        db.parts_sales.find({}).sort("created_at",-1).limit(limit).to_list(limit),
    )
    activity = []
    for s in sales_docs:
        activity.append({"type":"sale","time":s.get("created_at",""),"text":s.get("invoice_number",""),"sub":f"{s.get('customer_name','')} · ₹{s.get('total_amount',0):,.0f}"})
    for j in job_docs:
        activity.append({"type":"service","time":j.get("created_at",""),"text":j.get("job_number","")+" "+j.get("status",""),"sub":f"{j.get('vehicle_number','')} · {j.get('customer_name','')}"})
    for b in bill_docs:
        activity.append({"type":"parts","time":b.get("created_at",""),"text":b.get("bill_number",""),"sub":f"{b.get('customer_name','')} · ₹{b.get('grand_total',0):,.0f}"})
    activity.sort(key=lambda x: x["time"], reverse=True)
    return activity[:limit]


# ═══════════════════════════════════════════════════════════════════════════════
#  REPORTS
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/reports/revenue")
async def revenue_report(months: int = Query(6, ge=1, le=24), current_user=Depends(require_admin)):
    # sale_date can be: "08 Apr 2026", "2026-04-24", "2024-07-01 00:00:00", "01/04/2026"
    # Strategy: try multiple formats, fall back to created_at substring
    pipeline = [
        {"$addFields": {
            # Try "DD Mon YYYY" e.g. "08 Apr 2026"
            "p1": {"$dateFromString": {"dateString": "$sale_date", "format": "%d %b %Y", "onError": None, "onNull": None}},
            # Try "YYYY-MM-DD" or "YYYY-MM-DD HH:MM:SS"
            "p2": {"$dateFromString": {"dateString": {"$substr": ["$sale_date", 0, 10]}, "format": "%Y-%m-%d", "onError": None, "onNull": None}},
            # Try "DD/MM/YYYY"
            "p3": {"$dateFromString": {"dateString": "$sale_date", "format": "%d/%m/%Y", "onError": None, "onNull": None}},
        }},
        {"$addFields": {
            "parsed_date": {
                "$ifNull": ["$p1", {"$ifNull": ["$p2", {"$ifNull": ["$p3", None]}]}]
            }
        }},
        {"$addFields": {
            "month_key": {"$cond": [
                {"$ne": ["$parsed_date", None]},
                {"$dateToString": {"format": "%Y-%m", "date": "$parsed_date"}},
                {"$substr": ["$created_at", 0, 7]}
            ]}
        }},
        {"$group": {"_id": "$month_key", "sales": {"$sum": "$total_amount"}, "count": {"$sum": 1}}},
        {"$sort": {"_id": -1}}, {"$limit": months},
    ]
    sales_by_month = await db.sales.aggregate(pipeline).to_list(months)
    svc_pipeline   = [{"$addFields":{"month_key":{"$substr":["$created_at",0,7]}}},{"$group":{"_id":"$month_key","service":{"$sum":"$grand_total"},"svc_count":{"$sum":1}}},{"$sort":{"_id":-1}},{"$limit":months}]
    svc_by_month   = await db.service_bills.aggregate(svc_pipeline).to_list(months)
    parts_pipeline = [{"$addFields":{"month_key":{"$substr":["$created_at",0,7]}}},{"$group":{"_id":"$month_key","parts":{"$sum":"$grand_total"},"parts_count":{"$sum":1}}},{"$sort":{"_id":-1}},{"$limit":months}]
    parts_by_month = await db.parts_sales.aggregate(parts_pipeline).to_list(months)
    return {"sales":oids(sales_by_month),"service":oids(svc_by_month),"parts":oids(parts_by_month)}

@api_router.get("/reports/daily-closing")
async def daily_closing_report(date: Optional[str] = Query(None), current_user=Depends(require_admin)):
    """Daily closing tally.
    Matches on parsed date (not raw string) — handles 'DD Mon YYYY', 'YYYY-MM-DD',
    'DD/MM/YYYY', and 'YYYY-MM-DDTHH:MM:SS' formats stored across collections.
    Amounts grouped by payment mode.
    """
    # Accept either "DD Mon YYYY" (from frontend) or "YYYY-MM-DD" (ISO)
    target_str = date or utcnow().strftime("%d %b %Y")
    # Parse target to a single canonical date object for range match
    target_dt = None
    for fmt in ("%d %b %Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            target_dt = datetime.strptime(target_str.strip(), fmt)
            break
        except ValueError:
            continue
    if target_dt is None:
        raise HTTPException(status_code=400, detail=f"Invalid date: {target_str}")
    day_start = target_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end   = day_start + timedelta(days=1)

    async def get_totals(collection, date_field, amount_field):
        """Aggregate a collection: parse the date field defensively, match the target day,
        group by payment_mode, sum the amount field."""
        pipeline = [
            {"$addFields": {
                "p1": {"$dateFromString": {"dateString": f"${date_field}", "format": "%d %b %Y", "onError": None, "onNull": None}},
                "p2": {"$dateFromString": {"dateString": {"$substr": [f"${date_field}", 0, 10]}, "format": "%Y-%m-%d", "onError": None, "onNull": None}},
                "p3": {"$dateFromString": {"dateString": f"${date_field}", "format": "%d/%m/%Y", "onError": None, "onNull": None}},
            }},
            {"$addFields": {"parsed": {"$ifNull": ["$p1", {"$ifNull": ["$p2", "$p3"]}]}}},
            {"$match": {"parsed": {"$gte": day_start, "$lt": day_end}}},
            {"$group": {"_id": {"$toLower": "$payment_mode"}, "total": {"$sum": amount_field}}},
        ]
        return await db[collection].aggregate(pipeline).to_list(None)

    sales_r, service_r, parts_sales_r, parts_bills_r = await asyncio.gather(
        get_totals("sales",         "sale_date",   "$total_amount"),
        get_totals("service_bills", "bill_date",   "$grand_total"),
        get_totals("parts_sales",   "sale_date",   "$grand_total"),
        get_totals("parts_bills",   "created_at",  "$grand_total"),   # walk-in bills stamp created_at only
    )

    # Merge parts_sales + parts_bills under one "Parts" bucket per payment mode
    parts_r: dict = {}
    for item in parts_sales_r + parts_bills_r:
        m = item["_id"] or "unknown"
        parts_r[m] = parts_r.get(m, 0) + item["total"]
    parts_merged = [{"_id": m, "total": t} for m, t in parts_r.items()]

    summary = {}
    for source, data in [("Sales", sales_r), ("Service", service_r), ("Parts", parts_merged)]:
        for item in data:
            mode = (item["_id"] or "unknown").title()
            if mode not in summary:
                summary[mode] = {"total": 0, "Sales": 0, "Service": 0, "Parts": 0}
            summary[mode][source] += item["total"]
            summary[mode]["total"] += item["total"]

    result = [{"payment_mode": k, **v} for k, v in summary.items()]
    result.sort(key=lambda x: 0 if x["payment_mode"] == "Cash" else 1)
    return {"date": target_str, "breakdown": result, "grand_total": sum(r["total"] for r in result)}


@api_router.get("/reports/monthly-counts")
async def monthly_counts(months: int = Query(6, ge=1, le=24), current_user=Depends(verify_token)):
    """Monthly transaction counts — vehicle sales, service bills, parts bills + parts sales."""
    # Sales: parse sale_date (multi-format like GSTR1 does)
    sales_pipe = [
        {"$addFields": {
            "p1": {"$dateFromString": {"dateString": "$sale_date", "format": "%d %b %Y", "onError": None, "onNull": None}},
            "p2": {"$dateFromString": {"dateString": {"$substr": ["$sale_date", 0, 10]}, "format": "%Y-%m-%d", "onError": None, "onNull": None}},
            "p3": {"$dateFromString": {"dateString": "$sale_date", "format": "%d/%m/%Y", "onError": None, "onNull": None}},
        }},
        {"$addFields": {"parsed": {"$ifNull": ["$p1", {"$ifNull": ["$p2", "$p3"]}]}}},
        {"$addFields": {"month_key": {"$cond": [{"$ne": ["$parsed", None]},
            {"$dateToString": {"format": "%Y-%m", "date": "$parsed"}},
            {"$substr": ["$created_at", 0, 7]}]}}},
        {"$group": {"_id": "$month_key", "count": {"$sum": 1}}},
        {"$sort": {"_id": -1}}, {"$limit": months},
    ]
    # Service bills: use bill_date ("DD Mon YYYY") — same field daily-closing uses
    svc_pipe = [
        {"$addFields": {
            "p1": {"$dateFromString": {"dateString": "$bill_date", "format": "%d %b %Y", "onError": None, "onNull": None}},
        }},
        {"$addFields": {"month_key": {"$cond": [{"$ne": ["$p1", None]},
            {"$dateToString": {"format": "%Y-%m", "date": "$p1"}},
            {"$substr": ["$created_at", 0, 7]}]}}},
        {"$group": {"_id": "$month_key", "count": {"$sum": 1}}},
        {"$sort": {"_id": -1}}, {"$limit": months},
    ]
    # Parts bills (walk-in billing — primary source)
    parts_bills_pipe = [
        {"$addFields": {"month_key": {"$substr": ["$created_at", 0, 7]}}},
        {"$group": {"_id": "$month_key", "count": {"$sum": 1}}},
        {"$sort": {"_id": -1}}, {"$limit": months},
    ]
    # Parts sales (counter bills with stock deduction — secondary source)
    parts_sales_pipe = [
        {"$addFields": {"month_key": {"$substr": ["$created_at", 0, 7]}}},
        {"$group": {"_id": "$month_key", "count": {"$sum": 1}}},
        {"$sort": {"_id": -1}}, {"$limit": months},
    ]
    sales_r, svc_r, pb_r, ps_r = await asyncio.gather(
        db.sales.aggregate(sales_pipe).to_list(months),
        db.service_bills.aggregate(svc_pipe).to_list(months),
        db.parts_bills.aggregate(parts_bills_pipe).to_list(months),
        db.parts_sales.aggregate(parts_sales_pipe).to_list(months),
    )
    # Merge parts_bills + parts_sales counts per month
    parts_map: dict = {}
    for doc in pb_r:
        m = doc["_id"] or ""
        if m: parts_map[m] = parts_map.get(m, 0) + doc["count"]
    for doc in ps_r:
        m = doc["_id"] or ""
        if m: parts_map[m] = parts_map.get(m, 0) + doc["count"]
    parts_merged = sorted(
        [{"month": m, "count": c} for m, c in parts_map.items()],
        key=lambda x: x["month"], reverse=True
    )[:months]
    return {
        "sales":   [{"month": d["_id"], "count": d["count"]} for d in sales_r if d["_id"]],
        "service": [{"month": d["_id"], "count": d["count"]} for d in svc_r   if d["_id"]],
        "parts":   parts_merged,
    }

@api_router.post("/migrations/backfill-service-dates")
async def backfill_service_dates(current_user=Depends(require_admin)):
    """
    One-time migration: sets created_at = parsed check_in_date for imported service jobs
    where created_at is today (import artifact). Fixes Service Due calculations.
    """
    updated = 0; skipped = 0; errors = 0
    today_prefix = utcnow().strftime("%Y-%m-%dT")

    async for job in db.service_jobs.find(
        {"_imported": True, "check_in_date": {"$exists": True, "$ne": ""}},
        {"_id": 1, "check_in_date": 1, "created_at": 1}
    ):
        check_in = job.get("check_in_date", "")
        try:
            for fmt in ("%d/%m/%Y", "%d %b %Y", "%Y-%m-%d"):
                try:
                    parsed = datetime.strptime(check_in, fmt); break
                except ValueError:
                    continue
            else:
                skipped += 1; continue
            await db.service_jobs.update_one(
                {"_id": job["_id"]},
                {"$set": {"created_at": parsed.isoformat()}}
            )
            updated += 1
        except Exception:
            errors += 1

    return {"updated": updated, "skipped": skipped, "errors": errors,
            "message": f"✅ {updated} service jobs backfilled with correct dates. Run once only."}


@api_router.post("/migrations/backfill-delivered-milestones")
async def backfill_delivered_milestones(
    include_pending_completed: bool = Query(False, description="If true, also include sales with status=completed"),
    current_user=Depends(require_admin),
):
    """
    One-time migration: sets all 5 milestones to True for every sale with
    status='delivered' (and optionally 'completed'). Idempotent — safe to
    re-run. Uses a single bulk update.
    """
    all_true = {k: True for k in SALE_MILESTONE_KEYS}
    status_filter = ["delivered"]
    if include_pending_completed:
        status_filter.append("completed")

    result = await db.sales.update_many(
        {"status": {"$in": status_filter}},
        {"$set": {"milestones": all_true}},
    )
    return {
        "matched":  result.matched_count,
        "modified": result.modified_count,
        "status_filter": status_filter,
        "message": f"✅ {result.modified_count} delivered sales marked with all 5 milestones complete.",
    }


@api_router.post("/migrations/backfill-sale-addresses")
async def backfill_sale_addresses(current_user=Depends(require_admin)):
    """
    One-time migration: copies customer.address → sale.customer_address
    for every sale that is missing the address field or has it blank.
    Also copies back into the customers collection if the customer has no address.
    """
    updated = 0
    skipped = 0
    no_address = 0

    async for sale in db.sales.find(
        {"$or": [{"customer_address": {"$exists": False}}, {"customer_address": ""}, {"customer_address": None}]},
        {"_id": 1, "customer_id": 1, "customer_mobile": 1}
    ):
        # Try by customer_id first, fall back to mobile
        customer = None
        if sale.get("customer_id"):
            try:
                customer = await db.customers.find_one({"_id": obj_id(sale["customer_id"])}, {"address": 1})
            except Exception:
                pass
        if not customer and sale.get("customer_mobile"):
            customer = await db.customers.find_one({"mobile": sale["customer_mobile"]}, {"address": 1})

        address = (customer or {}).get("address", "").strip() if customer else ""

        if address:
            await db.sales.update_one(
                {"_id": sale["_id"]},
                {"$set": {"customer_address": address}}
            )
            updated += 1
        else:
            no_address += 1

        await asyncio.sleep(0) if updated % 100 == 0 else None

    return {
        "updated": updated,
        "no_customer_address": no_address,
        "message": f"✅ {updated} sales updated with address. {no_address} customers have no address on file — import template with addresses to fill those."
    }


@api_router.get("/reports/brand-sales")
async def brand_sales_report(current_user=Depends(require_admin)):
    pipeline = [{"$group":{"_id":"$vehicle_brand","units":{"$sum":1},"revenue":{"$sum":"$total_amount"}}},{"$sort":{"units":-1}}]
    docs = await db.sales.aggregate(pipeline).to_list(None)
    return [{"brand":d["_id"],"units":d["units"],"revenue":d["revenue"]} for d in docs]

@api_router.get("/reports/top-parts")
async def top_parts_report(limit: int = Query(10), current_user=Depends(require_admin)):
    pipeline = [{"$unwind":"$items"},{"$group":{"_id":"$items.name","qty":{"$sum":"$items.qty"},"revenue":{"$sum":"$items.total"}}},{"$sort":{"qty":-1}},{"$limit":limit}]
    docs = await db.parts_sales.aggregate(pipeline).to_list(limit)
    return [{"name":d["_id"],"qty_sold":d["qty"],"revenue":round(d["revenue"],2)} for d in docs]


# ─── GSTR-1 Export (CA flat format) ────────────────────────────────────────────
# Single-sheet flat report matching accountant's template.
# 3 header rows + 34-col table + totals row.
# 18% GST inclusive back-calc, CGST+SGST split, DD/MM/YYYY dates.
# One row per bill (service_bills = SERVICES BILL, sales = SALES BILL).

_GSTR1_GST_RATE = 18.0  # Service + showroom charges

def _parse_month_range(month: str):
    """month = 'YYYY-MM' → (start_dt, end_dt_exclusive, label)."""
    from datetime import datetime as _dt
    y, m = month.split("-")
    y, m = int(y), int(m)
    start = _dt(y, m, 1)
    end   = _dt(y + (1 if m == 12 else 0), 1 if m == 12 else m + 1, 1)
    return start, end, month

def _split_gst_inclusive(total: float, rate: float):
    """Back-calc taxable + CGST + SGST from GST-inclusive total.
    Rounds to 2dp per invoice. Karnataka intra-state → equal CGST/SGST."""
    total = float(total or 0)
    taxable = round(total / (1 + rate / 100), 2)
    gst_total = round(total - taxable, 2)
    cgst = round(gst_total / 2, 2)
    sgst = round(gst_total - cgst, 2)
    return taxable, cgst, sgst

async def _gstr1_fetch_sales(start, end):
    """Sales in month. sale_date format varies — parse defensively."""
    pipeline = [
        {"$addFields": {
            "p1": {"$dateFromString": {"dateString": "$sale_date", "format": "%d %b %Y", "onError": None, "onNull": None}},
            "p2": {"$dateFromString": {"dateString": {"$substr": ["$sale_date", 0, 10]}, "format": "%Y-%m-%d", "onError": None, "onNull": None}},
            "p3": {"$dateFromString": {"dateString": "$sale_date", "format": "%d/%m/%Y", "onError": None, "onNull": None}},
        }},
        {"$addFields": {"parsed": {"$ifNull": ["$p1", {"$ifNull": ["$p2", "$p3"]}]}}},
        {"$match": {"parsed": {"$gte": start, "$lt": end}}},
        {"$sort": {"parsed": 1}},
    ]
    return await db.sales.aggregate(pipeline).to_list(20000)

async def _gstr1_fetch_service_bills(start, end):
    """Service bills in month. bill_date format 'DD Mon YYYY'."""
    pipeline = [
        {"$addFields": {
            "p1": {"$dateFromString": {"dateString": "$bill_date", "format": "%d %b %Y", "onError": None, "onNull": None}},
            "p2": {"$dateFromString": {"dateString": {"$substr": ["$created_at", 0, 10]}, "format": "%Y-%m-%d", "onError": None, "onNull": None}},
        }},
        {"$addFields": {"parsed": {"$ifNull": ["$p1", "$p2"]}}},
        {"$match": {"parsed": {"$gte": start, "$lt": end}}},
        {"$sort": {"parsed": 1}},
    ]
    return await db.service_bills.aggregate(pipeline).to_list(20000)

def _fmt_ddmmyyyy(parsed):
    """MongoDB Motor returns datetime for parsed date. Fallback to blank."""
    from datetime import datetime as _dt
    if isinstance(parsed, _dt):
        return parsed.strftime("%d/%m/%Y")
    return ""

def _customer_display(name, mobile):
    name = (name or "").strip()
    mob  = (mobile or "").strip()
    if mob and mob not in name:
        return f"{name} {mob}".strip()
    return name

def _build_gstr1_workbook(month_label, sales, service_bills, period_from, period_to):
    """Build CA's flat single-sheet GST report."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime as _dt

    wb = Workbook()
    ws = wb.active
    ws.title = "GST"

    title_font  = Font(name="Arial", bold=True, size=12)
    sub_font    = Font(name="Arial", size=10)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1F3864")
    body_font   = Font(name="Arial", size=9)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin        = Side(border_style="thin", color="999999")
    box         = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header rows (3)
    ws.cell(row=1, column=1, value="M M MOTORS").font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=34)
    ws.cell(row=2, column=1, value="Bengaluru Main Road, Beside Ruchi Bakery, Malur, 563130  Ph : 7026263123  GSTIN: 29CUJPM6814P1ZQ").font = sub_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=34)
    period_str = f"Sales Invoice       From {period_from.strftime('%d/%m/%Y')} To {(period_to - timedelta(days=1)).strftime('%d/%m/%Y')}"
    ws.cell(row=3, column=1, value=period_str).font = sub_font
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=34)

    # Column header row (row 4, matching CA template exactly)
    headers = [
        "Branch Name","FormatName","Date","Bill No.","Bill/Reff. Date","Bill Reff. No.",
        "Customer Name","Party GSTIN","Item","AdditionalDescription","Item Group",
        "Quantity","Free Quantity","Rate","Item MRP","Billed Qty","Billed UOM","Billed Rate",
        "TaxAmount","Unit","Gross Amount","Disc %","Discount Amount","Taxable Value",
        "Pre Tax","Post Tax","Round Off","Net Amount","SGST Amount","CGST Amount",
        "IGST Amount","HSN/SAC Code","Cess Amount","Bill Amount",
    ]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = box

    # Data rows
    rows = []  # list of (parsed_date, format_name, dict of values)

    # Service bills → SERVICES BILL
    for b in service_bills:
        total = float(b.get("net_total") or b.get("grand_total") or 0)
        if total <= 0:
            continue
        taxable, cgst, sgst = _split_gst_inclusive(total, _GSTR1_GST_RATE)
        tax_total = round(cgst + sgst, 2)
        bill_amt  = round(taxable + tax_total)  # CA rounds to whole rupee
        round_off = round(bill_amt - (taxable + tax_total), 2)
        rows.append({
            "parsed":       b.get("parsed"),
            "format_name":  "SERVICES BILL",
            "date":         _fmt_ddmmyyyy(b.get("parsed")),
            "bill_no":      b.get("bill_number", ""),
            "customer":     _customer_display(b.get("customer_name",""), b.get("customer_mobile","")),
            "gstin":        (b.get("customer_gstin","") or "").strip().upper(),
            "item":         "Service",
            "taxable":      taxable,
            "tax_total":    tax_total,
            "cgst":         cgst,
            "sgst":         sgst,
            "round_off":    round_off,
            "bill_amount":  bill_amt,
        })

    # Sales → SALES BILL
    for s in sales:
        total = float(s.get("total_amount") or s.get("sale_price") or 0)
        if total <= 0:
            continue
        taxable, cgst, sgst = _split_gst_inclusive(total, _GSTR1_GST_RATE)
        tax_total = round(cgst + sgst, 2)
        bill_amt  = round(taxable + tax_total)
        round_off = round(bill_amt - (taxable + tax_total), 2)
        rows.append({
            "parsed":       s.get("parsed"),
            "format_name":  "SALES BILL",
            "date":         _fmt_ddmmyyyy(s.get("parsed")),
            "bill_no":      s.get("invoice_number", ""),
            "customer":     _customer_display(s.get("customer_name",""), s.get("customer_mobile","")),
            "gstin":        (s.get("customer_gstin","") or "").strip().upper(),
            "item":         "Showroom Charges/CONSULTATION",
            "taxable":      taxable,
            "tax_total":    tax_total,
            "cgst":         cgst,
            "sgst":         sgst,
            "round_off":    round_off,
            "bill_amount":  bill_amt,
        })

    # Sort by parsed date then format
    rows.sort(key=lambda r: (r["parsed"] or _dt.min, r["format_name"]))

    r_idx = 5  # data starts row 5
    for r in rows:
        ws.cell(row=r_idx, column=1,  value="Main Location")
        ws.cell(row=r_idx, column=2,  value=r["format_name"])
        ws.cell(row=r_idx, column=3,  value=r["date"])
        ws.cell(row=r_idx, column=4,  value=r["bill_no"])
        ws.cell(row=r_idx, column=5,  value="")   # Bill/Reff. Date
        ws.cell(row=r_idx, column=6,  value="")   # Bill Reff. No.
        ws.cell(row=r_idx, column=7,  value=r["customer"])
        ws.cell(row=r_idx, column=8,  value=r["gstin"])
        ws.cell(row=r_idx, column=9,  value=r["item"])
        ws.cell(row=r_idx, column=10, value="")   # AdditionalDescription
        ws.cell(row=r_idx, column=11, value="Main Group")
        ws.cell(row=r_idx, column=12, value=1)      # Quantity
        ws.cell(row=r_idx, column=13, value=0)      # Free Quantity
        ws.cell(row=r_idx, column=14, value=r["taxable"])   # Rate
        ws.cell(row=r_idx, column=15, value=0.00)   # Item MRP
        ws.cell(row=r_idx, column=16, value=1)      # Billed Qty
        ws.cell(row=r_idx, column=17, value="Amt")  # Billed UOM
        ws.cell(row=r_idx, column=18, value=r["taxable"])   # Billed Rate
        ws.cell(row=r_idx, column=19, value=r["tax_total"]) # TaxAmount
        ws.cell(row=r_idx, column=20, value="Amt")  # Unit
        ws.cell(row=r_idx, column=21, value=r["taxable"])   # Gross Amount
        ws.cell(row=r_idx, column=22, value=0)      # Disc %
        ws.cell(row=r_idx, column=23, value=0.00)   # Discount Amount
        ws.cell(row=r_idx, column=24, value=r["taxable"])   # Taxable Value
        ws.cell(row=r_idx, column=25, value=0)      # Pre Tax
        ws.cell(row=r_idx, column=26, value=0)      # Post Tax
        ws.cell(row=r_idx, column=27, value=r["round_off"]) # Round Off
        ws.cell(row=r_idx, column=28, value=r["taxable"])   # Net Amount
        ws.cell(row=r_idx, column=29, value=r["sgst"])      # SGST Amount
        ws.cell(row=r_idx, column=30, value=r["cgst"])      # CGST Amount
        ws.cell(row=r_idx, column=31, value=0)      # IGST Amount
        ws.cell(row=r_idx, column=32, value="")     # HSN/SAC Code
        ws.cell(row=r_idx, column=33, value=0)      # Cess Amount
        ws.cell(row=r_idx, column=34, value=r["bill_amount"])
        for col_idx in range(1, 35):
            c = ws.cell(row=r_idx, column=col_idx)
            c.font = body_font
            c.border = box
        r_idx += 1

    # Totals row (formulas, not hardcoded)
    if rows:
        from openpyxl.utils import get_column_letter as _gcl
        first_data_row = 5
        last_data_row  = r_idx - 1
        totals_row     = r_idx
        # Only sum numeric columns per CA template: Quantity(12), Free Qty(13),
        # TaxAmount(19), Gross Amount(21), Disc Amt(23), Taxable(24), Pre Tax(25),
        # Post Tax(26), Round Off(27), Net Amount(28), SGST(29), CGST(30),
        # IGST(31), Cess(33), Bill Amount(34)
        sum_cols = [12, 13, 19, 21, 23, 24, 25, 26, 27, 28, 29, 30, 31, 33, 34]
        for col_idx in sum_cols:
            col_letter = _gcl(col_idx)
            ws.cell(row=totals_row, column=col_idx,
                    value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
        for col_idx in range(1, 35):
            c = ws.cell(row=totals_row, column=col_idx)
            c.font = Font(name="Arial", size=9, bold=True)
            c.border = box
            c.fill = PatternFill("solid", fgColor="FFF2CC")

    # Column widths
    from openpyxl.utils import get_column_letter as _gcl2
    widths = {
        1:14, 2:14, 3:12, 4:12, 5:12, 6:12, 7:26, 8:16, 9:22, 10:16, 11:14,
        12:10, 13:10, 14:12, 15:10, 16:10, 17:10, 18:12, 19:12, 20:8,
        21:14, 22:8, 23:12, 24:14, 25:10, 26:10, 27:10, 28:14, 29:12, 30:12,
        31:12, 32:14, 33:10, 34:14,
    }
    for col_idx, w in widths.items():
        ws.column_dimensions[_gcl2(col_idx)].width = w

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@api_router.get("/reports/gstr1")
async def gstr1_export(
    month: str = Query(..., regex=r"^\d{4}-\d{2}$"),
    current_user=Depends(require_admin),
):
    """GST monthly export in CA's flat format (YYYY-MM). Returns xlsx blob.
    Single sheet: Business header + 34-col flat table + totals row.
    18% inclusive GST back-calc, CGST/SGST equal split, Karnataka intra-state."""
    try:
        start, end, label = _parse_month_range(month)
    except Exception:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")
    sales, service_bills = await asyncio.gather(
        _gstr1_fetch_sales(start, end),
        _gstr1_fetch_service_bills(start, end),
    )
    buf = _build_gstr1_workbook(label, sales, service_bills, start, end)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="GST_{label}.xlsx"'},
    )


@api_router.get("/reports/complimentary")
async def complimentary_report(
    date_from: Optional[str] = Query(None),
    date_to:   Optional[str] = Query(None),
    current_user=Depends(verify_token),
):
    """Complimentary (free) items given across service_bills and parts_bills.
    Groups by part_number / name. Shows MRP value forgone."""
    match: dict = {"items.complimentary": True}
    if date_from or date_to:
        rng: dict = {}
        if date_from: rng["$gte"] = date_from
        if date_to:   rng["$lte"] = date_to
        match["created_at"] = rng

    pipeline = [
        {"$match": match},
        {"$unwind": "$items"},
        {"$match": {"items.complimentary": True}},
        {"$group": {
            "_id": {"pn": "$items.part_number", "name": {"$ifNull": ["$items.name", "$items.description"]}},
            "qty":       {"$sum": "$items.qty"},
            "mrp_value": {"$sum": {"$multiply": [{"$ifNull": ["$items.mrp", 0]}, "$items.qty"]}},
            "occurrences": {"$sum": 1},
        }},
        {"$sort": {"qty": -1}},
    ]
    sb, pb = await asyncio.gather(
        db.service_bills.aggregate(pipeline).to_list(500),
        db.parts_bills.aggregate(pipeline).to_list(500),
    )
    merged: dict = {}
    for d in (sb + pb):
        key = d["_id"].get("pn") or d["_id"].get("name")
        row = merged.setdefault(key, {
            "part_number": d["_id"].get("pn",""),
            "name":        d["_id"].get("name",""),
            "qty": 0, "mrp_value": 0.0, "occurrences": 0,
        })
        row["qty"]         += d["qty"]
        row["mrp_value"]   += d["mrp_value"]
        row["occurrences"] += d["occurrences"]
    result = sorted(merged.values(), key=lambda r: r["qty"], reverse=True)
    return {"items": result, "total_value": round(sum(r["mrp_value"] for r in result), 2)}


# ═══════════════════════════════════════════════════════════════════════════════
#  IMPORT PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════

import io
import csv as csv_module
from openpyxl import load_workbook
from openpyxl import Workbook
from fastapi import File, Form, UploadFile
from fastapi.responses import StreamingResponse

EXPENSE_CATEGORIES = [
    "Staff Salaries", "Rent & Utilities", "Vehicle Purchase",
    "Parts & Consumables", "RTO & Insurance", "Transport & Logistics",
    "Marketing & Advertising", "Bank Charges & Loan EMI",
    "Equipment & Maintenance", "Miscellaneous",
]

import_router = APIRouter(prefix="/api/import", tags=["import"])

def safe(val, default=""):
    if val is None or str(val).strip() in ("","None","nan"):
        return default
    v = str(val).strip()
    try:
        f = float(v)
        if f == int(f):
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return v

def safe_float(val, default=0.0) -> float:
    try:
        return float(str(val).replace(",","").strip())
    except (ValueError, TypeError):
        return default

def safe_int(val, default=0) -> int:
    try:
        return int(float(str(val).replace(",","").strip()))
    except (ValueError, TypeError):
        return default

def read_file(content: bytes, filename: str) -> list:
    name = (filename or "").lower()
    rows = []
    if name.endswith(".csv"):
        text   = content.decode("utf-8-sig", errors="replace")
        reader = csv_module.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip().lower().replace(" ","_").replace("/","_"): safe(v) for k,v in row.items()})
    else:
        wb   = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        # Find the data sheet — prefer "Data" sheet, skip "Instructions"
        ws = None
        for sheet_name in wb.sheetnames:
            sn = sheet_name.lower()
            if "data" in sn or "📥" in sn:
                ws = wb[sheet_name]; break
        # Fallback: pick the last sheet (Instructions is first in our templates)
        if ws is None:
            ws = wb[wb.sheetnames[-1]] if len(wb.sheetnames) > 1 else wb.active
        data = list(ws.values)
        wb.close()
        if len(data) < 2:
            return []
        # Find header row (first row with actual column names, skip banner rows)
        header_row_idx = 0
        for idx, row in enumerate(data):
            non_empty = [c for c in row if c is not None and str(c).strip()]
            if len(non_empty) >= 2:
                header_row_idx = idx
                break
        headers_raw = data[header_row_idx]
        headers = [
            str(h).strip().lower()
              .replace(" ","_").replace("/","_")
              .replace("*","").replace("📥","").strip()
            if h else f"col{i}"
            for i,h in enumerate(headers_raw)
        ]
        for row in data[header_row_idx + 1:]:
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            row_dict = {}
            for i in range(min(len(headers), len(row))):
                cell_val = row[i]
                # Handle Excel datetime objects
                if hasattr(cell_val, 'strftime'):
                    cell_val = cell_val.strftime("%d/%m/%Y")
                row_dict[headers[i]] = safe(cell_val)
            rows.append(row_dict)
    return rows

def result_summary(inserted, skipped, errors):
    return {
        "inserted":      inserted,
        "skipped_count": len(skipped),
        "error_count":   len(errors),
        "skipped":       skipped[:100],
        "errors":        errors[:100],
        "summary":       f"✅ {inserted} imported, ⏭ {len(skipped)} skipped, ❌ {len(errors)} errors",
    }

TEMPLATES = {
    "customers": {
        "cols":["name","mobile","care_of","email","address","tags"],
        "rows":[
            ["Ravi Kumar","9876543210","Srinivas","ravi@example.com","12 MG Road, Bengaluru","VIP"],
            ["Meena Shetty","9845123456","","","45 Koramangala, Bengaluru","Loyal"],
            ["ABC Infra Ltd","9900112233","","accounts@abc.com","HSR Layout, Bengaluru","Corporate"],
        ]
    },
    "vehicles": {
        "cols":["brand","model","variant","color","chassis_number","engine_number","vehicle_number","key_number","type","status","inbound_date","inbound_location","return_date","returned_location"],
        "rows":[
            ["HONDA","Activa 6G","STD","Pearl Black","ME4JF502RH7000001","JF50E7000001","KA01HH1234","K001","new","in_stock","01/04/2026","Showroom","",""],
            ["HERO","Splendor+","Self Start","Heavy Grey","MBLHA10EVHM000002","HA10EAHM00002","KA03AB5678","K002","new","in_stock","05/04/2026","Showroom","",""],
            ["BAJAJ","Pulsar 150","Drum","Black Red","MD2DHDZZXRCB12345","DHZDRCB12345","","","used","returned","10/03/2026","Showroom","01/04/2026","Showroom"],
        ]
    },
    "sales": {
        "cols":["customer_name","customer_mobile","care_of","vehicle_brand","vehicle_model","chassis_number",
                "engine_number","vehicle_number","vehicle_color","vehicle_variant","sale_price",
                "rto","financier","payment_mode","nominee_name","nominee_relation","nominee_age",
                "sale_date","customer_address"],
        "rows":[
            ["Ravi Kumar","9876543210","Srinivas","HONDA","Activa 6G","ME4JF502RH7000001","JF50E7000001",
             "KA01HH1234","Pearl Black","STD","80500","KA07","","Cash",
             "Balakrishna","Father","54","08/04/2026","12 MG Road, Bengaluru"],
            ["Priya Nair","9845001122","","HERO","Splendor+","MBLHA10EVHM000002","HA10EAHM00002",
             "","Heavy Grey","Self Start","73200","","HDFC Bank","Finance",
             "Suresh Nair","Husband","42","15/04/2026",""],
        ]
    },
    "service": {
        "cols":["customer_name","customer_mobile","vehicle_number","brand","model",
                "odometer_km","complaint","technician","check_in_date","status","amount","notes"],
        "rows":[
            ["Ravi Kumar","9876543210","KA01HH1234","HONDA","Activa 6G",
             "8420","Engine noise, oil change","Suresh","07/04/2026","delivered","350",""],
            ["Meena Shetty","9845123456","KA03AB5678","HERO","Splendor+",
             "12500","Routine service, brake pad","Arun","10/04/2026","delivered","500","Chain also tightened"],
        ]
    },
    "parts": {
        "cols":["part_number","name","category","brand","compatible_with","stock",
                "reorder_level","purchase_price","selling_price","gst_rate","hsn_code","location"],
        "rows":[
            ["30050-KWB-901","Spark Plug (Iridium)","Engine","NGK","HONDA,TVS","24","10","180","280","18","8511","A1-R2"],
            ["15400-PLM-A01","Oil Filter","Filters","Honda","HONDA","18","15","120","195","18","8421","A2-R1"],
            ["06435-KZR-305","Brake Pad Set (Front)","Brakes","Honda","HONDA","12","8","250","390","18","8708","B1-R3"],
        ]
    },
    "staff": {
        "cols":["name","mobile","email","username","role","salary","join_date"],
        "rows":[
            ["Rajesh Kumar","9845001122","rajesh@mmmotors.com","rajesh_k","sales","18000","01/03/2023"],
            ["Arun Shetty","9566001122","arun@mmmotors.com","arun_s","service_advisor","20000","01/04/2023"],
            ["Suresh B","9900334455","","suresh_b","technician","16000","15/06/2023"],
        ]
    },
    "expenses": {
        "cols":["date","category","sub_category","amount","description","vendor","payment_mode","receipt_no","notes"],
        "rows":[
            ["01/04/2026","Staff Salaries","","72000","April salaries — 4 staff","","Cash","",""],
            ["01/04/2026","Rent & Utilities","Rent","25000","Showroom monthly rent","Landlord","Bank","R001",""],
            ["05/04/2026","Parts & Consumables","","8500","Engine oil, grease, consumables","Spare parts supplier","Cash","",""],
            ["10/04/2026","Transport & Logistics","","3200","Vehicle delivery charges","Transport vendor","UPI","",""],
            ["15/04/2026","Marketing & Advertising","","5000","Social media ads","","UPI","",""],
        ]
    },
}

@import_router.get("/template/{entity}")
async def download_template(entity: str, current_user=Depends(verify_token)):
    if entity not in TEMPLATES:
        raise HTTPException(status_code=404, detail=f"No template for '{entity}'")
    t  = TEMPLATES[entity]
    wb = Workbook(); ws = wb.active; ws.title = entity
    ws.append(t["cols"])
    for row in t["rows"]:
        ws.append(row + [""]*(len(t["cols"])-len(row)))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition":f'attachment; filename="template_{entity}.xlsx"'})

@import_router.post("/preview/{entity}")
async def preview_import(entity: str, file: UploadFile = File(...), current_user=Depends(verify_token)):
    if entity not in TEMPLATES:
        raise HTTPException(status_code=404, detail=f"Unknown entity: {entity}")
    content = await file.read()
    rows    = read_file(content, file.filename or "")
    return {"entity":entity,"total_rows":len(rows),"columns_found":list(rows[0].keys()) if rows else [],"preview":rows[:10],"template_cols":TEMPLATES[entity]["cols"]}

@import_router.post("/customers")
async def import_customers(request: Request, file: UploadFile = File(...), mode: str = Form("skip"), current_user=Depends(verify_token)):
    content = await file.read(); rows = read_file(content, file.filename or "")
    if not rows: raise HTTPException(status_code=400, detail="File is empty or could not be parsed")
    inserted, skipped, errors = 0, [], []

    # Dedup by mobile (primary) — same mobile = same customer
    existing_mobiles = set()
    async for doc in db.customers.find({}, {"mobile": 1}):
        if doc.get("mobile"): existing_mobiles.add(str(doc["mobile"]))

    to_insert = []
    to_update = []
    for i, row in enumerate(rows):
        if i % 200 == 0: await asyncio.sleep(0)
        rn = i + 2
        try:
            name   = safe(row.get("name"))
            mobile = safe(row.get("mobile"))
            if not name:   skipped.append({"row": rn, "reason": "Missing name"});   continue
            if not mobile: skipped.append({"row": rn, "reason": "Missing mobile"}); continue
            doc = {
                "name":       name,
                "mobile":     mobile,
                "care_of":    safe(row.get("care_of", "")),
                "email":      safe(row.get("email")),
                "address":    safe(row.get("address")),
                "tags":       [t.strip() for t in safe(row.get("tags","")).split(",") if t.strip()],
                "created_at": utcnow().isoformat(),
            }
            if mobile in existing_mobiles:
                if mode == "overwrite":
                    to_update.append(doc)
                else:
                    skipped.append({"row": rn, "reason": f"Mobile {mobile} already exists"})
            else:
                to_insert.append(doc)
                existing_mobiles.add(mobile)  # prevent intra-file dupes
        except Exception as e:
            errors.append({"row": rn, "error": str(e)})

    # Bulk insert
    if to_insert:
        try:
            result = await db.customers.insert_many(to_insert, ordered=False)
            inserted += len(result.inserted_ids)
        except Exception as e:
            errors.append({"row": "bulk", "error": str(e)})

    # Updates (sequential — need to match by mobile)
    for doc in to_update:
        try:
            await db.customers.update_one(
                {"mobile": doc["mobile"]},
                {"$set": {k: v for k, v in doc.items() if k != "created_at"}}
            )
            inserted += 1
        except Exception as e:
            errors.append({"row": "update", "error": str(e)})

    return result_summary(inserted, skipped, errors)

@import_router.post("/vehicles")
async def import_vehicles(request: Request, file: UploadFile = File(...), mode: str = Form("skip"), current_user=Depends(verify_token)):
    content = await file.read(); rows = read_file(content, file.filename or "")
    inserted, skipped, errors = 0, [], []
    existing_chassis = set()
    existing_veh_numbers = set()
    async for doc in db.vehicles.find({}, {"chassis_number":1,"vehicle_number":1}):
        if doc.get("chassis_number"): existing_chassis.add(str(doc["chassis_number"]))
        if doc.get("vehicle_number"): existing_veh_numbers.add(str(doc["vehicle_number"]).upper())
    to_insert, to_update = [], []
    for i, row in enumerate(rows):
        if i % 200 == 0: await asyncio.sleep(0)
        rn = i + 2
        try:
            chassis = safe(row.get("chassis_number","")).upper().replace(" ","")
            brand   = safe(row.get("brand","")).upper()
            model   = safe(row.get("model",""))
            veh_no  = safe(row.get("vehicle_number","")).upper()
            if not chassis: skipped.append({"row":rn,"reason":"Missing chassis_number"}); continue
            if not brand:   skipped.append({"row":rn,"reason":"Missing brand"});          continue
            if not model:   skipped.append({"row":rn,"reason":"Missing model"});          continue
            doc = {"brand":brand,"model":model,"variant":safe(row.get("variant")),"color":safe(row.get("color")),"chassis_number":chassis,"engine_number":safe(row.get("engine_number")),"vehicle_number":safe(row.get("vehicle_number")),"key_number":safe(row.get("key_number")),"type":safe(row.get("type","new")).lower() or "new","status":safe(row.get("status","in_stock")).lower() or "in_stock","inbound_date":safe(row.get("inbound_date","")),"inbound_location":safe(row.get("inbound_location","")),"return_date":safe(row.get("return_date","")),"returned_location":safe(row.get("returned_location","")),"created_at":utcnow().isoformat()}
            if chassis in existing_chassis:
                if mode=="overwrite": to_update.append(doc)
                else: skipped.append({"row":rn,"reason":f"Chassis {chassis} already imported"})
            elif veh_no and veh_no in existing_veh_numbers:
                skipped.append({"row":rn,"reason":f"Vehicle number {veh_no} already imported"})
            else:
                to_insert.append(doc)
                existing_chassis.add(chassis)
                if veh_no: existing_veh_numbers.add(veh_no)
        except Exception as e:
            errors.append({"row":rn,"error":str(e)})
    if to_insert:
        try:
            r = await db.vehicles.insert_many(to_insert, ordered=False); inserted += len(r.inserted_ids)
        except Exception as e: errors.append({"row":"bulk","error":str(e)})
    for doc in to_update:
        try:
            await db.vehicles.update_one({"chassis_number":doc["chassis_number"]},{"$set":{k:v for k,v in doc.items() if k!="created_at"}}); inserted += 1
        except Exception as e: errors.append({"row":"update","error":str(e)})
    return result_summary(inserted, skipped, errors)

@import_router.post("/sales")
async def import_sales(request: Request, file: UploadFile = File(...), mode: str = Form("skip"), current_user=Depends(verify_token)):
    content = await file.read(); rows = read_file(content, file.filename or "")
    inserted, skipped, errors = 0, [], []

    # Pre-load existing dedup keys
    existing_chassis = set()   # primary dedup: chassis_number
    existing_sale_keys = set() # fallback dedup: mobile|model|sale_date (for chassis-less records)
    chassis_to_sale = {}       # chassis -> {_id, address} for address patching
    fallback_to_sale = {}      # fallback_key -> {_id, address} for address patching
    async for doc in db.sales.find({}, {"chassis_number":1,"customer_mobile":1,"vehicle_model":1,"sale_date":1,"customer_address":1}):
        ch  = str(doc.get("chassis_number","")).strip()
        mob = str(doc.get("customer_mobile","")).strip()
        mdl = str(doc.get("vehicle_model","")).strip().lower()
        dt  = str(doc.get("sale_date","")).strip()
        addr = str(doc.get("customer_address","")).strip()
        sale_ref = {"id": str(doc["_id"]), "address": addr}
        if ch:
            existing_chassis.add(ch)
            chassis_to_sale[ch] = sale_ref
        if mob and mdl and dt:
            fk = f"{mob}|{mdl}|{dt}"
            existing_sale_keys.add(fk)
            fallback_to_sale[fk] = sale_ref
    existing_inv = set()
    async for doc in db.sales.find({}, {"invoice_number": 1}):
        if doc.get("invoice_number"): existing_inv.add(str(doc["invoice_number"]))
    customer_cache = {}
    async for doc in db.customers.find({}, {"_id":1,"mobile":1}):
        if doc.get("mobile"): customer_cache[str(doc["mobile"])] = str(doc["_id"])

    # Insert row-by-row so:
    #   1. Each row gets its own error (no "Row —" bulk mystery)
    #   2. Customer is only created if the sale insert succeeds
    for i, row in enumerate(rows):
        if i % 20 == 0: await asyncio.sleep(0)
        rn = i + 2
        try:
            name   = safe(row.get("customer_name"))
            mobile = safe(row.get("customer_mobile"))
            brand  = safe(row.get("vehicle_brand","")).upper()
            model  = safe(row.get("vehicle_model",""))
            price  = safe_float(row.get("sale_price", 0))
            chassis= safe(row.get("chassis_number","")).upper().replace(" ","")

            # Validate required fields
            if not name or not mobile or not brand or not model or not price:
                missing = [f for f,v in [("customer_name",name),("customer_mobile",mobile),("vehicle_brand",brand),("vehicle_model",model),("sale_price",price)] if not v]
                skipped.append({"row":rn,"reason":f"Missing: {', '.join(missing)}"}); continue

            # Dedup: primary by chassis, fallback by mobile+model+date
            sale_date_val = safe(row.get("sale_date","")) or utcnow().strftime("%d %b %Y")
            fallback_key  = f"{mobile}|{model.lower()}|{sale_date_val}"
            import_address = safe(row.get("customer_address",""))

            if chassis and chassis in existing_chassis:
                # Sale exists — patch address if missing in DB but present in import row
                ref = chassis_to_sale.get(chassis, {})
                if import_address and not ref.get("address"):
                    await db.sales.update_one(
                        {"_id": obj_id(ref["id"])},
                        {"$set": {"customer_address": import_address}}
                    )
                    # Also patch the customer record
                    await db.customers.update_one(
                        {"mobile": mobile, "address": {"$in": ["", None]}},
                        {"$set": {"address": import_address}}
                    )
                    skipped.append({"row":rn,"reason":f"Chassis {chassis} already imported — address updated"})
                else:
                    skipped.append({"row":rn,"reason":f"Chassis {chassis} already imported"})
                continue

            if not chassis and fallback_key in existing_sale_keys:
                # Sale exists — patch address if missing
                ref = fallback_to_sale.get(fallback_key, {})
                if import_address and not ref.get("address"):
                    await db.sales.update_one(
                        {"_id": obj_id(ref["id"])},
                        {"$set": {"customer_address": import_address}}
                    )
                    await db.customers.update_one(
                        {"mobile": mobile, "address": {"$in": ["", None]}},
                        {"$set": {"address": import_address}}
                    )
                    skipped.append({"row":rn,"reason":f"Sale for {name} already imported — address updated"})
                else:
                    skipped.append({"row":rn,"reason":f"Sale for {name} ({model}, {sale_date_val}) already imported"})
                continue

            # Resolve or create customer — only after sale validation passes
            new_customer_id = None
            if mobile in customer_cache:
                cust_id = customer_cache[mobile]
            else:
                r = await db.customers.insert_one({
                    "name":name,"mobile":mobile,
                    "care_of":safe(row.get("care_of","")),
                    "email":"","address":safe(row.get("customer_address","")),
                    "tags":[],"created_at":utcnow().isoformat()
                })
                cust_id = str(r.inserted_id)
                new_customer_id = cust_id  # track so we can delete on failure
                customer_cache[mobile] = cust_id

            discount  = safe_float(row.get("discount",0))
            insurance = safe_float(row.get("insurance",0))
            rto       = safe(row.get("rto",""))          # RTO office code e.g. KA07 — stored as text
            total     = round(price - discount + insurance, 2)
            inv_no    = await next_sequence("invoice")

            doc = {
                "invoice_number": inv_no,
                "customer_id":    cust_id,
                "customer_name":  name,
                "customer_mobile":mobile,
                "care_of":        safe(row.get("care_of","")),
                "customer_address":safe(row.get("customer_address","")),
                "vehicle_brand":  brand,
                "vehicle_model":  model,
                "chassis_number": chassis,
                "engine_number":  safe(row.get("engine_number")),
                "vehicle_number": safe(row.get("vehicle_number")),
                "vehicle_color":  safe(row.get("vehicle_color")),
                "vehicle_variant":safe(row.get("vehicle_variant")),
                "sale_price":     price,
                "discount":       discount,
                "insurance":      insurance,
                "rto":            rto,  # RTO office code e.g. KA07
                "total_amount":   total,
                "amount_in_words":amount_in_words(total),
                "financier":      safe(row.get("financier","")),
                "payment_mode":   safe(row.get("payment_mode","Cash")),
                "nominee": {
                    "name":     safe(row.get("nominee_name")),
                    "relation": safe(row.get("nominee_relation")),
                    "age":      safe(row.get("nominee_age")),
                },
                "sale_date":  safe(row.get("sale_date","")) or utcnow().strftime("%d %b %Y"),
                "status":     "delivered",
                "created_at": utcnow().isoformat(),
                "_imported":  True,
            }

            # Insert sale individually — clear error per row if it fails
            try:
                await db.sales.insert_one(doc)
                inserted += 1
                if chassis: existing_chassis.add(chassis)
                existing_inv.add(inv_no)
                existing_sale_keys.add(fallback_key)
            except Exception as sale_err:
                # Roll back the newly created customer to avoid orphans
                if new_customer_id:
                    await db.customers.delete_one({"_id": obj_id(new_customer_id)})
                    del customer_cache[mobile]
                errors.append({"row": rn, "error": str(sale_err)})

        except Exception as e:
            errors.append({"row": rn, "error": str(e)})

    return result_summary(inserted, skipped, errors)

@import_router.post("/service")
async def import_service(request: Request, file: UploadFile = File(...), mode: str = Form("skip"), current_user=Depends(verify_token)):
    content = await file.read(); rows = read_file(content, file.filename or "")
    inserted, skipped, errors = 0, [], []
    # Build dedup set: vehicle_number+check_in_date
    existing_keys = set()
    async for doc in db.service_jobs.find({}, {"vehicle_number":1,"check_in_date":1}):
        if doc.get("vehicle_number") and doc.get("check_in_date"):
            existing_keys.add(f"{doc['vehicle_number']}|{doc['check_in_date']}")
    # Build customer cache
    customer_cache = {}
    async for doc in db.customers.find({}, {"_id":1,"mobile":1}):
        if doc.get("mobile"): customer_cache[str(doc["mobile"])] = str(doc["_id"])
    to_insert = []
    for i, row in enumerate(rows):
        if i % 100 == 0: await asyncio.sleep(0)
        rn = i + 2
        try:
            name     = safe(row.get("customer_name",""))
            mobile   = safe(row.get("customer_mobile",""))
            veh_no   = safe(row.get("vehicle_number","")).upper()
            complaint= safe(row.get("complaint","")) or "Service"
            check_in = safe(row.get("check_in_date","")) or utcnow().strftime("%d %b %Y")
            amount   = safe_float(row.get("amount",0))
            dedup_key = f"{veh_no}|{check_in}"
            if veh_no and check_in and dedup_key in existing_keys:
                skipped.append({"row":rn,"reason":f"Job for {veh_no} on {check_in} exists"}); continue
            if mobile in customer_cache:
                cust_id = customer_cache[mobile]
            elif name:
                r = await db.customers.insert_one({"name":name,"mobile":mobile,"email":"","address":"","tags":[],"created_at":utcnow().isoformat()})
                cust_id = str(r.inserted_id); customer_cache[mobile] = cust_id
            else: cust_id = ""
            status = safe(row.get("status","delivered")).lower()
            if status not in ("pending","in_progress","ready","delivered"): status = "delivered"
            job_no = await next_sequence("job")
            # Parse check_in_date → ISO for created_at so service_due uses real date
            try:
                parsed_checkin = datetime.strptime(check_in, "%d/%m/%Y")
            except ValueError:
                try:
                    parsed_checkin = datetime.strptime(check_in, "%d %b %Y")
                except ValueError:
                    try:
                        parsed_checkin = datetime.strptime(check_in, "%Y-%m-%d")
                    except ValueError:
                        parsed_checkin = utcnow()
            created_iso = parsed_checkin.isoformat()
            to_insert.append({"job_number":job_no,"customer_id":cust_id,"customer_name":name or "","customer_mobile":mobile or "","vehicle_number":veh_no or "","brand":safe(row.get("brand","")).upper(),"model":safe(row.get("model","")),"odometer_km":safe_int(row.get("odometer_km",0)),"complaint":complaint,"technician":safe(row.get("technician","")),"check_in_date":check_in,"status":status,"grand_total":amount,"notes":safe(row.get("notes","")),"created_at":created_iso,"_imported":True})
            if veh_no: existing_keys.add(dedup_key)
        except Exception as e:
            errors.append({"row":rn,"error":str(e)})
    if to_insert:
        try:
            r = await db.service_jobs.insert_many(to_insert, ordered=False); inserted += len(r.inserted_ids)
        except Exception as e: errors.append({"row":"bulk","error":str(e)})
    return result_summary(inserted, skipped, errors)

@import_router.post("/parts")
async def import_parts(request: Request, file: UploadFile = File(...), mode: str = Form("skip"), current_user=Depends(verify_token)):
    content = await file.read(); rows = read_file(content, file.filename or "")
    inserted, skipped, errors = 0, [], []
    existing_parts = {}
    async for doc in db.spare_parts.find({}, {"part_number":1,"stock":1,"selling_price":1,"purchase_price":1}):
        if doc.get("part_number"): existing_parts[str(doc["part_number"])] = doc
    to_insert, to_update = [], []
    for i, row in enumerate(rows):
        if i % 200 == 0: await asyncio.sleep(0)
        rn = i + 2
        try:
            part_no = safe(row.get("part_number","")).strip()
            name    = safe(row.get("name",""))
            if not part_no: skipped.append({"row":rn,"reason":"Missing part_number"}); continue
            if not name:    skipped.append({"row":rn,"reason":"Missing name"});        continue
            compat_raw = safe(row.get("compatible_with",""))
            doc = {"part_number":part_no,"name":name,"category":safe(row.get("category","")),"brand":safe(row.get("brand","")),"compatible_with":[c.strip().upper() for c in compat_raw.split(",") if c.strip()],"stock":safe_int(row.get("stock",0)),"reorder_level":safe_int(row.get("reorder_level",5)),"purchase_price":safe_float(row.get("purchase_price",0)),"selling_price":safe_float(row.get("selling_price",0)),"gst_rate":safe_float(row.get("gst_rate",18)),"hsn_code":safe(row.get("hsn_code","")),"location":safe(row.get("location","")),"created_at":utcnow().isoformat()}
            if part_no in existing_parts:
                if mode=="overwrite": to_update.append(doc)
                else: skipped.append({"row":rn,"reason":f"Part {part_no} already exists"})
            else:
                to_insert.append(doc); existing_parts[part_no] = doc
        except Exception as e:
            errors.append({"row":rn,"error":str(e)})
    if to_insert:
        try:
            r = await db.spare_parts.insert_many(to_insert, ordered=False); inserted += len(r.inserted_ids)
        except Exception as e: errors.append({"row":"bulk","error":str(e)})
    for doc in to_update:
        try:
            await db.spare_parts.update_one({"part_number":doc["part_number"]},{"$set":{k:v for k,v in doc.items() if k!="created_at"}}); inserted += 1
        except Exception as e: errors.append({"row":"update","error":str(e)})
    return result_summary(inserted, skipped, errors)

@import_router.post("/staff")
async def import_staff(file: UploadFile = File(...), mode: str = Form("skip"), current_user=Depends(require_admin)):
    content = await file.read(); rows = read_file(content, file.filename or "")
    inserted, skipped, errors = 0, [], []
    valid_roles = {"owner","sales","service_advisor","parts_counter","technician"}
    for i, row in enumerate(rows):
        rn = i+2
        try:
            name=safe(row.get("name","")); username=safe(row.get("username","")).strip().lower(); role=safe(row.get("role","sales")).strip().lower()
            if not name or not username: skipped.append({"row":rn,"reason":"Missing name or username"}); continue
            if role not in valid_roles:  skipped.append({"row":rn,"reason":f"Invalid role: {role}"});   continue
            existing = await db.users.find_one({"username":username})
            if existing:
                if mode=="overwrite":
                    await db.users.update_one({"username":username},{"$set":{"name":name,"role":role,"salary":safe_float(row.get("salary",0))}})
                    inserted += 1
                else: skipped.append({"row":rn,"reason":f"Username {username} already exists"})
                continue
            await db.users.insert_one({"username":username,"name":name,"mobile":safe(row.get("mobile","")),"email":safe(row.get("email","")),"role":role,"password":pwd_ctx.hash("mm@123456"),"salary":safe_float(row.get("salary",0)),"join_date":safe(row.get("join_date","")),"status":"active","created_at":utcnow().isoformat()})
            inserted += 1
        except Exception as e:
            traceback.print_exc(); errors.append({"row":rn,"error":str(e)})
    return result_summary(inserted, skipped, errors)

@import_router.post("/expenses")
async def import_expenses(request: Request, file: UploadFile = File(...), mode: str = Form("skip"), current_user=Depends(verify_token)):
    content = await file.read(); rows = read_file(content, file.filename or "")
    inserted, skipped, errors = 0, [], []
    to_insert = []
    valid_cats = set(EXPENSE_CATEGORIES)
    # Dedup expenses by date+category+amount+description fingerprint
    existing_expense_keys = set()
    async for doc in db.expenses.find({}, {"date":1,"category":1,"amount":1,"description":1}):
        key = f"{doc.get('date','')}|{doc.get('category','')}|{doc.get('amount',0)}|{doc.get('description','')}"
        existing_expense_keys.add(key)
    for i, row in enumerate(rows):
        if i % 200 == 0: await asyncio.sleep(0)
        rn = i + 2
        try:
            date   = safe(row.get("date", "")).strip()
            cat    = safe(row.get("category", "")).strip()
            amount = safe_float(row.get("amount", 0))
            if not date:   skipped.append({"row": rn, "reason": "Missing date"}); continue
            if not amount: skipped.append({"row": rn, "reason": "Missing amount"}); continue
            if not cat:    skipped.append({"row": rn, "reason": "Missing category"}); continue
            # Dedup check
            exp_key = f"{date}|{cat}|{amount}|{safe(row.get('description',''))}"
            if exp_key in existing_expense_keys:
                skipped.append({"row": rn, "reason": "Expense already imported (same date, category, amount, description)"}); continue
            existing_expense_keys.add(exp_key)
            # Auto-match category if close
            if cat not in valid_cats:
                match = next((c for c in valid_cats if c.lower().startswith(cat.lower()[:4])), "Miscellaneous")
                cat = match
            to_insert.append({
                "date":         date,
                "category":     cat,
                "sub_category": safe(row.get("sub_category", "")),
                "amount":       amount,
                "description":  safe(row.get("description", "")),
                "vendor":       safe(row.get("vendor", "")),
                "payment_mode": safe(row.get("payment_mode", "Cash")) or "Cash",
                "receipt_no":   safe(row.get("receipt_no", "")),
                "notes":        safe(row.get("notes", "")),
                "created_by":   current_user.get("name", "imported"),
                "created_at":   utcnow().isoformat(),
            })
        except Exception as e:
            errors.append({"row": rn, "error": str(e)})
    if to_insert:
        try:
            r = await db.expenses.insert_many(to_insert, ordered=False)
            inserted += len(r.inserted_ids)
        except Exception as e:
            errors.append({"row": "bulk", "error": str(e)})
    return result_summary(inserted, skipped, errors)

@import_router.delete("/clear/{entity}")
async def clear_entity(entity: str, current_user=Depends(require_admin)):
    entity_map = {
        "customers":  ("customers","",  ""),
        "vehicles":   ("vehicles","",   ""),
        "sales":      ("sales","invoice","invoice_number"),
        "service":    ("service_jobs","job","job_number"),
        "parts":      ("spare_parts","",""),
        "parts_sales":("parts_sales","part_bill","bill_number"),
    }
    if entity not in entity_map:
        raise HTTPException(status_code=400, detail=f"Cannot clear '{entity}'")
    coll, counter_name, counter_field = entity_map[entity]
    result = await db[coll].delete_many({})
    if counter_name and counter_field:
        await _sync_counter(counter_name, coll, counter_field)
    return {"entity":entity,"deleted":result.deleted_count}

@import_router.get("/counts")
async def import_counts(current_user=Depends(verify_token)):
    counts = await asyncio.gather(
        db.customers.count_documents({}), db.vehicles.count_documents({}),
        db.sales.count_documents({}), db.service_jobs.count_documents({}),
        db.spare_parts.count_documents({}), db.parts_sales.count_documents({}),
        db.users.count_documents({}),
    )
    return {"customers":counts[0],"vehicles":counts[1],"sales":counts[2],"service_jobs":counts[3],"spare_parts":counts[4],"parts_sales":counts[5],"users":counts[6]}



# ═══════════════════════════════════════════════════════════════════════════════
#  Debt Ledger
# ═══════════════════════════════════════════════════════════════════════════════

class DebtCreate(BaseModel):
    customer_id:   str
    amount:        float = Field(..., ge=0)
    description:   Optional[str] = ""
    due_date:      Optional[str] = ""
    source:        Optional[str] = "manual"   # manual | sale | service

class PaymentCreate(BaseModel):
    amount:     float = Field(..., gt=0)
    notes:      Optional[str] = ""
    paid_date:  Optional[str] = ""

@api_router.get("/debts")
async def list_debts(
    customer_id: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 200,
    current_user=Depends(verify_token)
):
    q: dict = {}
    if customer_id: q["customer_id"] = customer_id
    if status:      q["status"] = status
    cursor = db.debts.find(q).sort("created_at", -1).limit(limit)
    return oids(await cursor.to_list(length=limit))

@api_router.post("/debts", status_code=201)
async def create_debt(body: DebtCreate, current_user=Depends(verify_token)):
    customer = await db.customers.find_one({"_id": obj_id(body.customer_id)})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    doc = {
        "customer_id":    body.customer_id,
        "customer_name":  customer["name"],
        "customer_mobile": customer.get("mobile", ""),
        "amount":         body.amount,
        "paid":           0.0,
        "balance":        body.amount,
        "description":    body.description or "",
        "due_date":       body.due_date or "",
        "source":         body.source or "manual",
        "status":         "pending",
        "payments":       [],
        "created_at":     utcnow().isoformat(),
    }
    result = await db.debts.insert_one(doc)
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)
    return doc

@api_router.get("/debts/summary")
async def debt_summary(current_user=Depends(verify_token)):
    pipeline = [
        {"$group": {
            "_id": "$status",
            "total_amount":  {"$sum": "$amount"},
            "total_balance": {"$sum": "$balance"},
            "count":         {"$sum": 1},
        }}
    ]
    rows = await db.debts.aggregate(pipeline).to_list(length=20)
    return rows

@api_router.post("/debts/{debt_id}/payments", status_code=201)
async def add_payment(debt_id: str, body: PaymentCreate, current_user=Depends(verify_token)):
    debt = await db.debts.find_one({"_id": obj_id(debt_id)})
    if not debt:
        raise HTTPException(status_code=404, detail="Debt not found")
    payment = {
        "amount":    body.amount,
        "notes":     body.notes or "",
        "paid_date": body.paid_date or utcnow().strftime("%Y-%m-%d"),
        "recorded_at": utcnow().isoformat(),
        "recorded_by": current_user.get("name", ""),
    }
    new_paid    = round(debt.get("paid", 0) + body.amount, 2)
    new_balance = round(debt["amount"] - new_paid, 2)
    new_status  = "paid" if new_balance <= 0 else ("partial" if new_paid > 0 else "pending")
    await db.debts.update_one(
        {"_id": obj_id(debt_id)},
        {"$push": {"payments": payment},
         "$set":  {"paid": new_paid, "balance": max(new_balance, 0), "status": new_status}}
    )
    updated = await db.debts.find_one({"_id": obj_id(debt_id)})
    return oid(updated)

@api_router.put("/debts/{debt_id}")
async def update_debt(debt_id: str, body: dict, current_user=Depends(verify_token)):
    allowed = {"description", "due_date", "amount", "status"}
    update  = {k: v for k, v in body.items() if k in allowed}
    if not update:
        raise HTTPException(status_code=400, detail="Nothing to update")
    # recalc balance if amount changed
    debt = await db.debts.find_one({"_id": obj_id(debt_id)})
    if "amount" in update:
        update["balance"] = max(round(update["amount"] - debt.get("paid", 0), 2), 0)
        update["status"]  = "paid" if update["balance"] <= 0 else ("partial" if debt.get("paid",0)>0 else "pending")
    await db.debts.update_one({"_id": obj_id(debt_id)}, {"$set": update})
    updated = await db.debts.find_one({"_id": obj_id(debt_id)})
    return oid(updated)

@api_router.delete("/debts/{debt_id}")
async def delete_debt(debt_id: str, current_user=Depends(require_admin)):
    result = await db.debts.delete_one({"_id": obj_id(debt_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"message": "Deleted"}



# ═══════════════════════════════════════════════════════════════════════════════
#  Expenses
# ═══════════════════════════════════════════════════════════════════════════════


class ExpenseCreate(BaseModel):
    date:         str
    category:     str
    sub_category: Optional[str] = ""
    amount:       float = Field(..., ge=0)
    description:  Optional[str] = ""
    vendor:       Optional[str] = ""
    payment_mode: Optional[str] = "Cash"
    receipt_no:   Optional[str] = ""
    notes:        Optional[str] = ""

class ExpenseUpdate(BaseModel):
    date:         Optional[str]   = None
    category:     Optional[str]   = None
    sub_category: Optional[str]   = None
    amount:       Optional[float] = Field(None, ge=0)
    description:  Optional[str]   = None
    vendor:       Optional[str]   = None
    payment_mode: Optional[str]   = None
    receipt_no:   Optional[str]   = None
    notes:        Optional[str]   = None

@api_router.get("/expenses")
async def list_expenses(
    month:    Optional[str] = Query(None),   # "2026-04"
    category: Optional[str] = Query(None),
    search:   Optional[str] = Query(None),
    limit:    int           = Query(500, ge=1, le=2000),
    current_user=Depends(verify_token),
):
    q: dict = {}
    if month:
        if not re.fullmatch(r"\d{4}-\d{2}", month):
            raise HTTPException(status_code=400, detail="month must be YYYY-MM")
        q["date"] = {"$regex": f"^{re.escape(month)}"}
    if category: q["category"] = category
    if search:
        s = re.escape(search)
        q["$or"] = [
            {"description": {"$regex": s, "$options": "i"}},
            {"vendor":      {"$regex": s, "$options": "i"}},
            {"category":    {"$regex": s, "$options": "i"}},
        ]
    docs = await db.expenses.find(q).sort("date", -1).limit(limit).to_list(limit)
    total = await db.expenses.count_documents(q)
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})

@api_router.post("/expenses", status_code=201)
async def create_expense(body: ExpenseCreate, current_user=Depends(verify_token)):
    doc = {
        "date":         body.date,
        "category":     body.category,
        "sub_category": body.sub_category or "",
        "amount":       body.amount,
        "description":  body.description or "",
        "vendor":       body.vendor or "",
        "payment_mode": body.payment_mode or "Cash",
        "receipt_no":   body.receipt_no or "",
        "notes":        body.notes or "",
        "created_by":   current_user.get("name", ""),
        "created_at":   utcnow().isoformat(),
    }
    result = await db.expenses.insert_one(doc)
    doc["id"] = str(result.inserted_id); doc.pop("_id", None)
    return doc

@api_router.put("/expenses/{expense_id}")
async def update_expense(expense_id: str, body: ExpenseUpdate, current_user=Depends(verify_token)):
    update = {k: v for k, v in body.dict().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="Nothing to update")
    await db.expenses.update_one({"_id": obj_id(expense_id)}, {"$set": update})
    updated = await db.expenses.find_one({"_id": obj_id(expense_id)})
    return oid(updated)

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, current_user=Depends(require_admin)):
    result = await db.expenses.delete_one({"_id": obj_id(expense_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"message": "Deleted"}

@api_router.get("/expenses/stats/summary")
async def expense_stats(
    months: int = Query(10, ge=1, le=24),
    current_user=Depends(verify_token),
):
    """Monthly expense totals by category for the last N months."""
    pipeline = [
        {"$addFields": {"month_key": {"$substr": ["$date", 0, 7]}}},
        {"$group": {
            "_id": {"month": "$month_key", "category": "$category"},
            "total": {"$sum": "$amount"},
        }},
        {"$sort": {"_id.month": -1}},
        {"$limit": months * len(EXPENSE_CATEGORIES)},
    ]
    rows = await db.expenses.aggregate(pipeline).to_list(500)
    # Reshape into {month: {category: total}}
    result: dict = {}
    for r in rows:
        m = r["_id"]["month"]
        c = r["_id"]["category"]
        result.setdefault(m, {})[c] = r["total"]
    return result

@api_router.get("/reports/pnl")
async def profit_and_loss(
    months: int = Query(10, ge=1, le=24),
    current_user=Depends(require_admin),
):
    """Month-by-month P&L: Revenue (sales+service+parts) minus Expenses."""
    # ── Revenue pipelines ──────────────────────────────────────────
    def month_from_sale_date():
        return {"$addFields": {
            "parsed": {"$dateFromString": {"dateString": "$sale_date", "format": "%d %b %Y", "onError": None, "onNull": None}},
        }, "$addFields": {
            "month_key": {"$cond": [
                {"$ne": ["$parsed", None]},
                {"$dateToString": {"format": "%Y-%m", "date": "$parsed"}},
                {"$substr": ["$created_at", 0, 7]},
            ]}
        }}

    sales_pipe = [
        {"$addFields": {
            "parsed": {"$dateFromString": {"dateString": "$sale_date", "format": "%d %b %Y", "onError": None, "onNull": None}},
        }},
        {"$addFields": {
            "month_key": {"$cond": [
                {"$ne": ["$parsed", None]},
                {"$dateToString": {"format": "%Y-%m", "date": "$parsed"}},
                {"$substr": ["$created_at", 0, 7]},
            ]}
        }},
        {"$group": {"_id": "$month_key", "revenue": {"$sum": "$total_amount"}, "count": {"$sum": 1}}},
        {"$sort": {"_id": -1}}, {"$limit": months},
    ]
    svc_pipe = [
        {"$addFields": {"month_key": {"$substr": ["$created_at", 0, 7]}}},
        {"$group": {"_id": "$month_key", "revenue": {"$sum": "$grand_total"}}},
        {"$sort": {"_id": -1}}, {"$limit": months},
    ]
    parts_pipe = [
        {"$addFields": {"month_key": {"$substr": ["$created_at", 0, 7]}}},
        {"$group": {"_id": "$month_key", "revenue": {"$sum": "$grand_total"}}},
        {"$sort": {"_id": -1}}, {"$limit": months},
    ]
    exp_pipe = [
        {"$addFields": {"month_key": {"$substr": ["$date", 0, 7]}}},
        {"$group": {
            "_id": "$month_key",
            "total": {"$sum": "$amount"},
            "by_category": {"$push": {"cat": "$category", "amt": "$amount"}},
        }},
        {"$sort": {"_id": -1}}, {"$limit": months},
    ]

    sales_rev, svc_rev, parts_rev, exp_data = await asyncio.gather(
        db.sales.aggregate(sales_pipe).to_list(months),
        db.service_bills.aggregate(svc_pipe).to_list(months),
        db.parts_sales.aggregate(parts_pipe).to_list(months),
        db.expenses.aggregate(exp_pipe).to_list(months),
    )

    # Build month set
    all_months = sorted(set(
        [r["_id"] for r in sales_rev] +
        [r["_id"] for r in svc_rev] +
        [r["_id"] for r in parts_rev] +
        [r["_id"] for r in exp_data]
    ), reverse=True)[:months]

    def to_map(rows): return {r["_id"]: r["revenue"] for r in rows}
    sm, vm, pm = to_map(sales_rev), to_map(svc_rev), to_map(parts_rev)
    em = {r["_id"]: r["total"] for r in exp_data}
    ec = {r["_id"]: r["by_category"] for r in exp_data}

    result = []
    for m in all_months:
        sales_r = sm.get(m, 0)
        svc_r   = vm.get(m, 0)
        parts_r = pm.get(m, 0)
        revenue = sales_r + svc_r + parts_r
        expenses = em.get(m, 0)
        profit   = revenue - expenses
        margin   = round(profit / revenue * 100, 1) if revenue else 0

        # Category breakdown
        cat_map: dict = {}
        for item in ec.get(m, []):
            cat_map[item["cat"]] = cat_map.get(item["cat"], 0) + item["amt"]

        result.append({
            "month":       m,
            "sales_rev":   round(sales_r, 2),
            "service_rev": round(svc_r, 2),
            "parts_rev":   round(parts_r, 2),
            "revenue":     round(revenue, 2),
            "expenses":    round(expenses, 2),
            "profit":      round(profit, 2),
            "margin":      margin,
            "expense_by_category": cat_map,
        })
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  Full Data Backup Export
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/backup/export")
async def export_backup(current_user=Depends(require_admin)):
    """
    Export all collections as a ZIP containing individual Excel files.
    One file per entity — ready to re-import via the Import page.
    """
    zip_bytes, filename = await _build_backup_zip()
    headers_resp = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "application/zip",
    }
    return StreamingResponse(iter([zip_bytes]), headers=headers_resp, media_type="application/zip")


async def _build_backup_zip() -> tuple[bytes, str]:
    """Build full backup ZIP. Returns (bytes, filename). Used by /backup/export."""
    import zipfile, io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    GOLD = 'FFB8860B'; WHITE = 'FFFFFFFF'

    def make_sheet(wb, title, rows, headers):
        ws = wb.create_sheet(title)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font      = Font(name='Arial', bold=True, color=WHITE, size=9)
            c.fill      = PatternFill('solid', start_color=GOLD)
            c.alignment = Alignment(horizontal='center', vertical='center')
        for ri, row in enumerate(rows, 2):
            for ci, h in enumerate(headers, 1):
                val = row.get(h, '')
                if isinstance(val, (dict, list)):
                    val = str(val)
                ws.cell(row=ri, column=ci, value=val)
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        return ws

    # ── Fetch all collections ──────────────────────────────────────────────────
    (
        customers, vehicles, sales, service_jobs, parts, expenses, debts, staff
    ) = await asyncio.gather(
        db.customers.find({}).to_list(100000),
        db.vehicles.find({}).to_list(100000),
        db.sales.find({}).to_list(100000),
        db.service_jobs.find({}).to_list(100000),
        db.spare_parts.find({}).to_list(100000),
        db.expenses.find({}).to_list(100000),
        db.debts.find({}).to_list(100000),
        db.users.find({}, {"password": 0}).to_list(1000),
    )

    # Convert ObjectIds to strings
    def clean(docs):
        out = []
        for d in docs:
            row = {}
            for k, v in d.items():
                if k == '_id': row['_id'] = str(v)
                elif hasattr(v, '__str__') and 'ObjectId' in type(v).__name__: row[k] = str(v)
                elif isinstance(v, datetime): row[k] = v.isoformat()
                elif isinstance(v, dict): row[k] = str(v)
                elif isinstance(v, list): row[k] = ', '.join(str(x) for x in v)
                else: row[k] = v
            out.append(row)
        return out

    customers_c    = clean(customers)
    vehicles_c     = clean(vehicles)
    sales_c        = clean(sales)
    service_c      = clean(service_jobs)
    parts_c        = clean(parts)
    expenses_c     = clean(expenses)
    debts_c        = clean(debts)
    staff_c        = clean(staff)

    # ── Build separate Excel files ─────────────────────────────────────────────
    def make_workbook(rows, headers, sheet_name):
        wb = Workbook(); wb.remove(wb.active)
        make_sheet(wb, sheet_name, rows, headers)
        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf.read()

    today = utcnow().strftime("%Y-%m-%d")

    files = [
        (f"customers_{today}.xlsx",    make_workbook(customers_c, ['name','mobile','care_of','email','address','tags','created_at'], 'Customers')),
        (f"vehicles_{today}.xlsx",     make_workbook(vehicles_c,  ['brand','model','variant','color','chassis_number','engine_number','vehicle_number','key_number','type','status','inbound_date','inbound_location','return_date','returned_location','customer_name','customer_mobile','created_at'], 'Vehicles')),
        (f"sales_{today}.xlsx",        make_workbook(sales_c,     ['invoice_number','sale_date','customer_name','customer_mobile','care_of','customer_address','vehicle_brand','vehicle_model','vehicle_variant','vehicle_color','vehicle_number','chassis_number','engine_number','sale_price','discount','insurance','rto','total_amount','payment_mode','nominee','status','created_at'], 'Sales')),
        (f"service_{today}.xlsx",      make_workbook(service_c,   ['job_number','check_in_date','customer_name','customer_mobile','vehicle_number','brand','model','odometer_km','complaint','technician','status','grand_total','estimated_delivery','notes','created_at'], 'Service')),
        (f"spare_parts_{today}.xlsx",  make_workbook(parts_c,     ['part_number','name','category','brand','compatible_with','stock','reorder_level','purchase_price','selling_price','gst_rate','hsn_code','location','created_at'], 'Parts')),
        (f"expenses_{today}.xlsx",     make_workbook(expenses_c,  ['date','category','sub_category','amount','description','vendor','payment_mode','receipt_no','notes','created_by','created_at'], 'Expenses')),
        (f"debts_{today}.xlsx",        make_workbook(debts_c,     ['customer_name','customer_mobile','amount','paid','balance','description','due_date','status','source','created_at'], 'Debts')),
        (f"staff_{today}.xlsx",        make_workbook(staff_c,     ['name','username','mobile','email','role','salary','join_date','status','allowed_pages','created_at'], 'Staff')),
    ]

    # ── Pack into ZIP ──────────────────────────────────────────────────────────
    zip_buf = _io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname, data in files:
            zf.writestr(f"MMMotors_Backup_{today}/{fname}", data)

    zip_buf.seek(0)
    return zip_buf.read(), f"MMMotors_Backup_{today}.zip"

# ═══════════════════════════════════════════════════════════════════════════════
#  Accident Estimates
# ═══════════════════════════════════════════════════════════════════════════════

@api_router.get("/accident-estimates")
async def list_accident_estimates(
    skip:  int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    current_user=Depends(verify_token),
):
    cursor = db.accident_estimates.find().sort("created_at", -1).skip(skip).limit(limit)
    docs   = await cursor.to_list(limit)
    total  = await db.accident_estimates.count_documents({})
    return JSONResponse(content=oids(docs), headers={"X-Total-Count": str(total)})


@api_router.post("/accident-estimates", status_code=201)
async def create_accident_estimate(body: AccidentEstimateCreate, current_user=Depends(verify_token)):
    est_number = await next_sequence("accident_estimate")   # already "EST-0001"
    ts         = utcnow().isoformat()
    doc = {
        **body.model_dump(),
        "estimate_number": est_number,
        "created_by":      current_user["username"],
        "created_at":      ts,
        "updated_at":      ts,
    }
    result  = await db.accident_estimates.insert_one(doc)
    created = await db.accident_estimates.find_one({"_id": result.inserted_id})
    return oid(created)


@api_router.get("/accident-estimates/{est_id}")
async def get_accident_estimate(est_id: str, current_user=Depends(verify_token)):
    doc = await db.accident_estimates.find_one({"_id": obj_id(est_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Estimate not found")
    return oid(doc)


@api_router.put("/accident-estimates/{est_id}")
async def update_accident_estimate(est_id: str, body: AccidentEstimateUpdate, current_user=Depends(verify_token)):
    existing = await db.accident_estimates.find_one({"_id": obj_id(est_id)})
    if not existing:
        # Estimate was deleted — create fresh so frontend gets a new id back
        est_number = await next_sequence("accident_estimate")
        ts  = utcnow().isoformat()
        doc = {
            **body.model_dump(exclude_none=True),
            "estimate_number": est_number,
            "created_by":      current_user["username"],
            "created_at":      ts,
            "updated_at":      ts,
        }
        result  = await db.accident_estimates.insert_one(doc)
        created = await db.accident_estimates.find_one({"_id": result.inserted_id})
        return oid(created)
    updates = body.model_dump(exclude_none=True)
    updates["updated_at"] = utcnow().isoformat()
    await db.accident_estimates.update_one({"_id": obj_id(est_id)}, {"$set": updates})
    return oid(await db.accident_estimates.find_one({"_id": obj_id(est_id)}))


@api_router.delete("/accident-estimates/{est_id}", status_code=204)
async def delete_accident_estimate(est_id: str, current_user=Depends(require_admin)):
    result = await db.accident_estimates.delete_one({"_id": obj_id(est_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Estimate not found")


app.include_router(api_router)
app.include_router(import_router)

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 10000))
    uvicorn.run("server:app", host="0.0.0.0", port=port)
